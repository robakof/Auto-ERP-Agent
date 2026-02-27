"""
E-06: Inspekcja struktury arkuszy XLSM — Relacje, Słownik, Przykładowe.
Cel: ustalić indeksy kolumn potrzebne do build_index.py.
"""

import openpyxl
import os
import sys

XLSM_PATH = os.getenv(
    "XLSM_PATH",
    r"C:\Users\dcyprowski\Desktop\Automatyzacja konfiguracji systemu"
    r"\erp_docs\raw\Przetwarzanie bazy XL pod zapytania LLM - testowanie makro.xlsm",
)

SHEETS = ["Relacje", "Słownik wartości kolumn", "Przykładowe wartości kolumn"]


def inspect_sheet(wb, sheet_name: str, data_rows: int = 3):
    try:
        ws = wb[sheet_name]
    except KeyError:
        print(f"  BRAK ARKUSZA: '{sheet_name}'")
        print(f"  Dostępne: {wb.sheetnames}\n")
        return

    print(f"\n=== {sheet_name} ===")
    print(f"  Wymiary: {ws.dimensions}")

    for i, row in enumerate(ws.iter_rows(max_row=1 + data_rows, values_only=True)):
        prefix = "NAGŁÓWEK" if i == 0 else f"     [{i}]"
        cells = [repr(v)[:25] if v is not None else "None" for v in row[:12]]
        print(f"  {prefix}: {cells}")


if __name__ == "__main__":
    print(f"Wczytywanie: {XLSM_PATH}\n")
    try:
        wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=False, data_only=True)
    except FileNotFoundError:
        print(f"BŁĄD: plik nie istnieje: {XLSM_PATH}")
        sys.exit(1)

    print(f"Arkusze w pliku: {wb.sheetnames}\n")
    for sheet in SHEETS:
        inspect_sheet(wb, sheet)

    print("\n[E-06] Zakończono — uzupełnij COL_* w build_index.py na podstawie powyższych indeksów")
