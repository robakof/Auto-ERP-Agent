"""
E-03: Weryfikacja SQLite FTS5 z polskimi znakami i przydatnosci dla schematu ERP
Cel: sprawdzic czy wyszukiwanie po wlasnych nazwach kolumn zwraca uzyteczne wyniki
"""

import sqlite3
import openpyxl
import sys
import os

XLSM_PATH = r"C:\Users\dcyprowski\Desktop\Automatyzacja konfiguracji systemu\erp_docs\raw\Przetwarzanie bazy XL pod zapytania LLM - testowanie makro.xlsm"
DB_PATH = ":memory:"  # Tylko na potrzeby eksperymentu

# Kolumny arkusza Tabele (wiersz 1 = naglowek)
COL_TABELE = {
    "table_id": 0,        # Numer tabeli
    "table_label": 1,     # Nazwa własna tabeli
    "table_name": 2,      # Nazwa tabeli (CDN.XXX)
    "prefix": 3,          # Prefiks
    "description": 5,     # Opis tabeli
}

# Kolumny arkusza Kolumny (wiersz 5 = naglowek, dane od wiersza 6)
COL_KOLUMNY = {
    "table_id": 0,        # Numer tabeli
    "table_label": 1,     # Nazwa własna tabeli
    "table_name": 2,      # Nazwa tabeli
    "col_id": 3,          # Numer kolumny
    "col_label": 4,       # Nazwa własna kolumny
    "col_name": 5,        # Nazwa kolumny (SQL)
    "role": 6,            # Rola kolumny
    "is_useful": 7,       # Czy użyteczna
    "preferred": 8,       # Preferowana do raportów
    "data_type": 9,       # Typ danych
    "description": 11,    # Opis kolumny
}


def load_excel():
    print("[1] Wczytywanie Excel...")
    wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=False, data_only=True)
    return wb


def parse_tables(wb):
    ws = wb["Tabele"]
    tables = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # naglowek
            continue
        if row[COL_TABELE["table_id"]] is None:
            continue
        table_name = row[COL_TABELE["table_name"]]
        if not table_name:
            continue
        tables[table_name] = {
            "table_id": row[COL_TABELE["table_id"]],
            "table_label": row[COL_TABELE["table_label"]] or "",
            "table_name": table_name,
            "description": row[COL_TABELE["description"]] or "",
        }
    print(f"    Wczytano {len(tables)} tabel")
    return tables


def parse_columns(wb):
    ws = wb["Kolumny"]
    columns = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        # Szukaj wiersza naglowkowego
        if not header_found:
            if row[0] == "Numer tabeli":
                header_found = True
            continue
        if row[COL_KOLUMNY["table_id"]] is None:
            continue
        col_name = row[COL_KOLUMNY["col_name"]]
        if not col_name:
            continue
        columns.append({
            "table_name": row[COL_KOLUMNY["table_name"]] or "",
            "table_label": row[COL_KOLUMNY["table_label"]] or "",
            "col_name": col_name,
            "col_label": row[COL_KOLUMNY["col_label"]] or "",
            "is_useful": str(row[COL_KOLUMNY["is_useful"]] or ""),
            "preferred": str(row[COL_KOLUMNY["preferred"]] or ""),
            "data_type": row[COL_KOLUMNY["data_type"]] or "",
            "description": row[COL_KOLUMNY["description"]] or "",
        })
    print(f"    Wczytano {len(columns)} kolumn")
    return columns


def build_fts_index(conn, tables, columns):
    print("[2] Budowanie indeksu FTS5...")
    conn.execute("""
        CREATE VIRTUAL TABLE columns_fts USING fts5(
            table_name,
            table_label,
            col_name,
            col_label,
            description,
            is_useful UNINDEXED,
            preferred UNINDEXED,
            data_type UNINDEXED,
            tokenize = 'unicode61'
        )
    """)
    for col in columns:
        table_info = tables.get(col["table_name"], {})
        table_label = col["table_label"] or table_info.get("table_label", "")
        table_desc = table_info.get("description", "")
        # Lacz opis tabeli i kolumny dla bogatszego indeksu
        full_desc = " ".join(filter(None, [col["description"], table_desc]))
        conn.execute("""
            INSERT INTO columns_fts VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            col["table_name"], table_label, col["col_name"],
            col["col_label"], full_desc,
            col["is_useful"], col["preferred"], col["data_type"]
        ))
    conn.commit()
    print(f"    Zaindeksowano {len(columns)} rekordow")


def search(conn, query, limit=5):
    rows = conn.execute("""
        SELECT table_name, table_label, col_name, col_label,
               description, is_useful, preferred,
               rank
        FROM columns_fts
        WHERE columns_fts MATCH ?
        ORDER BY rank
        LIMIT ?
    """, (query, limit)).fetchall()
    return rows


def print_results(results, query):
    if not results:
        print(f"    BRAK WYNIKOW dla: '{query}'")
        return
    for r in results:
        useful = "[OK]" if r[5] not in ("", "None", None) else "[ ]"
        label = r[3] if r[3] and r[3] != "None" else r[2]
        table = r[1] if r[1] and r[1] not in ("", " ", "None") else r[0]
        desc = (r[4] or "")[:60]
        print(f"    {useful} {table} | {label} ({r[2]}) | {desc}")


def test_polish_characters(conn):
    print("\n[3] Test polskich znakow i deklinacji:")
    cases = [
        ("kontrahent", "szuka 'kontrahent'"),
        ("kontrahenta", "deklinacja dopelniacz"),
        ("Kontrahent", "wielka litera"),
        ("KONTRAHENT", "wszystkie wielkie"),
    ]
    for query, desc in cases:
        results = search(conn, query, limit=2)
        status = "OK" if results else "BRAK"
        print(f"    [{status}] '{query}' ({desc}) -> {len(results)} wynikow")


def test_business_queries(conn):
    print("\n[4] Testy zapytan biznesowych (kryterium: top 5 zawiera odpowiednia kolumne):")
    queries = [
        ("kontrahent zamowienie", "Kolumna laczaca zamowienie z kontrahentem"),
        ("status dokumentu", "Status/typ dokumentu w jakiejs tabeli"),
        ("magazyn towar", "Kolumna towaru na magazynie"),
        ("data wystawienia", "Data wystawienia dokumentu"),
        ("nazwa kontrahenta", "Nazwa kontrahenta"),
    ]
    for query, expected in queries:
        results = search(conn, query, limit=5)
        print(f"\n  Zapytanie: '{query}' (oczekiwane: {expected})")
        print_results(results, query)


def test_useful_flag_filtering(conn):
    print("\n[5] Test filtrowania po fladze uzytecznosci:")
    # Ile kolumn jest oznaczonych jako uzyteczne?
    total = conn.execute("SELECT COUNT(*) FROM columns_fts").fetchone()[0]
    useful = conn.execute(
        "SELECT COUNT(*) FROM columns_fts WHERE is_useful NOT IN ('', 'None', 'Relacja')"
    ).fetchone()[0]
    preferred = conn.execute(
        "SELECT COUNT(*) FROM columns_fts WHERE preferred NOT IN ('', 'None', 'Nie')"
    ).fetchone()[0]
    print(f"    Wszystkich kolumn:       {total}")
    print(f"    Oznaczonych uzytecznych: {useful} ({round(100*useful/total)}%)")
    print(f"    Preferowanych:           {preferred} ({round(100*preferred/total)}%)")

    # Porownaj wyniki z filtem i bez
    query = "kontrahent"
    all_results = search(conn, query, limit=5)
    useful_results = conn.execute("""
        SELECT table_name, table_label, col_name, col_label, description, is_useful, preferred, rank
        FROM columns_fts
        WHERE columns_fts MATCH ?
          AND is_useful NOT IN ('', 'None', 'Relacja')
        ORDER BY rank LIMIT 5
    """, (query,)).fetchall()
    print(f"\n    Zapytanie '{query}' bez filtra: {len(all_results)} wynikow")
    print_results(all_results, query)
    print(f"\n    Zapytanie '{query}' tylko uzyteczne: {len(useful_results)} wynikow")
    print_results(useful_results, query)


if __name__ == "__main__":
    try:
        wb = load_excel()
        tables = parse_tables(wb)
        columns = parse_columns(wb)

        conn = sqlite3.connect(DB_PATH)
        build_fts_index(conn, tables, columns)

        test_polish_characters(conn)
        test_business_queries(conn)
        test_useful_flag_filtering(conn)

        conn.close()
        print("\n[E-03] Zakonczone — oceń wyniki powyzej")

    except Exception as e:
        import traceback
        print(f"\n[E-03] BLAD: {e}")
        traceback.print_exc()
        sys.exit(1)
