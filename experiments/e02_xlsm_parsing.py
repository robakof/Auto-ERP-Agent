"""
E-02: Weryfikacja parsowania pliku xlsm przez openpyxl
Cel: sprawdzić czy data_only=True zwraca wartości zamiast formuł XLOOKUP
"""

import openpyxl
import sys

XLSM_PATH = r"C:\Users\dcyprowski\Desktop\Automatyzacja konfiguracji systemu\erp_docs\raw\Przetwarzanie bazy XL pod zapytania LLM - testowanie makro.xlsm"

SHEETS_TO_CHECK = ["Tabele", "Kolumny", "Relacje", "Słownik wartości kolumn", "Przykładowe wartości kolumn", "Opcje kolumn"]


def classify_cell(value):
    if value is None:
        return "empty"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "value"


def analyze_sheet(wb, sheet_name):
    try:
        ws = wb[sheet_name]
    except KeyError:
        # próba z inną nazwą (encoding)
        matches = [s for s in wb.sheetnames if sheet_name[:6] in s]
        if not matches:
            print(f"  BRAK ARKUSZA: {sheet_name}")
            return
        ws = wb[matches[0]]
        sheet_name = matches[0]

    stats = {"value": 0, "formula": 0, "empty": 0}
    sample_formulas = []
    sample_values = []

    for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
        for cell in row:
            kind = classify_cell(cell)
            stats[kind] += 1
            if kind == "formula" and len(sample_formulas) < 2:
                sample_formulas.append(str(cell)[:60])
            if kind == "value" and len(sample_values) < 3:
                sample_values.append(repr(cell)[:60])

    total = sum(stats.values())
    pct_formula = round(100 * stats["formula"] / max(total, 1), 1)
    pct_value = round(100 * stats["value"] / max(total, 1), 1)

    status = "OK" if pct_formula == 0 else ("UWAGA" if pct_formula < 50 else "PROBLEM")
    print(f"  [{status}] {sheet_name}: {pct_value}% wartości, {pct_formula}% formuł (z 50 pierwszych wierszy)")
    if sample_values:
        print(f"         Przykład wartości: {sample_values}")
    if sample_formulas:
        print(f"         Przykład formuły:  {sample_formulas}")


def test_header_row(wb):
    """Sprawdź czy nagłówki w kluczowych arkuszach są czytelne."""
    print("\n[3] Nagłówki kluczowych arkuszy:")
    for name in ["Tabele", "Kolumny"]:
        matches = [s for s in wb.sheetnames if name in s]
        if not matches:
            continue
        ws = wb[matches[0]]
        for row in ws.iter_rows(max_row=10, values_only=True):
            if any(v is not None and not str(v).startswith("=") for v in row):
                print(f"  {matches[0]}: {[str(v)[:20] if v else None for v in row[:8]]}")
                break


if __name__ == "__main__":
    print("[1] Wczytywanie xlsm z data_only=True...")
    try:
        wb = openpyxl.load_workbook(XLSM_PATH, keep_vba=False, data_only=True)
        print(f"    OK — arkusze: {wb.sheetnames}\n")
    except Exception as e:
        print(f"    BŁĄD: {e}")
        sys.exit(1)

    print("[2] Analiza zawartości arkuszy (wartości vs formuły):")
    for sheet in SHEETS_TO_CHECK:
        analyze_sheet(wb, sheet)

    test_header_row(wb)

    print("\n[E-02] Zakończono — oceń wyniki powyżej")
