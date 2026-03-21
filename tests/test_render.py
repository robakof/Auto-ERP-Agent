"""Tests for tools/render.py — universal renderer for mrowisko.db views."""

import json
import subprocess
import sys
from pathlib import Path

import pytest
import openpyxl

from tools.lib.agent_bus import AgentBus

CLI = str(Path(__file__).parent.parent / "tools" / "render.py")
PYTHON = sys.executable


def run(args: list[str], db_path: str) -> subprocess.CompletedProcess:
    return subprocess.run(
        [PYTHON, CLI, "--db", db_path] + args,
        capture_output=True, text=True, encoding="utf-8",
    )


@pytest.fixture
def db(tmp_path):
    path = str(tmp_path / "test.db")
    bus = AgentBus(db_path=path)
    bus.add_backlog_item("Fix bot", "opis", area="Bot", value="wysoka", effort="mala")
    bus.add_backlog_item("Fix arch", "opis arch", area="Arch", value="srednia", effort="srednia")
    bus.add_suggestion("erp_specialist", "sugestia testowa")
    bus.send_message("erp_specialist", "developer", "wiadomosc testowa")
    bus.add_session_log("developer", "log testowy")
    return path, tmp_path


class TestBacklogMd:
    def test_outputs_md_table(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["backlog", "--format", "md", "--output", str(out)], path)
        assert r.returncode == 0
        content = out.read_text(encoding="utf-8")
        assert "Fix bot" in content
        assert "Fix arch" in content
        assert "##" in content
        assert "**status:**" in content

    def test_filter_area(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["backlog", "--format", "md", "--area", "Bot", "--output", str(out)], path)
        assert r.returncode == 0
        content = out.read_text(encoding="utf-8")
        assert "Fix bot" in content
        assert "Fix arch" not in content

    def test_filter_status(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["backlog", "--format", "md", "--status", "planned", "--output", str(out)], path)
        assert r.returncode == 0
        assert "Fix bot" in out.read_text(encoding="utf-8")

    def test_default_status_excludes_done(self, db):
        path, tmp = db
        bus = AgentBus(db_path=path)
        item_id = bus.add_backlog_item("Done item", "gotowe", area="Dev", value="niska", effort="mala")
        bus.update_backlog_status(item_id, "done")
        out = tmp / "out_default.md"
        r = run(["backlog", "--format", "md", "--output", str(out)], path)
        assert r.returncode == 0
        assert "Done item" not in out.read_text(encoding="utf-8")

    def test_status_all_includes_done(self, db):
        path, tmp = db
        bus = AgentBus(db_path=path)
        item_id = bus.add_backlog_item("Done item2", "gotowe", area="Dev", value="niska", effort="mala")
        bus.update_backlog_status(item_id, "done")
        out = tmp / "out_all.md"
        r = run(["backlog", "--format", "md", "--status", "all", "--output", str(out)], path)
        assert r.returncode == 0
        assert "Done item2" in out.read_text(encoding="utf-8")


class TestBacklogXlsx:
    def test_creates_xlsx(self, db):
        path, tmp = db
        out = tmp / "out.xlsx"
        r = run(["backlog", "--format", "xlsx", "--output", str(out)], path)
        assert r.returncode == 0
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "Fix bot" in values
        assert "Fix arch" in values

    def test_xlsx_has_header(self, db):
        path, tmp = db
        out = tmp / "out.xlsx"
        run(["backlog", "--format", "xlsx", "--output", str(out)], path)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert any(h and "title" in h.lower() for h in headers)


class TestBacklogJson:
    def test_outputs_json(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["backlog", "--format", "json", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert "data" in data
        assert data["count"] == 2
        titles = [i["title"] for i in data["data"]]
        assert "Fix bot" in titles


class TestSuggestions:
    def test_md(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["suggestions", "--format", "md", "--output", str(out)], path)
        assert r.returncode == 0
        assert "sugestia testowa" in out.read_text(encoding="utf-8")

    def test_json(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["suggestions", "--format", "json", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["count"] == 1

    def test_filter_author(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["suggestions", "--format", "json", "--author", "analyst", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["count"] == 0


class TestInbox:
    def test_md(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["inbox", "--role", "developer", "--format", "md", "--output", str(out)], path)
        assert r.returncode == 0
        assert "wiadomosc testowa" in out.read_text(encoding="utf-8")

    def test_json(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["inbox", "--role", "developer", "--format", "json", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["count"] == 1


class TestSessionLog:
    def test_json(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["session-log", "--role", "developer", "--format", "json", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["count"] == 1

    def test_md(self, db):
        path, tmp = db
        out = tmp / "out.md"
        r = run(["session-log", "--role", "developer", "--format", "md", "--output", str(out)], path)
        assert r.returncode == 0
        assert "log testowy" in out.read_text(encoding="utf-8")


class TestMessages:
    def test_all_messages_json(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["messages", "--format", "json", "--output", str(out)], path)
        assert r.returncode == 0
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["count"] == 1
        assert data["data"][0]["sender"] == "erp_specialist"

    def test_filter_sender(self, db):
        path, tmp = db
        out = tmp / "out.json"
        r = run(["messages", "--format", "json", "--sender", "analyst", "--output", str(out)], path)
        assert r.returncode == 0
        assert json.loads(out.read_text(encoding="utf-8"))["count"] == 0

    def test_xlsx(self, db):
        path, tmp = db
        out = tmp / "out.xlsx"
        r = run(["messages", "--format", "xlsx", "--output", str(out)], path)
        assert r.returncode == 0
        assert out.exists()


class TestDefaultOutput:
    def test_default_filename(self, db, tmp_path):
        path, _ = db
        import os
        os.chdir(tmp_path)
        r = run(["backlog", "--format", "json"], path)
        assert r.returncode == 0
        assert (tmp_path / "views" / "backlog.json").exists()
