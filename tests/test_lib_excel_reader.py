"""Testy jednostkowe dla tools/lib/excel_reader.py."""

import pytest
from openpyxl import Workbook

from tools.lib.excel_reader import ExcelReader


# ── Helpers ──────────────────────────────────────────────────────────────────


def make_xlsx(tmp_path, sheet_name: str, headers: list, rows: list, filename="test.xlsx") -> object:
    """Tworzy plik .xlsx z jednym arkuszem i zwraca jego ścieżkę."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    return path


# ── TestExcelReaderInit ───────────────────────────────────────────────────────


class TestExcelReaderInit:
    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ExcelReader(tmp_path / "missing.xlsx")

    def test_context_manager(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["A"], [[1]])
        with ExcelReader(path) as reader:
            assert reader is not None


# ── TestReadStats ─────────────────────────────────────────────────────────────


class TestReadStats:
    def test_happy_path(self, tmp_path):
        path = make_xlsx(tmp_path, "Dane", ["Kod", "Ilosc"], [["A", 1], ["B", 2], ["A", 3]])
        with ExcelReader(path) as r:
            result = r.read_stats()
        assert result["ok"] is True
        assert result["data"]["row_count"] == 3
        assert len(result["data"]["columns"]) == 2

    def test_distinct_count(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["X"], [["a"], ["b"], ["a"]])
        with ExcelReader(path) as r:
            result = r.read_stats()
        col = result["data"]["columns"][0]
        assert col["distinct"] == 2
        assert col["null_count"] == 0

    def test_null_count(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["X"], [["a"], [None], ["b"]])
        with ExcelReader(path) as r:
            result = r.read_stats()
        col = result["data"]["columns"][0]
        assert col["null_count"] == 1

    def test_values_when_below_max_unique(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["X"], [["a"], ["b"]])
        with ExcelReader(path) as r:
            result = r.read_stats(max_unique=5)
        col = result["data"]["columns"][0]
        assert col["values"] is not None
        assert col["sample"] is None

    def test_sample_when_above_max_unique(self, tmp_path):
        rows = [[str(i)] for i in range(10)]
        path = make_xlsx(tmp_path, "S", ["X"], rows)
        with ExcelReader(path) as r:
            result = r.read_stats(max_unique=3)
        col = result["data"]["columns"][0]
        assert col["values"] is None
        assert col["sample"] is not None
        assert len(col["sample"]) <= 10

    def test_filter_columns(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["A", "B", "C"], [[1, 2, 3]])
        with ExcelReader(path) as r:
            result = r.read_stats(columns=["A", "C"])
        names = [c["name"] for c in result["data"]["columns"]]
        assert names == ["A", "C"]

    def test_sheet_by_name(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1.append(["X"])
        ws2 = wb.create_sheet("Second")
        ws2.append(["Y"])
        ws2.append(["val"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)
        with ExcelReader(path) as r:
            result = r.read_stats(sheet_name="Second")
        assert result["ok"] is True
        assert result["data"]["columns"][0]["name"] == "Y"

    def test_sheet_not_found(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["A"], [[1]])
        with ExcelReader(path) as r:
            result = r.read_stats(sheet_name="Missing")
        assert result["ok"] is False
        assert result["error"]["type"] == "SHEET_NOT_FOUND"

    def test_empty_sheet(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        with ExcelReader(path) as r:
            result = r.read_stats()
        assert result["ok"] is True
        assert result["data"]["row_count"] == 0


# ── TestReadRows ──────────────────────────────────────────────────────────────


class TestReadRows:
    def test_happy_path(self, tmp_path):
        path = make_xlsx(tmp_path, "Plan", ["CDN_Pole", "Alias", "Komentarz"],
                         [["Rez_GIDTyp", "Typ", "ok"], ["Rez_GIDNumer", "ID", ""]])
        with ExcelReader(path) as r:
            result = r.read_rows()
        assert result["ok"] is True
        assert result["data"]["row_count"] == 2
        assert result["data"]["columns"] == ["CDN_Pole", "Alias", "Komentarz"]
        assert result["data"]["rows"][0] == ["Rez_GIDTyp", "Typ", "ok"]

    def test_filter_columns(self, tmp_path):
        path = make_xlsx(tmp_path, "Plan", ["A", "B", "C"],
                         [["a1", "b1", "c1"], ["a2", "b2", "c2"]])
        with ExcelReader(path) as r:
            result = r.read_rows(columns=["A", "C"])
        assert result["data"]["columns"] == ["A", "C"]
        assert result["data"]["rows"][0] == ["a1", "c1"]

    def test_sheet_not_found(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["A"], [[1]])
        with ExcelReader(path) as r:
            result = r.read_rows(sheet_name="Missing")
        assert result["ok"] is False
        assert result["error"]["type"] == "SHEET_NOT_FOUND"

    def test_empty_sheet(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        with ExcelReader(path) as r:
            result = r.read_rows()
        assert result["ok"] is True
        assert result["data"]["row_count"] == 0
        assert result["data"]["rows"] == []

    def test_none_values_preserved(self, tmp_path):
        path = make_xlsx(tmp_path, "S", ["A", "B"], [["val", None]])
        with ExcelReader(path) as r:
            result = r.read_rows()
        assert result["data"]["rows"][0] == ["val", None]
