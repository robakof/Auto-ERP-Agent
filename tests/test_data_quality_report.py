"""Testy dla tools/data_quality_report.py."""

import json
import sqlite3

import openpyxl
import pytest

import tools.data_quality_report as dqr


def _make_db(tmp_path, findings=None, records=None):
    """Tworzy SQLite z tabelami findings i records i zwraca ścieżkę."""
    db_path = tmp_path / "workdb.db"
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE findings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            column TEXT,
            observation TEXT,
            rows_affected INTEGER,
            created_at TEXT
        );
        CREATE TABLE records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            column TEXT,
            data TEXT,
            created_at TEXT
        );
    """)
    if findings:
        conn.executemany(
            "INSERT INTO findings (column, observation, rows_affected, created_at) VALUES (?, ?, ?, ?)",
            findings,
        )
    if records:
        conn.executemany(
            "INSERT INTO records (column, data, created_at) VALUES (?, ?, ?)",
            records,
        )
    conn.commit()
    conn.close()
    return db_path


class TestGenerateReport:
    def test_happy_path_creates_xlsx(self, tmp_path):
        db_path = _make_db(tmp_path,
                           findings=[("Telefon", "Email w telefonie.", 47, "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        result = dqr.generate_report(db_path, out_path)
        assert result["ok"] is True
        assert out_path.exists()

    def test_obserwacje_sheet_has_findings(self, tmp_path):
        db_path = _make_db(tmp_path, findings=[
            ("Telefon", "Email w telefonie.", 47, "2026-03-10"),
            ("NIP", "Brak NIP.", 12, "2026-03-10"),
        ])
        out_path = tmp_path / "report.xlsx"
        dqr.generate_report(db_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        assert "Obserwacje" in wb.sheetnames
        ws = wb["Obserwacje"]
        assert ws.max_row == 3  # nagłówek + 2 wiersze

    def test_obserwacje_columns(self, tmp_path):
        db_path = _make_db(tmp_path,
                           findings=[("Telefon", "Obs.", 5, "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        dqr.generate_report(db_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Obserwacje"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Kolumna" in headers
        assert "Obserwacja" in headers
        assert "Liczba_rekordow" in headers
        assert "Data_analizy" in headers

    def test_rekordy_sheet_exists(self, tmp_path):
        db_path = _make_db(tmp_path,
                           findings=[("Telefon", "Obs.", 1, "2026-03-10")],
                           records=[("Telefon", json.dumps({"Kod": "ABC", "Telefon": "x@y.pl"}), "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        dqr.generate_report(db_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        assert "Rekordy" in wb.sheetnames

    def test_rekordy_dynamic_columns_from_json(self, tmp_path):
        db_path = _make_db(tmp_path,
                           records=[
                               ("Telefon", json.dumps({"Kod": "ABC", "Nazwa": "Firma", "Telefon": "x@y.pl"}), "2026-03-10"),
                           ])
        out_path = tmp_path / "report.xlsx"
        dqr.generate_report(db_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Rekordy"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Kod" in headers
        assert "Nazwa" in headers
        assert "Telefon" in headers
        assert "Kolumna" in headers

    def test_rekordy_row_values(self, tmp_path):
        db_path = _make_db(tmp_path,
                           records=[("Telefon", json.dumps({"Kod": "ABC", "Telefon": "x@y.pl"}), "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        dqr.generate_report(db_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Rekordy"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        kod_col = headers.index("Kod") + 1
        assert ws.cell(2, kod_col).value == "ABC"

    def test_empty_findings_still_creates_file(self, tmp_path):
        db_path = _make_db(tmp_path)
        out_path = tmp_path / "report.xlsx"
        result = dqr.generate_report(db_path, out_path)
        assert result["ok"] is True
        assert out_path.exists()

    def test_returns_counts(self, tmp_path):
        db_path = _make_db(tmp_path,
                           findings=[("A", "Obs1.", 1, "2026-03-10"), ("B", "Obs2.", 2, "2026-03-10")],
                           records=[("A", json.dumps({"Kod": "X"}), "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        result = dqr.generate_report(db_path, out_path)
        assert result["data"]["findings_count"] == 2
        assert result["data"]["records_count"] == 1

    def test_db_not_found_returns_error(self, tmp_path):
        result = dqr.generate_report(tmp_path / "brak.db", tmp_path / "report.xlsx")
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"

    def test_creates_parent_directory(self, tmp_path):
        db_path = _make_db(tmp_path)
        out_path = tmp_path / "subdir" / "report.xlsx"
        result = dqr.generate_report(db_path, out_path)
        assert result["ok"] is True
        assert out_path.exists()


class TestMain:
    def test_happy_path_prints_json(self, tmp_path, capsys):
        db_path = _make_db(tmp_path, findings=[("Telefon", "Obs.", 1, "2026-03-10")])
        out_path = tmp_path / "report.xlsx"
        with pytest.MonkeyPatch().context() as mp:
            mp.setattr("sys.argv", ["dqr.py", "--db", str(db_path), "--output", str(out_path)])
            dqr.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is True

    def test_db_not_found_prints_error(self, tmp_path, capsys):
        with pytest.MonkeyPatch().context() as mp:
            mp.setattr("sys.argv", ["dqr.py", "--db", str(tmp_path / "brak.db"), "--output", str(tmp_path / "r.xlsx")])
            dqr.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"
