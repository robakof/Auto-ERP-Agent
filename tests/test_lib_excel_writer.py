"""Testy jednostkowe dla tools/lib/excel_writer.py."""

import pytest
from openpyxl import load_workbook

from tools.lib.excel_writer import ExcelWriter


class TestExcelWriter:
    def test_save_creates_file(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["A", "B"], [[1, "x"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        assert out.exists()

    def test_creates_parent_dir(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["A"], [[1]])
        out = tmp_path / "sub" / "nested" / "out.xlsx"
        w.save(out)
        assert out.exists()

    def test_header_values(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["Kod", "Ilosc"], [["ABC", 10]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Kod"
        assert ws.cell(row=1, column=2).value == "Ilosc"

    def test_data_rows(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["X"], [["a"], ["b"], ["c"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "a"
        assert ws.cell(row=4, column=1).value == "c"

    def test_empty_rows(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["Col1", "Col2"], [])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value is None

    def test_first_sheet_uses_given_name(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Wynik", ["A"], [[1]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        assert "Wynik" in wb.sheetnames

    def test_multiple_sheets(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Plan", ["P"], [["p1"]])
        w.add_sheet("Wynik", ["W"], [["w1"]])
        w.add_sheet("Surówka", ["S"], [["s1"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        assert wb.sheetnames == ["Plan", "Wynik", "Surówka"]

    def test_multiple_sheets_correct_data(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("A", ["col"], [["val_a"]])
        w.add_sheet("B", ["col"], [["val_b"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        assert wb["A"].cell(row=2, column=1).value == "val_a"
        assert wb["B"].cell(row=2, column=1).value == "val_b"

    def test_freeze_panes(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["A"], [[1]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_header_is_bold(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Dane", ["Kolumna"], [[1]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True

    def test_sheet_has_excel_table(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Rezerwacje", ["Kod", "Nazwa"], [["A", "X"], ["B", "Y"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb["Rezerwacje"]
        assert len(ws.tables) > 0

    def test_excel_table_covers_all_data(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Plan", ["A", "B", "C"], [["x", "y", "z"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        ws = wb["Plan"]
        table = list(ws.tables.values())[0]
        assert table.ref == "A1:C2"

    def test_multiple_sheets_each_has_table(self, tmp_path):
        w = ExcelWriter()
        w.add_sheet("Arkusz1", ["X"], [["v1"]])
        w.add_sheet("Arkusz2", ["Y"], [["v2"]])
        out = tmp_path / "out.xlsx"
        w.save(out)
        wb = load_workbook(out)
        assert len(wb["Arkusz1"].tables) > 0
        assert len(wb["Arkusz2"].tables) > 0
