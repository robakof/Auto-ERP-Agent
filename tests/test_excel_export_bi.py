"""Testy dla tools/excel_export_bi.py."""

import json
from pathlib import Path
from unittest.mock import patch

import pyodbc
import pytest
from openpyxl import Workbook

from tests.conftest import make_mock_conn
from tools.lib.sql_client import SqlClient
import tools.excel_export_bi as eb


class TestExportBiView:
    def test_happy_path(self, tmp_path):
        mock_conn, _ = make_mock_conn(["Kod", "Ilosc"], [["A", 1]])
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view("SELECT Kod, Ilosc FROM t", "Test", output_path=output)
        assert result["ok"] is True
        assert "Plan" in result["data"]["sheets"]
        assert "Wynik" in result["data"]["sheets"]
        assert output.exists()

    def test_validation_error(self, tmp_path):
        output = tmp_path / "out.xlsx"
        result = eb.export_bi_view("DELETE FROM t", "Test", output_path=output)
        assert result["ok"] is False
        assert result["error"]["type"] == "VALIDATION_ERROR"

    def test_sql_error(self, tmp_path):
        mock_conn = make_mock_conn([], [])[0]
        mock_conn.cursor.return_value.execute.side_effect = pyodbc.Error(
            "42S02", "[42S02] Invalid object name 'X'"
        )
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view("SELECT * FROM X", "Test", output_path=output)
        assert result["ok"] is False
        assert result["error"]["type"] == "SQL_ERROR"

    def test_plan_sheet_from_xlsx(self, tmp_path):
        # plan jako Excel
        plan_wb = Workbook()
        ws = plan_wb.active
        ws.title = "Plan"
        ws.append(["CDN_Pole", "Alias"])
        ws.append(["Rez_GIDNumer", "ID_Rezerwacji"])
        plan_path = tmp_path / "plan.xlsx"
        plan_wb.save(plan_path)

        mock_conn, _ = make_mock_conn(["X"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view("SELECT X FROM t", "Test", plan_path=plan_path, output_path=output)
        assert result["ok"] is True
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "Plan" in wb.sheetnames
        assert wb["Plan"].cell(row=1, column=1).value == "CDN_Pole"

    def test_plan_sheet_fallback_when_no_plan(self, tmp_path):
        mock_conn, _ = make_mock_conn(["X"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view("SELECT X FROM t", "Test", output_path=output)
        assert result["ok"] is True
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "Plan" in wb.sheetnames

    def test_surowka_sheet_added_when_source_table(self, tmp_path):
        mock_conn, _ = make_mock_conn(["X"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view(
                "SELECT X FROM t", "Test", source_table="CDN.Rezerwacje", output_path=output
            )
        assert result["ok"] is True
        assert "Surówka" in result["data"]["sheets"]

    def test_default_output_in_exports_dir(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            with patch("tools.excel_export_bi.EXPORTS_DIR", tmp_path):
                result = eb.export_bi_view("SELECT n FROM t", "Widok")
        assert result["ok"] is True
        assert "Widok" in result["data"]["path"]

    def test_row_count_in_result(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1], [2], [3]])
        output = tmp_path / "out.xlsx"
        with patch.object(SqlClient, "get_connection", return_value=mock_conn):
            result = eb.export_bi_view("SELECT n FROM t", "Test", output_path=output)
        assert result["data"]["row_count"] == 3


class TestMain:
    def test_valid_call_prints_json(self, tmp_path, capsys):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("sys.argv", [
            "excel_export_bi.py", "--sql", "SELECT n FROM t",
            "--view-name", "Test", "--output", str(output)
        ]):
            with patch.object(SqlClient, "get_connection", return_value=mock_conn):
                eb.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is True

    def test_file_flag_reads_sql_from_file(self, tmp_path, capsys):
        sql_file = tmp_path / "draft.sql"
        sql_file.write_text("SELECT n FROM t", encoding="utf-8")
        output = tmp_path / "out.xlsx"
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        with patch("sys.argv", [
            "excel_export_bi.py", "--file", str(sql_file),
            "--view-name", "Test", "--output", str(output)
        ]):
            with patch.object(SqlClient, "get_connection", return_value=mock_conn):
                eb.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is True

    def test_file_flag_missing_file_returns_error(self, tmp_path, capsys):
        with patch("sys.argv", [
            "excel_export_bi.py", "--file", str(tmp_path / "brak.sql"),
            "--view-name", "Test"
        ]):
            eb.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"

    def test_missing_sql_and_file_returns_error(self, capsys):
        with patch("sys.argv", ["excel_export_bi.py", "--view-name", "Test"]):
            eb.main()
        result = json.loads(capsys.readouterr().out)
        assert result["ok"] is False
        assert result["error"]["type"] == "MISSING_ARGUMENT"
