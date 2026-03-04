"""Testy jednostkowe dla tools/export_bi_view.py. Połączenie z bazą mockowane."""

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pyodbc
import pytest
from openpyxl import load_workbook

import tools.export_bi_view as bv


# ── Helpers ───────────────────────────────────────────────────────────────────


def make_mock_conn(columns: list[str], rows: list[list]) -> MagicMock:
    mock_cursor = MagicMock()
    mock_cursor.description = [(c, None, None, None, None, None, None) for c in columns]
    mock_cursor.fetchall.return_value = [tuple(r) for r in rows]
    mock_conn = MagicMock()
    mock_conn.cursor.return_value = mock_cursor
    return mock_conn


def make_mock_conn_multi(calls: list[tuple[list[str], list[list]]]) -> MagicMock:
    """Mock conn obsługujący kolejne wywołania execute() z różnymi danymi."""
    mock_cursor = MagicMock()
    side_effects = []
    descriptions = []
    fetchalls = []
    for cols, rows in calls:
        descriptions.append([(c, None, None, None, None, None, None) for c in cols])
        fetchalls.append([tuple(r) for r in rows])

    call_count = [0]

    def set_description_and_fetchall(_sql):
        idx = call_count[0]
        mock_cursor.description = descriptions[idx]
        mock_cursor.fetchall.return_value = fetchalls[idx]
        call_count[0] += 1

    mock_cursor.execute.side_effect = set_description_and_fetchall
    mock_conn = MagicMock()
    mock_conn.cursor.return_value = mock_cursor
    return mock_conn


# ── TestValidateQuery ─────────────────────────────────────────────────────────


class TestValidateQuery:
    def test_valid_select(self):
        assert bv.validate_query("SELECT * FROM BI.Rezerwacje") is None

    def test_blocks_delete(self):
        result = bv.validate_query("DELETE FROM t")
        assert result is not None
        assert "DELETE" in result

    def test_blocks_multiple_statements(self):
        assert bv.validate_query("SELECT 1; SELECT 2") is not None

    def test_not_select(self):
        assert bv.validate_query("SHOW TABLES") is not None


# ── TestFillPlanSheet ─────────────────────────────────────────────────────────


class TestFillPlanSheet:
    def test_placeholder_when_no_plan(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        bv._fill_plan_sheet(ws, None)
        assert ws.cell(row=1, column=1).value == "CDN_Pole"

    def test_reads_md_file(self, tmp_path):
        from openpyxl import Workbook
        plan = tmp_path / "plan.md"
        plan.write_text("# Plan\nWiersz 2", encoding="utf-8")
        wb = Workbook()
        ws = wb.active
        bv._fill_plan_sheet(ws, plan)
        assert ws.cell(row=1, column=1).value == "# Plan"
        assert ws.cell(row=2, column=1).value == "Wiersz 2"

    def test_placeholder_when_plan_missing(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        bv._fill_plan_sheet(ws, tmp_path / "brak.md")
        # Brak pliku → szablon
        assert ws.cell(row=1, column=1).value == "CDN_Pole"


# ── TestExportBiView ──────────────────────────────────────────────────────────


class TestExportBiView:
    def test_creates_two_sheets_without_source_table(self, tmp_path):
        mock_conn = make_mock_conn(["ID", "Kod"], [[1, "ABC"]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            result = bv.export_bi_view("SELECT ID, Kod FROM t", "Widok", output_path=output)
        assert result["ok"] is True
        assert result["data"]["sheets"] == ["Plan", "Wynik"]
        wb = load_workbook(output)
        assert set(wb.sheetnames) == {"Plan", "Wynik"}

    def test_creates_three_sheets_with_source_table(self, tmp_path):
        mock_conn = make_mock_conn_multi([
            (["ID"], [[1]]),
            (["RawCol"], [["raw"]]),
        ])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            result = bv.export_bi_view(
                "SELECT ID FROM t", "Widok",
                source_table="CDN.Tabela", output_path=output,
            )
        assert result["ok"] is True
        assert result["data"]["sheets"] == ["Plan", "Wynik", "Surówka"]
        wb = load_workbook(output)
        assert "Surówka" in wb.sheetnames

    def test_wynik_sheet_has_data(self, tmp_path):
        mock_conn = make_mock_conn(["A", "B"], [["x", 1], ["y", 2]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            bv.export_bi_view("SELECT A, B FROM t", "Test", output_path=output)
        wb = load_workbook(output)
        ws = wb["Wynik"]
        assert ws.cell(row=1, column=1).value == "A"
        assert ws.cell(row=2, column=1).value == "x"

    def test_view_name_in_auto_filename(self, tmp_path):
        mock_conn = make_mock_conn(["n"], [[1]])
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            with patch("tools.export_bi_view.EXPORTS_DIR", tmp_path):
                result = bv.export_bi_view("SELECT n FROM t", "Rezerwacje")
        assert result["ok"] is True
        assert "Rezerwacje" in result["data"]["path"]

    def test_injects_top_when_missing(self, tmp_path):
        mock_conn = make_mock_conn(["n"], [[1]])
        mock_cursor = mock_conn.cursor.return_value
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            bv.export_bi_view("SELECT n FROM t", "Test", output_path=output)
        executed = mock_cursor.execute.call_args_list[0][0][0]
        assert "TOP" in executed.upper()

    def test_no_top_injection_when_present(self, tmp_path):
        mock_conn = make_mock_conn(["n"], [[1]])
        mock_cursor = mock_conn.cursor.return_value
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            bv.export_bi_view("SELECT TOP 50 n FROM t", "Test", output_path=output)
        executed = mock_cursor.execute.call_args_list[0][0][0]
        assert executed.upper().count("TOP") == 1

    def test_validation_error(self, tmp_path):
        output = tmp_path / "out.xlsx"
        result = bv.export_bi_view("DELETE FROM t", "Test", output_path=output)
        assert result["ok"] is False
        assert result["error"]["type"] == "VALIDATION_ERROR"
        assert not output.exists()

    def test_sql_error(self, tmp_path):
        mock_conn = MagicMock()
        mock_conn.cursor.return_value.execute.side_effect = pyodbc.Error(
            "42S02", "[42S02] Invalid object name 'Brak'"
        )
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            result = bv.export_bi_view("SELECT * FROM Brak", "Test", output_path=output)
        assert result["ok"] is False
        assert result["error"]["type"] == "SQL_ERROR"

    def test_row_count_in_result(self, tmp_path):
        mock_conn = make_mock_conn(["n"], [[i] for i in range(7)])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            result = bv.export_bi_view("SELECT n FROM t", "Test", output_path=output)
        assert result["data"]["row_count"] == 7

    def test_plan_sheet_from_md(self, tmp_path):
        plan = tmp_path / "plan.md"
        plan.write_text("# Mój plan\nLinia 2", encoding="utf-8")
        mock_conn = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            bv.export_bi_view("SELECT n FROM t", "Test", plan_path=plan, output_path=output)
        wb = load_workbook(output)
        ws = wb["Plan"]
        assert ws.cell(row=1, column=1).value == "# Mój plan"

    def test_duration_ms_present(self, tmp_path):
        mock_conn = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
            result = bv.export_bi_view("SELECT n FROM t", "Test", output_path=output)
        assert isinstance(result["meta"]["duration_ms"], int)
        assert result["meta"]["duration_ms"] >= 0


# ── TestMain ──────────────────────────────────────────────────────────────────


class TestMain:
    def test_view_name_required(self, capsys):
        with patch("sys.argv", ["export_bi_view.py", "--sql", "SELECT 1"]):
            with pytest.raises(SystemExit):
                bv.main()

    def test_sql_required(self, capsys):
        with patch("sys.argv", ["export_bi_view.py", "--view-name", "Test"]):
            with pytest.raises(SystemExit):
                bv.main()

    def test_valid_outputs_json(self, tmp_path, capsys):
        mock_conn = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("sys.argv", [
            "export_bi_view.py",
            "--sql", "SELECT n FROM t",
            "--view-name", "Test",
            "--output", str(output),
        ]):
            with patch("tools.export_bi_view.get_connection", return_value=mock_conn):
                bv.main()
        out = capsys.readouterr().out
        parsed = json.loads(out)
        assert parsed["ok"] is True
        assert parsed["data"]["view_name"] == "Test"
