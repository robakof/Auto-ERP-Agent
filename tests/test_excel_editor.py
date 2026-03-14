"""Tests for ExcelEditor — modifying existing xlsx files."""

import json
from pathlib import Path

import openpyxl
import pytest

from tools.lib.excel_editor import ExcelEditor


@pytest.fixture
def sample_xlsx(tmp_path) -> Path:
    """Excel z arkuszem Dane: kolumny CDN_Pole, Uwzglednic, Komentarz."""
    path = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dane"
    ws.append(["CDN_Pole", "Uwzglednic", "Komentarz"])
    ws.append(["Twr_GIDNumer", "TAK", None])
    ws.append(["Twr_Kod", "TAK", None])
    ws.append(["Twr_Nazwa", "TAK", None])
    wb.save(path)
    return path


class TestWriteCells:
    def test_writes_values_by_key(self, sample_xlsx):
        data = {"Twr_GIDNumer": "klucz techniczny", "Twr_Nazwa": "ok"}
        editor = ExcelEditor(sample_xlsx)
        result = editor.write_cells(
            sheet="Dane",
            key_column="CDN_Pole",
            target_column="Komentarz",
            data=data,
        )
        editor.save(sample_xlsx)
        assert result["ok"] is True
        assert result["data"]["updated"] == 2

        wb = openpyxl.load_workbook(sample_xlsx)
        ws = wb["Dane"]
        rows = {row[0].value: row[2].value for row in ws.iter_rows(min_row=2)}
        assert rows["Twr_GIDNumer"] == "klucz techniczny"
        assert rows["Twr_Nazwa"] == "ok"
        assert rows["Twr_Kod"] is None  # niezmieniony

    def test_missing_key_skipped(self, sample_xlsx):
        data = {"NIEISTNIEJACE_POLE": "wartosc"}
        editor = ExcelEditor(sample_xlsx)
        result = editor.write_cells(
            sheet="Dane",
            key_column="CDN_Pole",
            target_column="Komentarz",
            data=data,
        )
        assert result["ok"] is True
        assert result["data"]["updated"] == 0
        assert result["data"]["skipped"] == 1

    def test_sheet_not_found(self, sample_xlsx):
        editor = ExcelEditor(sample_xlsx)
        result = editor.write_cells(
            sheet="NIEISTNIEJACY",
            key_column="CDN_Pole",
            target_column="Komentarz",
            data={"X": "Y"},
        )
        assert result["ok"] is False
        assert result["error"]["type"] == "SHEET_NOT_FOUND"

    def test_key_column_not_found(self, sample_xlsx):
        editor = ExcelEditor(sample_xlsx)
        result = editor.write_cells(
            sheet="Dane",
            key_column="NIEISTNIEJACA",
            target_column="Komentarz",
            data={"X": "Y"},
        )
        assert result["ok"] is False
        assert result["error"]["type"] == "COLUMN_NOT_FOUND"

    def test_target_column_not_found(self, sample_xlsx):
        editor = ExcelEditor(sample_xlsx)
        result = editor.write_cells(
            sheet="Dane",
            key_column="CDN_Pole",
            target_column="NIEISTNIEJACA",
            data={"X": "Y"},
        )
        assert result["ok"] is False
        assert result["error"]["type"] == "COLUMN_NOT_FOUND"

    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            ExcelEditor(tmp_path / "brak.xlsx")

    def test_context_manager(self, sample_xlsx):
        data = {"Twr_Kod": "sprawdzony"}
        with ExcelEditor(sample_xlsx) as editor:
            editor.write_cells(
                sheet="Dane",
                key_column="CDN_Pole",
                target_column="Komentarz",
                data=data,
            )
            editor.save(sample_xlsx)

        wb = openpyxl.load_workbook(sample_xlsx)
        ws = wb["Dane"]
        rows = {row[0].value: row[2].value for row in ws.iter_rows(min_row=2)}
        assert rows["Twr_Kod"] == "sprawdzony"
