"""Testy dla generate_for_akronimy w xl_attribute_template.py."""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

import tools.xl_attribute_template as xt

_ATTRS = [
    ["WAGA", 3, 0, 0],
    ["KOLOR", 4, 1, 1],
]
_PRODUCTS = [["PROD-01", "Produkt 01"], ["PROD-02", "Produkt 02"]]
_VALS = [
    ["PROD-01", "WAGA", "1.5"],
    ["PROD-01", "KOLOR", "czerwony"],
    ["PROD-01", "KOLOR", "niebieski"],
]


def _mock_client(products=None, vals=None, attrs=None):
    client = MagicMock()
    cursor = MagicMock()
    client.get_connection.return_value.cursor.return_value = cursor

    _attrs = attrs if attrs is not None else _ATTRS
    client.execute.return_value = {"ok": True, "rows": _attrs}

    _products = products if products is not None else _PRODUCTS
    _vals_rows = vals if vals is not None else _VALS

    cursor.fetchall.side_effect = [_products, _vals_rows]
    return client


class TestGenerateForAkronimy:
    def test_empty_akronimy_returns_error(self, tmp_path):
        result = xt.generate_for_akronimy([], tmp_path / "out.xlsx")
        assert result["ok"] is False
        assert result["error"]["type"] == "EMPTY_INPUT"

    def test_creates_excel_file(self, tmp_path):
        out = tmp_path / "out.xlsx"
        with patch("tools.xl_attribute_template.SqlClient", return_value=_mock_client()):
            result = xt.generate_for_akronimy(["PROD-01", "PROD-02"], out)
        assert result["ok"] is True
        assert out.exists()

    def test_prefills_existing_values(self, tmp_path):
        out = tmp_path / "out.xlsx"
        with patch("tools.xl_attribute_template.SqlClient", return_value=_mock_client()):
            xt.generate_for_akronimy(["PROD-01", "PROD-02"], out)
        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb.active
        rows = {(ws.cell(r, 1).value, ws.cell(r, 2).value): ws.cell(r, 4).value
                for r in range(2, ws.max_row + 1)}
        assert rows.get(("PROD-01", "WAGA")) == "1.5"
        wb.close()

    def test_multivalue_in_columns(self, tmp_path):
        out = tmp_path / "out.xlsx"
        with patch("tools.xl_attribute_template.SqlClient", return_value=_mock_client()):
            xt.generate_for_akronimy(["PROD-01", "PROD-02"], out)
        wb = load_workbook(out, read_only=True, data_only=True)
        ws = wb.active
        kolor_row = next(
            r for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value == "PROD-01" and ws.cell(r, 2).value == "KOLOR"
        )
        assert ws.cell(kolor_row, 4).value == "czerwony"
        assert ws.cell(kolor_row, 5).value == "niebieski"
        wb.close()

    def test_not_found_reported(self, tmp_path):
        out = tmp_path / "out.xlsx"
        with patch("tools.xl_attribute_template.SqlClient",
                   return_value=_mock_client(products=[["PROD-01", "Produkt 01"]])):
            result = xt.generate_for_akronimy(["PROD-01", "BRAK-99"], out)
        assert "BRAK-99" in result["data"]["not_found"]

    def test_product_count_in_result(self, tmp_path):
        out = tmp_path / "out.xlsx"
        with patch("tools.xl_attribute_template.SqlClient", return_value=_mock_client()):
            result = xt.generate_for_akronimy(["PROD-01", "PROD-02"], out)
        assert result["data"]["products"] == 2

    def test_sql_error_returns_error(self, tmp_path):
        out = tmp_path / "out.xlsx"
        client = MagicMock()
        client.execute.return_value = {"ok": True, "rows": _ATTRS}
        client.get_connection.return_value.cursor.return_value.fetchall.side_effect = Exception("DB down")
        with patch("tools.xl_attribute_template.SqlClient", return_value=client):
            result = xt.generate_for_akronimy(["PROD-01"], out)
        assert result["ok"] is False
        assert result["error"]["type"] == "SQL_ERROR"
