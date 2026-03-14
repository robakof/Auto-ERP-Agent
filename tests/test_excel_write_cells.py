"""CLI tests for excel_write_cells.py."""

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

PYTHON = sys.executable
CLI = str(Path(__file__).parent.parent / "tools" / "excel_write_cells.py")


def run_cli(args: list[str]) -> dict:
    result = subprocess.run(
        [PYTHON, CLI] + args,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


@pytest.fixture
def sample_xlsx(tmp_path) -> Path:
    path = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dane"
    ws.append(["CDN_Pole", "Komentarz"])
    ws.append(["Twr_GIDNumer", None])
    ws.append(["Twr_Kod", None])
    wb.save(path)
    return path


@pytest.fixture
def data_file(tmp_path) -> Path:
    path = tmp_path / "data.json"
    path.write_text(json.dumps({"Twr_GIDNumer": "klucz techniczny"}), encoding="utf-8")
    return path


class TestExcelWriteCellsCLI:
    def test_writes_cells(self, sample_xlsx, data_file):
        result = run_cli([
            "--file", str(sample_xlsx),
            "--sheet", "Dane",
            "--key-column", "CDN_Pole",
            "--target-column", "Komentarz",
            "--data-file", str(data_file),
        ])
        assert result["ok"] is True
        assert result["data"]["updated"] == 1

        wb = openpyxl.load_workbook(sample_xlsx)
        rows = {r[0].value: r[1].value for r in wb["Dane"].iter_rows(min_row=2)}
        assert rows["Twr_GIDNumer"] == "klucz techniczny"
        assert rows["Twr_Kod"] is None

    def test_missing_data_file(self, sample_xlsx):
        result = run_cli([
            "--file", str(sample_xlsx),
            "--sheet", "Dane",
            "--key-column", "CDN_Pole",
            "--target-column", "Komentarz",
            "--data-file", "/nie/istnieje.json",
        ])
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"

    def test_missing_xlsx(self, tmp_path, data_file):
        result = run_cli([
            "--file", str(tmp_path / "brak.xlsx"),
            "--sheet", "Dane",
            "--key-column", "CDN_Pole",
            "--target-column", "Komentarz",
            "--data-file", str(data_file),
        ])
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"
