"""Testy jednostkowe dla tools/read_excel_stats.py."""

import json
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

import tools.read_excel_stats as rs


# ── Helpers ──────────────────────────────────────────────────────────────────


def make_xlsx(tmp_path, sheet_data: dict[str, list[list]], filename="test.xlsx") -> Path:
    """Tworzy plik xlsx z podanymi arkuszami. sheet_data: {nazwa: [[header...], [row...], ...]}"""
    wb = Workbook()
    first = True
    for sheet_name, rows in sheet_data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    return path


# ── TestReadStats ─────────────────────────────────────────────────────────────


class TestReadStats:
    def test_basic_stats(self, tmp_path):
        path = make_xlsx(tmp_path, {"Dane": [["A", "B"], [1, "x"], [2, "y"], [1, "x"]]})
        result = rs.read_stats(path)
        assert result["ok"] is True
        data = result["data"]
        assert data["row_count"] == 3
        assert len(data["columns"]) == 2

    def test_values_shown_when_lte_max_unique(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["Typ"], ["ZS"], ["ZZ"], ["ZS"]]})
        result = rs.read_stats(path, max_unique=5)
        col = result["data"]["columns"][0]
        assert col["distinct"] == 2
        assert set(col["values"]) == {"ZS", "ZZ"}
        assert col["sample"] is None

    def test_sample_shown_when_gt_max_unique(self, tmp_path):
        rows = [["ID"]] + [[i] for i in range(50)]
        path = make_xlsx(tmp_path, {"S": rows})
        result = rs.read_stats(path, max_unique=5)
        col = result["data"]["columns"][0]
        assert col["distinct"] == 50
        assert col["values"] is None
        assert len(col["sample"]) == 10

    def test_sample_capped_at_10(self, tmp_path):
        rows = [["X"]] + [[i] for i in range(100)]
        path = make_xlsx(tmp_path, {"S": rows})
        result = rs.read_stats(path, max_unique=3)
        col = result["data"]["columns"][0]
        assert len(col["sample"]) == 10

    def test_null_count(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["A"], [1], [None], [None], [2]]})
        result = rs.read_stats(path)
        col = result["data"]["columns"][0]
        assert col["null_count"] == 2
        assert col["total"] == 4

    def test_column_filter(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["A", "B", "C"], [1, 2, 3]]})
        result = rs.read_stats(path, columns=["A", "C"])
        names = [c["name"] for c in result["data"]["columns"]]
        assert names == ["A", "C"]

    def test_column_filter_unknown_skipped(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["A", "B"], [1, 2]]})
        result = rs.read_stats(path, columns=["A", "NIEISTNIEJACA"])
        names = [c["name"] for c in result["data"]["columns"]]
        assert names == ["A"]

    def test_custom_sheet(self, tmp_path):
        path = make_xlsx(tmp_path, {
            "Pierwszy": [["X"], [1]],
            "Drugi": [["Y"], [99]],
        })
        result = rs.read_stats(path, sheet_name="Drugi")
        assert result["data"]["sheet"] == "Drugi"
        assert result["data"]["columns"][0]["name"] == "Y"

    def test_sheet_not_found(self, tmp_path):
        path = make_xlsx(tmp_path, {"A": [["X"], [1]]})
        result = rs.read_stats(path, sheet_name="Nieistniejacy")
        assert result["ok"] is False
        assert result["error"]["type"] == "SHEET_NOT_FOUND"

    def test_file_not_found(self, tmp_path):
        result = rs.read_stats(tmp_path / "brak.xlsx")
        assert result["ok"] is False
        assert result["error"]["type"] == "FILE_NOT_FOUND"

    def test_empty_sheet(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pusty"
        path = tmp_path / "pusty.xlsx"
        wb.save(path)
        result = rs.read_stats(path)
        assert result["ok"] is True
        assert result["data"]["row_count"] == 0
        assert result["data"]["columns"] == []

    def test_distinct_order_preserved(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["V"], ["c"], ["a"], ["b"], ["a"], ["c"]]})
        result = rs.read_stats(path)
        col = result["data"]["columns"][0]
        assert col["values"] == ["c", "a", "b"]

    def test_all_nulls(self, tmp_path):
        path = make_xlsx(tmp_path, {"S": [["A"], [None], [None]]})
        result = rs.read_stats(path)
        col = result["data"]["columns"][0]
        assert col["null_count"] == 2
        assert col["distinct"] == 0
        assert col["values"] == []

    def test_default_sheet_name_returned(self, tmp_path):
        path = make_xlsx(tmp_path, {"MojArkusz": [["X"], [1]]})
        result = rs.read_stats(path)
        assert result["data"]["sheet"] == "MojArkusz"


# ── TestMain ──────────────────────────────────────────────────────────────────


class TestMain:
    def test_outputs_valid_json(self, tmp_path, capsys):
        path = make_xlsx(tmp_path, {"S": [["A", "B"], [1, "x"]]})
        with patch("sys.argv", ["read_excel_stats.py", "--file", str(path)]):
            rs.main()
        out = capsys.readouterr().out
        parsed = json.loads(out)
        assert parsed["ok"] is True

    def test_columns_flag(self, tmp_path, capsys):
        path = make_xlsx(tmp_path, {"S": [["A", "B", "C"], [1, 2, 3]]})
        with patch("sys.argv", [
            "read_excel_stats.py", "--file", str(path), "--columns", "A,C"
        ]):
            rs.main()
        out = capsys.readouterr().out
        parsed = json.loads(out)
        names = [c["name"] for c in parsed["data"]["columns"]]
        assert names == ["A", "C"]

    def test_max_unique_flag(self, tmp_path, capsys):
        rows = [["X"]] + [[i] for i in range(30)]
        path = make_xlsx(tmp_path, {"S": rows})
        with patch("sys.argv", [
            "read_excel_stats.py", "--file", str(path), "--max-unique", "50"
        ]):
            rs.main()
        out = capsys.readouterr().out
        parsed = json.loads(out)
        col = parsed["data"]["columns"][0]
        # 30 distinct <= 50 → values shown
        assert col["values"] is not None
        assert col["sample"] is None
