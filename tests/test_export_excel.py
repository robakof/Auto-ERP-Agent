"""Testy jednostkowe dla tools/export_excel.py. Połączenie z bazą i zapis pliku są mockowane."""

import json
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pyodbc
import pytest
from openpyxl import load_workbook

import tools.export_excel as ee


# ── Helpers ─────────────────────────────────────────────────────────────────


def make_mock_conn(columns: list[str], rows: list[list]) -> tuple[MagicMock, MagicMock]:
    mock_cursor = MagicMock()
    mock_cursor.description = [(col, None, None, None, None, None, None) for col in columns]
    mock_cursor.fetchall.return_value = [tuple(row) for row in rows]
    mock_conn = MagicMock()
    mock_conn.cursor.return_value = mock_cursor
    return mock_conn, mock_cursor


# ── TestValidateQuery ────────────────────────────────────────────────────────


class TestValidateQuery:
    def test_valid_select(self):
        assert ee.validate_query("SELECT * FROM BI.Rezerwacje") is None

    def test_blocks_insert(self):
        result = ee.validate_query("INSERT INTO t VALUES (1)")
        assert result is not None
        assert "INSERT" in result

    def test_blocks_exec(self):
        result = ee.validate_query("EXEC sp_who")
        assert result is not None

    def test_blocks_multiple_statements(self):
        result = ee.validate_query("SELECT 1; SELECT 2")
        assert result is not None

    def test_not_select(self):
        result = ee.validate_query("SHOW TABLES")
        assert result is not None
        assert "NOT_SELECT" in result

    def test_semicolon_at_end_allowed(self):
        assert ee.validate_query("SELECT * FROM t;") is None


# ── TestWriteExcel ───────────────────────────────────────────────────────────


class TestWriteExcel:
    def test_creates_file(self, tmp_path):
        output = tmp_path / "test.xlsx"
        ee._write_excel(["A", "B"], [[1, "x"], [2, "y"]], output)
        assert output.exists()

    def test_header_row(self, tmp_path):
        output = tmp_path / "test.xlsx"
        ee._write_excel(["KodTowaru", "Ilosc"], [["ABC", 10]], output)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "KodTowaru"
        assert ws.cell(row=1, column=2).value == "Ilosc"

    def test_data_rows(self, tmp_path):
        output = tmp_path / "test.xlsx"
        ee._write_excel(["X"], [["a"], ["b"], ["c"]], output)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "a"
        assert ws.cell(row=4, column=1).value == "c"

    def test_empty_result(self, tmp_path):
        output = tmp_path / "test.xlsx"
        ee._write_excel(["Col1", "Col2"], [], output)
        assert output.exists()
        wb = load_workbook(output)
        ws = wb.active
        # Tylko nagłówek — wiersz 2 pusty
        assert ws.cell(row=2, column=1).value is None

    def test_creates_parent_dir(self, tmp_path):
        output = tmp_path / "subdir" / "nested" / "out.xlsx"
        ee._write_excel(["X"], [[1]], output)
        assert output.exists()


# ── TestExportToExcel ────────────────────────────────────────────────────────


class TestExportToExcel:
    def test_happy_path(self, tmp_path):
        mock_conn, _ = make_mock_conn(["KodTowaru", "Ilosc"], [["ABC", 5], ["XYZ", 10]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            result = ee.export_to_excel("SELECT KodTowaru, Ilosc FROM BI.Rezerwacje", output)
        assert result["ok"] is True
        assert result["data"]["row_count"] == 2
        assert result["data"]["columns"] == ["KodTowaru", "Ilosc"]
        assert result["data"]["path"] == str(output.resolve())
        assert output.exists()

    def test_injects_top_1000(self, tmp_path):
        mock_conn, mock_cursor = make_mock_conn(["n"], [[i] for i in range(10)])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            ee.export_to_excel("SELECT n FROM t", output)
        executed_sql = mock_cursor.execute.call_args[0][0]
        assert "TOP 1000" in executed_sql.upper()

    def test_no_top_injection_when_top_present(self, tmp_path):
        mock_conn, mock_cursor = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            ee.export_to_excel("SELECT TOP 50 n FROM t", output)
        executed_sql = mock_cursor.execute.call_args[0][0]
        assert "TOP 50" in executed_sql.upper()
        assert executed_sql.upper().count("TOP") == 1

    def test_validation_error(self, tmp_path):
        output = tmp_path / "out.xlsx"
        result = ee.export_to_excel("DELETE FROM t", output)
        assert result["ok"] is False
        assert result["error"]["type"] == "VALIDATION_ERROR"
        assert not output.exists()

    def test_sql_error(self, tmp_path):
        mock_conn = MagicMock()
        mock_conn.cursor.return_value.execute.side_effect = pyodbc.Error(
            "42S02", "[42S02] Invalid object name 'NonExistent'"
        )
        output = tmp_path / "out.xlsx"
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            result = ee.export_to_excel("SELECT * FROM NonExistent", output)
        assert result["ok"] is False
        assert result["error"]["type"] == "SQL_ERROR"

    def test_default_output_path(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            with patch("tools.export_excel.EXPORTS_DIR", tmp_path):
                result = ee.export_to_excel("SELECT n FROM t")
        assert result["ok"] is True
        assert result["data"]["path"].endswith(".xlsx")

    def test_view_name_in_filename(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            with patch("tools.export_excel.EXPORTS_DIR", tmp_path):
                result = ee.export_to_excel("SELECT n FROM t", view_name="Rezerwacje")
        assert result["ok"] is True
        assert "Rezerwacje" in result["data"]["path"]

    def test_default_prefix_when_no_view_name(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            with patch("tools.export_excel.EXPORTS_DIR", tmp_path):
                result = ee.export_to_excel("SELECT n FROM t")
        filename = Path(result["data"]["path"]).name
        assert filename.startswith("query_")

    def test_duration_ms_present(self, tmp_path):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("tools.export_excel.get_connection", return_value=mock_conn):
            result = ee.export_to_excel("SELECT n FROM t", output)
        assert isinstance(result["meta"]["duration_ms"], int)
        assert result["meta"]["duration_ms"] >= 0


# ── TestMain ─────────────────────────────────────────────────────────────────


class TestMain:
    def test_no_argument_shows_help(self, capsys):
        with patch("sys.argv", ["export_excel.py"]):
            with pytest.raises(SystemExit):
                ee.main()

    def test_valid_query_prints_json(self, tmp_path, capsys):
        mock_conn, _ = make_mock_conn(["n"], [[1]])
        output = tmp_path / "out.xlsx"
        with patch("sys.argv", ["export_excel.py", "SELECT n FROM t", "--output", str(output)]):
            with patch("tools.export_excel.get_connection", return_value=mock_conn):
                ee.main()
        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert result["ok"] is True

    def test_output_flag_sets_path(self, tmp_path, capsys):
        mock_conn, _ = make_mock_conn(["n"], [[42]])
        output = tmp_path / "custom.xlsx"
        with patch("sys.argv", ["export_excel.py", "SELECT n FROM t", "-o", str(output)]):
            with patch("tools.export_excel.get_connection", return_value=mock_conn):
                ee.main()
        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert result["ok"] is True
        assert output.exists()
