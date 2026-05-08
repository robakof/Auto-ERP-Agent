"""Testy dla tools/xl_fko_bulk.py."""

import sys
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

import tools.xl_fko_bulk as xb
from tools.xl_invoice_parser import FzInvoice, FzPozycja

_POZ = FzPozycja(nr=1, nazwa="Usługa A", ilosc=Decimal("1"), jm="szt",
                  cena_netto=Decimal("100.00"), wartosc_netto=Decimal("100.00"),
                  stawka_vat="23", wartosc_vat=Decimal("23.00"))

_INV = FzInvoice(
    nr_obcy="FK/1/2026",
    nip_sprzedawcy="1234567890",
    nazwa_sprzedawcy="Dostawca Sp. z o.o.",
    data_wystawienia=date(2026, 4, 1),
    termin_platnosci=date(2026, 4, 15),
    waluta="PLN",
    suma_netto=Decimal("100.00"),
    suma_vat=Decimal("23.00"),
    suma_brutto=Decimal("123.00"),
    pozycje=(_POZ,),
)

_PARSE_OK   = {"ok": True,  "data": _INV,  "error": None}
_PARSE_FAIL = {"ok": False, "data": None,  "error": "XML parse error: bad"}
_PARSE_KOR  = {"ok": False, "data": None,  "error": "Nieobsługiwany rodzaj faktury: KOR (obsługiwane: VAT)", "skip": True}
_SET_OK     = {"ok": True,  "data": {"doc_id": 1, "nr_obcy": "FK/1/2026", "action": "inserted"},
               "error": None, "meta": {"duration_ms": 10}}
_SET_SKIP   = {"ok": True,  "data": {"nr_obcy": "FK/1/2026", "action": "skipped"},
               "error": None, "meta": {"duration_ms": 5}}
_SET_FAIL   = {"ok": False, "data": None,
               "error": {"type": "CONTRACTOR_NOT_FOUND", "message": "NIP nie istnieje"},
               "meta": {"duration_ms": 5}}


def _xml_dir(tmp_path, n=1) -> Path:
    for i in range(n):
        (tmp_path / f"faktura_{i+1}.xml").write_text("<Faktura/>", encoding="utf-8")
    return tmp_path


class TestBulkImport:
    def test_empty_dir_returns_ok(self, tmp_path):
        result = xb.bulk_import(tmp_path)
        assert result["ok"] is True
        assert result["data"]["total"] == 0

    def test_inserted_counted(self, tmp_path):
        _xml_dir(tmp_path)
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["inserted"] == 1
        assert result["data"]["failed"] == 0
        assert result["data"]["skipped"] == 0

    def test_skipped_counted(self, tmp_path):
        _xml_dir(tmp_path)
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_SKIP):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["skipped"] == 1
        assert result["data"]["inserted"] == 0

    def test_parse_error_counted_as_failed(self, tmp_path):
        _xml_dir(tmp_path)
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_FAIL):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["failed"] == 1

    def test_set_error_counted_as_failed(self, tmp_path):
        _xml_dir(tmp_path)
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_FAIL):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["failed"] == 1

    def test_one_error_does_not_stop_others(self, tmp_path):
        _xml_dir(tmp_path, n=3)
        results_seq = [_PARSE_FAIL, _PARSE_OK, _PARSE_OK]
        with patch.object(xb, "parse_ksef_xml", side_effect=results_seq), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["total"] == 3
        assert result["data"]["failed"] == 1
        assert result["data"]["inserted"] == 2

    def test_only_xml_files_processed(self, tmp_path):
        (tmp_path / "faktura.xml").write_text("<x/>", encoding="utf-8")
        (tmp_path / "notatka.txt").write_text("ignore", encoding="utf-8")
        (tmp_path / "inne.xlsx").write_text("ignore", encoding="utf-8")
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["total"] == 1

    def test_report_created(self, tmp_path):
        _xml_dir(tmp_path)
        report = tmp_path / "raport.xlsx"
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            xb.bulk_import(tmp_path, report=report)
        assert report.exists()

    def test_report_has_header_and_data_rows(self, tmp_path):
        _xml_dir(tmp_path, n=2)
        report = tmp_path / "raport.xlsx"
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            xb.bulk_import(tmp_path, report=report)
        wb = load_workbook(report, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # 1 header + 2 data
        wb.close()

    def test_report_status_ok(self, tmp_path):
        _xml_dir(tmp_path)
        report = tmp_path / "raport.xlsx"
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            xb.bulk_import(tmp_path, report=report)
        wb = load_workbook(report, read_only=True, data_only=True)
        ws = wb.active
        statuses = [ws.cell(r, 5).value for r in range(2, ws.max_row + 1)]
        assert "OK" in statuses
        wb.close()

    def test_report_tab_name_fko(self, tmp_path):
        _xml_dir(tmp_path)
        report = tmp_path / "raport.xlsx"
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_OK), \
             patch.object(xb, "set_invoice", return_value=_SET_OK):
            xb.bulk_import(tmp_path, report=report)
        wb = load_workbook(report, read_only=True)
        assert wb.active.title == "Import FKo"
        wb.close()

    def test_kor_counted_as_skipped_not_failed(self, tmp_path):
        _xml_dir(tmp_path)
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_KOR):
            result = xb.bulk_import(tmp_path)
        assert result["data"]["skipped"] == 1
        assert result["data"]["failed"] == 0

    def test_kor_report_status_pominieto(self, tmp_path):
        _xml_dir(tmp_path)
        report = tmp_path / "raport.xlsx"
        with patch.object(xb, "parse_ksef_xml", return_value=_PARSE_KOR):
            xb.bulk_import(tmp_path, report=report)
        wb = load_workbook(report, read_only=True, data_only=True)
        ws = wb.active
        statuses = [ws.cell(r, 5).value for r in range(2, ws.max_row + 1)]
        assert "POMINIĘTO" in statuses
        assert "BŁĄD" not in statuses
        wb.close()
        wb.close()
