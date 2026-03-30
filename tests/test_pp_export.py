"""Testy pp_export.export_plan — weryfikacja struktury xlsx."""
import datetime
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.lib.pp_export import export_plan
from tools.lib.pp_gap import GapRow
from tools.lib.pp_schedule import ScheduleSlot

_DEMAND = [
    {"Nr_Zamowienia": "ZO/2025/001", "Data_Realizacji": datetime.date(2025, 6, 1),
     "Kontrahent_Kod": "KNT1", "Kontrahent_Nazwa": "Firma A",
     "Towar_Kod": "CZNI001", "Towar_Nazwa": "Znicz A", "Ilosc": 100,
     "Jednostka": "szt", "Opis": "Zamówienie X"},
]

_SUPPLY_ROWS = [
    {"Towar_Kod": "SZ0324", "Towar_Nazwa": "Szkło", "Jednostka": "szt", "Stan": 5000},
]

_SUPPLY = {"SZ0324": 5000.0}

_GAPS = [
    GapRow("SZ0324", "Szkło", 1020.0, 5000.0, 0.0),
    GapRow("WK0001", "Wkład", 1000.0, 200.0, 800.0),
]


def test_export_tworzy_4_arkusze(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out)
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 4
    assert "Zamówienia CZNI" in wb.sheetnames
    assert "Zapotrzebowanie surowców" in wb.sheetnames
    assert "Stany OTOR_SUR" in wb.sheetnames
    assert "Gap Analysis" in wb.sheetnames


def test_export_gap_czerwone_wiersze(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Gap Analysis"]
    # WK0001 brak=800 → powinien być czerwony (sortowanie Brak DESC → wiersz 2)
    red_rows = []
    for row_idx in range(2, len(_GAPS) + 2):
        fill = ws.cell(row=row_idx, column=1).fill
        if fill and fill.fgColor and fill.fgColor.rgb == "FFFF9999":
            red_rows.append(row_idx)
    assert len(red_rows) >= 1, "Powinien być co najmniej 1 czerwony wiersz"


_D = datetime.date

_SCHEDULE = [
    ScheduleSlot(
        date=_D(2026, 5, 4), czni_kod="CZNI001", czni_nazwa="Znicz A",
        qty=30.0, hours_needed=1.0, order_numbers=["ZO/001"],
        deadline=_D(2026, 5, 10), priority=False,
    ),
    ScheduleSlot(
        date=_D(2026, 5, 4), czni_kod="CZNI002", czni_nazwa="Znicz B",
        qty=60.0, hours_needed=2.0, order_numbers=["ZO/002"],
        deadline=_D(2026, 5, 3), priority=True,  # po deadline → czerwony Gantt
    ),
    ScheduleSlot(
        date=_D(2026, 5, 5), czni_kod="CZNI001", czni_nazwa="Znicz A",
        qty=30.0, hours_needed=1.0, order_numbers=["ZO/001"],
        deadline=_D(2026, 5, 10), priority=False,
    ),
]


def test_export_tworzy_7_arkuszy_ze_schedule(tmp_path):
    out = tmp_path / "test7.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities={"ZO/002"},
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 7
    assert "Zamówienia z priorytetem" in wb.sheetnames
    assert "Harmonogram" in wb.sheetnames
    assert "Gantt" in wb.sheetnames


def test_export_arkusz5_kolumna_priorytet(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities=set(),
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    ws = wb["Zamówienia z priorytetem"]
    headers = [ws.cell(1, c).value for c in range(1, 8)]
    assert "Priorytet" in headers


def test_export_arkusz6_harmonogram_dane(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities={"ZO/002"},
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    ws = wb["Harmonogram"]
    # Nagłówki
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert "Data" in headers
    assert "CZNI_Kod" in headers
    # Dane — 3 sloty
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 3


def test_export_arkusz6_zolte_priorytety(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities={"ZO/002"},
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    ws = wb["Harmonogram"]
    yellow_rows = [
        r for r in range(2, 5)
        if ws.cell(r, 1).fill.fgColor.rgb == "FFFFFF00"
    ]
    assert len(yellow_rows) >= 1


def test_export_arkusz7_gantt_wiersze_czni(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities=set(),
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    ws = wb["Gantt"]
    # Wiersz 1 = nagłówki (CZNI_Kod, daty)
    czni_kody = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert "CZNI001" in czni_kody
    assert "CZNI002" in czni_kody


def test_export_arkusz7_gantt_czerwony_po_deadline(tmp_path):
    out = tmp_path / "test.xlsx"
    export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out,
                schedule=_SCHEDULE, priorities=set(),
                start_date=_D(2026, 5, 4))
    wb = openpyxl.load_workbook(out)
    ws = wb["Gantt"]
    red_cells = []
    for row in ws.iter_rows(min_row=2):
        for cell in row[1:]:  # pomiń kolumnę A (CZNI_Kod)
            if cell.fill and cell.fill.fgColor.rgb == "FFFF0000":
                red_cells.append(cell.coordinate)
    assert len(red_cells) >= 1, "CZNI002 jest po deadline → czerwony"


def test_export_permission_error(tmp_path):
    out = tmp_path / "test.xlsx"
    out.write_bytes(b"")
    out.chmod(0o444)
    try:
        export_plan(_DEMAND, _SUPPLY, _GAPS, _SUPPLY_ROWS, out)
    except RuntimeError as e:
        assert "Zamknij plik" in str(e)
    except PermissionError:
        pass  # też akceptowalny wynik na Windows
    finally:
        out.chmod(0o666)
