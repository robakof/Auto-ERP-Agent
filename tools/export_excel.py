"""
export_excel.py — Narzędzie agenta: eksport wyników SQL do pliku Excel (.xlsx).

CLI:
    python tools/export_excel.py "SELECT TOP 100 ..." [--output SCIEZKA.xlsx]

Jeśli --output nie podano, plik zapisywany jest w katalogu exports/ (tworzony
automatycznie) z nazwą opartą na znaczniku czasu.

Output: JSON na stdout zgodny z kontraktem narzędzi agenta.

Guardrails identyczne z sql_query.py (blokada DML, wymuszenie TOP, timeout 30s).
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

DRIVER = "ODBC Driver 17 for SQL Server"
DEFAULT_TOP = 1000          # Excel może pomieścić więcej niż terminal
TIMEOUT_SECONDS = 30

BLOCKED_KEYWORDS = [
    "INSERT", "UPDATE", "DELETE", "DROP", "CREATE", "ALTER",
    "TRUNCATE", "EXEC", "EXECUTE", "SP_", "XP_",
    "OPENROWSET", "OPENQUERY", "SELECT INTO",
]

EXPORTS_DIR = Path(__file__).parent.parent / "exports"


# ── Walidacja ────────────────────────────────────────────────────────────────


def validate_query(sql: str) -> str | None:
    """Zwraca komunikat błędu lub None gdy zapytanie jest bezpieczne do wykonania."""
    upper = sql.upper()
    for keyword in BLOCKED_KEYWORDS:
        if keyword in upper:
            return f"BLOCKED_KEYWORD: '{keyword}' nie jest dozwolone"
    statements = [s.strip() for s in sql.split(";") if s.strip()]
    if len(statements) > 1:
        return "MULTIPLE_STATEMENTS: dozwolone tylko jedno zapytanie"
    if not upper.lstrip().startswith("SELECT"):
        return "NOT_SELECT: dozwolone tylko zapytania SELECT"
    return None


# ── Połączenie ───────────────────────────────────────────────────────────────


def get_connection() -> pyodbc.Connection:
    conn_str = (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={os.getenv('SQL_SERVER')};"
        f"DATABASE={os.getenv('SQL_DATABASE')};"
        f"UID={os.getenv('SQL_USERNAME')};"
        f"PWD={os.getenv('SQL_PASSWORD')};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=TIMEOUT_SECONDS)


# ── Excel ────────────────────────────────────────────────────────────────────


def _write_excel(columns: list[str], rows: list[list], output_path: Path) -> None:
    """Zapisuje dane do .xlsx z nagłówkiem pogrubionym i zamrożoną pierwszą linią."""
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    # Nagłówek
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Dane
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Zamroź nagłówek
    ws.freeze_panes = "A2"

    # Auto-szerokość kolumn (max 50 znaków)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            cell_val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── Główna logika ────────────────────────────────────────────────────────────


def export_to_excel(
    sql: str,
    output_path: Path | None = None,
    view_name: str | None = None,
) -> dict:
    """Wykonuje SQL i zapisuje wynik do .xlsx. Zwraca słownik per kontrakt JSON."""
    error = validate_query(sql)
    if error:
        return {
            "ok": False,
            "data": None,
            "error": {"type": "VALIDATION_ERROR", "message": error},
            "meta": {"duration_ms": 0, "row_count": 0},
        }

    if "TOP " not in sql.upper():
        sql = sql.replace("SELECT ", f"SELECT TOP {DEFAULT_TOP} ", 1)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = view_name.replace(" ", "_").replace("/", "_") if view_name else "query"
        output_path = EXPORTS_DIR / f"{prefix}_{timestamp}.xlsx"

    output_path = Path(output_path)

    start = time.monotonic()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        columns = [col[0] for col in cursor.description]
        rows = [list(row) for row in cursor.fetchall()]
        duration_ms = round((time.monotonic() - start) * 1000)

        _write_excel(columns, rows, output_path)

        return {
            "ok": True,
            "data": {
                "path": str(output_path.resolve()),
                "row_count": len(rows),
                "columns": columns,
            },
            "error": None,
            "meta": {"duration_ms": duration_ms, "row_count": len(rows)},
        }

    except pyodbc.Error as e:
        duration_ms = round((time.monotonic() - start) * 1000)
        message = e.args[1] if len(e.args) > 1 else str(e)
        return {
            "ok": False,
            "data": None,
            "error": {"type": "SQL_ERROR", "message": message},
            "meta": {"duration_ms": duration_ms, "row_count": 0},
        }

    except Exception as e:  # noqa: BLE001
        duration_ms = round((time.monotonic() - start) * 1000)
        return {
            "ok": False,
            "data": None,
            "error": {"type": "EXPORT_ERROR", "message": str(e)},
            "meta": {"duration_ms": duration_ms, "row_count": 0},
        }


# ── CLI ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Eksportuj wyniki SQL do pliku Excel."
    )
    parser.add_argument("sql", help='Zapytanie SELECT, np. "SELECT TOP 100 * FROM BI.Rezerwacje"')
    parser.add_argument(
        "--output", "-o",
        help="Ścieżka do pliku wyjściowego .xlsx (domyślnie: exports/{view_name}_TIMESTAMP.xlsx)",
        default=None,
    )
    parser.add_argument(
        "--view-name",
        help="Nazwa widoku używana w nazwie pliku (domyślnie: 'query')",
        default=None,
    )
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else None
    result = export_to_excel(args.sql, output_path, args.view_name)
    print(json.dumps(result, default=str, ensure_ascii=False))


if __name__ == "__main__":
    main()
