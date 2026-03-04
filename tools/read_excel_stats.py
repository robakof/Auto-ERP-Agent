"""
read_excel_stats.py — Statystyki kolumn z pliku Excel dla agenta.

CLI:
    python tools/read_excel_stats.py --file PATH.xlsx
                                     [--sheet NAZWA]
                                     [--max-unique N]
                                     [--columns col1,col2]

Dla każdej kolumny zwraca: total, distinct, null_count oraz:
  - values  — pełna lista unikalnych wartości (gdy distinct <= max_unique)
  - sample  — pierwsze 10 unikalnych (gdy distinct > max_unique)

Pozwala agentowi zweryfikować wynik eksportu bez ładowania danych do kontekstu.
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def read_stats(
    file_path: Path,
    sheet_name: str | None = None,
    max_unique: int = 20,
    columns: list[str] | None = None,
) -> dict:
    """Zwraca statystyki kolumn z pliku xlsx."""
    if not file_path.exists():
        return {
            "ok": False,
            "data": None,
            "error": {
                "type": "FILE_NOT_FOUND",
                "message": f"Plik nie istnieje: {file_path}",
            },
        }

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {
            "ok": False,
            "data": None,
            "error": {"type": "OPEN_ERROR", "message": str(e)},
        }

    if sheet_name is None:
        ws = wb.active
        sheet_name = ws.title
    else:
        if sheet_name not in wb.sheetnames:
            available = list(wb.sheetnames)
            wb.close()
            return {
                "ok": False,
                "data": None,
                "error": {
                    "type": "SHEET_NOT_FOUND",
                    "message": f"Arkusz '{sheet_name}' nie istnieje. Dostępne: {available}",
                },
            }
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {
            "ok": True,
            "data": {"sheet": sheet_name, "row_count": 0, "columns": []},
            "error": None,
        }

    headers = [
        str(h) if h is not None else f"Col{i + 1}"
        for i, h in enumerate(rows[0])
    ]
    data_rows = rows[1:]
    row_count = len(data_rows)

    if columns:
        col_indices = [headers.index(c) for c in columns if c in headers]
    else:
        col_indices = list(range(len(headers)))

    col_stats = []
    for idx in col_indices:
        col_name = headers[idx]
        values = [row[idx] if idx < len(row) else None for row in data_rows]

        null_count = sum(1 for v in values if v is None)
        non_null = [v for v in values if v is not None]

        # Unikalne wartości w kolejności pierwszego wystąpienia
        seen: dict = {}
        for v in non_null:
            key = str(v)
            if key not in seen:
                seen[key] = True
        distinct_values = list(seen.keys())
        distinct_count = len(distinct_values)

        stat: dict = {
            "name": col_name,
            "total": row_count,
            "distinct": distinct_count,
            "null_count": null_count,
        }

        if distinct_count <= max_unique:
            stat["values"] = distinct_values
            stat["sample"] = None
        else:
            stat["values"] = None
            stat["sample"] = distinct_values[:10]

        col_stats.append(stat)

    return {
        "ok": True,
        "data": {
            "sheet": sheet_name,
            "row_count": row_count,
            "columns": col_stats,
        },
        "error": None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Odczytaj statystyki kolumn z pliku Excel."
    )
    parser.add_argument("--file", required=True, help="Ścieżka do pliku .xlsx")
    parser.add_argument(
        "--sheet", default=None, help="Nazwa arkusza (domyślnie: pierwszy)"
    )
    parser.add_argument(
        "--max-unique",
        type=int,
        default=20,
        help="Próg unikalnych wartości (domyślnie: 20)",
    )
    parser.add_argument(
        "--columns",
        default=None,
        help="Kolumny oddzielone przecinkami (domyślnie: wszystkie)",
    )
    args = parser.parse_args()

    cols = [c.strip() for c in args.columns.split(",")] if args.columns else None
    result = read_stats(Path(args.file), args.sheet, args.max_unique, cols)
    print(json.dumps(result, default=str, ensure_ascii=False))


if __name__ == "__main__":
    main()
