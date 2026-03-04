"""
export_bi_view.py — Narzędzie agenta: eksport widoku BI do wieloarkuszowego pliku Excel.

CLI:
    python tools/export_bi_view.py --sql "SELECT ..." --view-name "NazwaWidoku"
                                   [--source-table CDN.XXX] [--plan PATH.md]
                                   [--output SCIEZKA.xlsx]

Arkusze:
  Plan    — tabela mapowania pól z pliku --plan (jeśli podano) lub pusty szablon
  Wynik   — wynik SQL (do 5000 wierszy)
  Surówka — SELECT TOP 200 * FROM source-table (jeśli podano --source-table)

Nazwa pliku (gdy brak --output): {view_name}_{YYYYMMDD_HHMMSS}.xlsx
"""

import argparse
import json
import os
import time
from datetime import datetime
from pathlib import Path

import pyodbc
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

DRIVER = "ODBC Driver 17 for SQL Server"
DEFAULT_TOP = 5000
RAW_TOP = 200
TIMEOUT_SECONDS = 30

BLOCKED_KEYWORDS = [
    "INSERT", "UPDATE", "DELETE", "DROP", "CREATE", "ALTER",
    "TRUNCATE", "EXEC", "EXECUTE", "SP_", "XP_",
    "OPENROWSET", "OPENQUERY", "SELECT INTO",
]

EXPORTS_DIR = Path(__file__).parent.parent / "exports"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


# ── Walidacja ─────────────────────────────────────────────────────────────────


def validate_query(sql: str) -> str | None:
    upper = sql.upper()
    for keyword in BLOCKED_KEYWORDS:
        if keyword in upper:
            return f"BLOCKED_KEYWORD: '{keyword}' nie jest dozwolone"
    statements = [s.strip() for s in sql.split(";") if s.strip()]
    if len(statements) > 1:
        return "MULTIPLE_STATEMENTS: dozwolone tylko jedno zapytanie"
    if not upper.lstrip().startswith("SELECT"):
        return "NOT_SELECT: dozwolone tylko zapytania SELECT"
    return None


# ── Połączenie ────────────────────────────────────────────────────────────────


def get_connection() -> pyodbc.Connection:
    conn_str = (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={os.getenv('SQL_SERVER')};"
        f"DATABASE={os.getenv('SQL_DATABASE')};"
        f"UID={os.getenv('SQL_USERNAME')};"
        f"PWD={os.getenv('SQL_PASSWORD')};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=TIMEOUT_SECONDS)


# ── Excel helpers ─────────────────────────────────────────────────────────────


def _style_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _auto_width(ws, columns: list[str], rows: list[list]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in rows:
            val = str(row[col_idx - 1]) if col_idx - 1 < len(row) and row[col_idx - 1] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def _fill_data_sheet(ws, columns: list[str], rows: list[list]) -> None:
    _style_header(ws, columns)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _auto_width(ws, columns, rows)


def _fill_plan_sheet(ws, plan_path: Path | None) -> None:
    if plan_path and plan_path.exists():
        content = plan_path.read_text(encoding="utf-8")
        for row_idx, line in enumerate(content.splitlines(), start=1):
            ws.cell(row=row_idx, column=1, value=line)
        ws.column_dimensions["A"].width = 120
    else:
        headers = ["CDN_Pole", "Alias_w_raporcie", "Transformacja", "Uzasadnienie"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.cell(row=2, column=1, value="-- brak pliku --plan; uzupełnij tabelę ręcznie --")
        for width, col in zip([25, 30, 40, 40], ["A", "B", "C", "D"]):
            ws.column_dimensions[col].width = width


# ── Główna logika ─────────────────────────────────────────────────────────────


def export_bi_view(
    sql: str,
    view_name: str,
    source_table: str | None = None,
    plan_path: Path | None = None,
    output_path: Path | None = None,
) -> dict:
    """Eksportuje widok BI do wieloarkuszowego .xlsx. Zwraca słownik per kontrakt JSON."""
    error = validate_query(sql)
    if error:
        return {
            "ok": False,
            "data": None,
            "error": {"type": "VALIDATION_ERROR", "message": error},
            "meta": {"duration_ms": 0},
        }

    if "TOP " not in sql.upper():
        sql = sql.replace("SELECT ", f"SELECT TOP {DEFAULT_TOP} ", 1)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = view_name.replace(" ", "_").replace("/", "_")
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        output_path = EXPORTS_DIR / f"{safe_name}_{timestamp}.xlsx"

    output_path = Path(output_path)

    start = time.monotonic()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(sql)
        result_columns = [col[0] for col in cursor.description]
        result_rows = [list(row) for row in cursor.fetchall()]

        raw_columns: list[str] | None = None
        raw_rows: list[list] | None = None
        if source_table:
            cursor.execute(f"SELECT TOP {RAW_TOP} * FROM {source_table}")
            raw_columns = [col[0] for col in cursor.description]
            raw_rows = [list(row) for row in cursor.fetchall()]

        duration_ms = round((time.monotonic() - start) * 1000)

        wb = Workbook()

        ws_plan = wb.active
        ws_plan.title = "Plan"
        _fill_plan_sheet(ws_plan, plan_path)

        ws_result = wb.create_sheet("Wynik")
        _fill_data_sheet(ws_result, result_columns, result_rows)

        sheets = ["Plan", "Wynik"]
        if raw_columns is not None:
            ws_raw = wb.create_sheet("Surówka")
            _fill_data_sheet(ws_raw, raw_columns, raw_rows)
            sheets.append("Surówka")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return {
            "ok": True,
            "data": {
                "path": str(output_path.resolve()),
                "view_name": view_name,
                "row_count": len(result_rows),
                "columns": result_columns,
                "sheets": sheets,
            },
            "error": None,
            "meta": {"duration_ms": duration_ms},
        }

    except pyodbc.Error as e:
        duration_ms = round((time.monotonic() - start) * 1000)
        message = e.args[1] if len(e.args) > 1 else str(e)
        return {
            "ok": False,
            "data": None,
            "error": {"type": "SQL_ERROR", "message": message},
            "meta": {"duration_ms": duration_ms},
        }

    except Exception as e:  # noqa: BLE001
        duration_ms = round((time.monotonic() - start) * 1000)
        return {
            "ok": False,
            "data": None,
            "error": {"type": "EXPORT_ERROR", "message": str(e)},
            "meta": {"duration_ms": duration_ms},
        }


# ── CLI ───────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Eksportuj widok BI do wieloarkuszowego pliku Excel."
    )
    parser.add_argument("--sql", required=True, help="Zapytanie SELECT")
    parser.add_argument("--view-name", required=True, help="Nazwa widoku (w nazwie pliku)")
    parser.add_argument(
        "--source-table", default=None,
        help="Tabela źródłowa do arkusza Surówka (np. CDN.Rezerwacje)",
    )
    parser.add_argument(
        "--plan", default=None,
        help="Ścieżka do pliku .md z planem mapowania (arkusz Plan)",
    )
    parser.add_argument("--output", "-o", default=None, help="Ścieżka do pliku .xlsx")
    args = parser.parse_args()

    result = export_bi_view(
        sql=args.sql,
        view_name=args.view_name,
        source_table=args.source_table,
        plan_path=Path(args.plan) if args.plan else None,
        output_path=Path(args.output) if args.output else None,
    )
    print(json.dumps(result, default=str, ensure_ascii=False))


if __name__ == "__main__":
    main()
