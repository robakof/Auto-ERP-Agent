"""
ofertowanie_export.py — eksport produktów z ERP do Excel.

Tryby:
  --group          produkty z grupy 10_OFERTY (domyślnie)
  --excel PATH     produkty z listy kodów w pliku Excel (kolumna 'Kod' lub pierwsza)

Użycie:
  python tools/ofertowanie_export.py
  python tools/ofertowanie_export.py --excel lista.xlsx
  python tools/ofertowanie_export.py --output wynik.xlsx
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from sql_query import run_query

PHOTOS_DIR = r"D:\UdzialySieciowe\ZDJĘCIA\ZDJĘCIA PRODUKTÓW\jpg do systemu"
GROUP_NUMER = 9139   # CDN.TwrGrupy: 10_OFERTY
GROUP_TYP   = -16

# Kolejność kolumn w wyjściowym Excelu
COLUMNS_ORDER = [
    "Kod",
    "Zdjęcie",
    "Nazwa",
    "Grupa kartoteki",
    "Kod EAN",
    "ZAKUP",
    "CZAS PALENIA / DZIAŁANIA",
    "GRAMATURA WKŁADU",
    "SZEROKOŚĆ NETTO PRODUKTU",
    "WYSOKOŚĆ NETTO PRODUKTU",
    "SZEROKOŚĆ BRUTTO OPAKOWANIA",
    "warstwa",
    "opak.",
    "karton",
    "paleta",
    "Czy zdjęcie załadowane do systemu",
    "Czy zdjęcie zrobione png",
    "Cena 100",
    "Cena Brico",
    "Cena PSB",
]

# Kolumny obliczane w Python (nie z SQL)
PYTHON_COLS = {"Zdjęcie", "Czy zdjęcie załadowane do systemu", "Czy zdjęcie zrobione png"}

_SQL_BODY = """
SELECT
    tk.Twr_Kod                                AS [Kod],
    tk.Twr_Nazwa                              AS [Nazwa],
    tg_path.[Grupa kartoteki]                 AS [Grupa kartoteki],
    tk.Twr_Ean                                AS [Kod EAN],
    a9.Atr_Wartosc                            AS [CZAS PALENIA / DZIAŁANIA],
    a10.Atr_Wartosc                           AS [GRAMATURA WKŁADU],
    a11.Atr_Wartosc                           AS [SZEROKOŚĆ NETTO PRODUKTU],
    a12.Atr_Wartosc                           AS [WYSOKOŚĆ NETTO PRODUKTU],
    a13.Atr_Wartosc                           AS [SZEROKOŚĆ BRUTTO OPAKOWANIA],
    CAST(jm_warstwa.TwJ_PrzeliczL AS BIGINT)  AS [warstwa],
    CAST(jm_opak.TwJ_PrzeliczL    AS BIGINT)  AS [opak.],
    CAST(jm_karton.TwJ_PrzeliczL  AS BIGINT)  AS [karton],
    CAST(jm_paleta.TwJ_PrzeliczL  AS BIGINT)  AS [paleta],
    ISNULL(zak.CenaZakupu, 0)                 AS [ZAKUP],
    tc100.TwC_Wartosc                         AS [Cena 100],
    tc_brico.TwC_Wartosc                      AS [Cena Brico],
    tc_psb.TwC_Wartosc                        AS [Cena PSB]
FROM CDN.TwrKarty tk
JOIN Produkty p
    ON tk.Twr_GIDNumer = p.Twr_GIDNumer
    AND tk.Twr_GIDTyp  = p.Twr_GIDTyp
LEFT JOIN CDN.Atrybuty a9
    ON a9.Atr_ObiNumer = tk.Twr_GIDNumer AND a9.Atr_ObiTyp = tk.Twr_GIDTyp AND a9.Atr_AtkId = 9
LEFT JOIN CDN.Atrybuty a10
    ON a10.Atr_ObiNumer = tk.Twr_GIDNumer AND a10.Atr_ObiTyp = tk.Twr_GIDTyp AND a10.Atr_AtkId = 10
LEFT JOIN CDN.Atrybuty a11
    ON a11.Atr_ObiNumer = tk.Twr_GIDNumer AND a11.Atr_ObiTyp = tk.Twr_GIDTyp AND a11.Atr_AtkId = 11
LEFT JOIN CDN.Atrybuty a12
    ON a12.Atr_ObiNumer = tk.Twr_GIDNumer AND a12.Atr_ObiTyp = tk.Twr_GIDTyp AND a12.Atr_AtkId = 12
LEFT JOIN CDN.Atrybuty a13
    ON a13.Atr_ObiNumer = tk.Twr_GIDNumer AND a13.Atr_ObiTyp = tk.Twr_GIDTyp AND a13.Atr_AtkId = 13
LEFT JOIN CDN.TwrJM jm_opak
    ON jm_opak.TwJ_TwrNumer = tk.Twr_GIDNumer AND jm_opak.TwJ_TwrTyp = tk.Twr_GIDTyp
    AND jm_opak.TwJ_JmZ = 'opak.'
LEFT JOIN CDN.TwrJM jm_karton
    ON jm_karton.TwJ_TwrNumer = tk.Twr_GIDNumer AND jm_karton.TwJ_TwrTyp = tk.Twr_GIDTyp
    AND jm_karton.TwJ_JmZ = 'karton'
LEFT JOIN CDN.TwrJM jm_warstwa
    ON jm_warstwa.TwJ_TwrNumer = tk.Twr_GIDNumer AND jm_warstwa.TwJ_TwrTyp = tk.Twr_GIDTyp
    AND jm_warstwa.TwJ_JmZ IN ('warstwa', 'warsta')
LEFT JOIN CDN.TwrJM jm_paleta
    ON jm_paleta.TwJ_TwrNumer = tk.Twr_GIDNumer AND jm_paleta.TwJ_TwrTyp = tk.Twr_GIDTyp
    AND jm_paleta.TwJ_JmZ = 'paleta'
LEFT JOIN (
    SELECT TwZ_TwrNumer,
           CAST(SUM(TwZ_KsiegowaNetto) / IIF(SUM(TwZ_Ilosc)=0,1,SUM(TwZ_Ilosc))
                AS DECIMAL(20,4)) AS CenaZakupu
    FROM CDN.TwrZasoby
    GROUP BY TwZ_TwrNumer
) zak ON zak.TwZ_TwrNumer = tk.Twr_GIDNumer
LEFT JOIN (
    SELECT TwC_TwrNumer, TwC_TwrTyp, MAX(TwC_Wartosc) AS TwC_Wartosc
    FROM CDN.TwrCeny
    WHERE TwC_TcnId = 1 AND TwC_KntNumer = 0
    GROUP BY TwC_TwrNumer, TwC_TwrTyp
) tc100 ON tc100.TwC_TwrNumer = tk.Twr_GIDNumer AND tc100.TwC_TwrTyp = tk.Twr_GIDTyp
LEFT JOIN (
    SELECT TwC_TwrNumer, TwC_TwrTyp, MAX(TwC_Wartosc) AS TwC_Wartosc
    FROM CDN.TwrCeny
    WHERE TwC_TcnId = 8 AND TwC_KntNumer = 0
    GROUP BY TwC_TwrNumer, TwC_TwrTyp
) tc_brico ON tc_brico.TwC_TwrNumer = tk.Twr_GIDNumer AND tc_brico.TwC_TwrTyp = tk.Twr_GIDTyp
LEFT JOIN (
    SELECT TwC_TwrNumer, TwC_TwrTyp, MAX(TwC_Wartosc) AS TwC_Wartosc
    FROM CDN.TwrCeny
    WHERE TwC_TcnId = 10 AND TwC_KntNumer = 0
    GROUP BY TwC_TwrNumer, TwC_TwrTyp
) tc_psb ON tc_psb.TwC_TwrNumer = tk.Twr_GIDNumer AND tc_psb.TwC_TwrTyp = tk.Twr_GIDTyp
LEFT JOIN GrupaNazwa tg_path ON tg_path.TwrNumer = tk.Twr_GIDNumer
ORDER BY tk.Twr_Kod
"""

# Rekurencyjny CTE budujący pełną ścieżkę grupy, np. \10_Oferty\2026\Dino.
# Wstrzykiwany po ostatnim CTE w _sql_group() i _sql_excel().
_SQL_GROUP_PATH_CTE = """,
GroupPath AS (
    -- Anchor: wszystkie grupy produktu (bridge GIDTyp=16)
    SELECT
        br.TwG_GIDNumer                              AS TwrNumer,
        grp.TwG_GIDNumer                             AS GrpNumer,
        grp.TwG_GrONumer                             AS ParentNumer,
        CAST(RTRIM(grp.TwG_Nazwa) AS NVARCHAR(500))  AS Path
    FROM CDN.TwrGrupy br
    JOIN CDN.TwrGrupy grp
        ON grp.TwG_GIDTyp = -16 AND grp.TwG_GIDNumer = br.TwG_GrONumer
    WHERE br.TwG_GIDTyp = 16
    UNION ALL
    -- Rekurencja: idź w górę hierarchii
    SELECT
        gp.TwrNumer,
        par.TwG_GIDNumer,
        par.TwG_GrONumer,
        CAST(RTRIM(par.TwG_Nazwa) + '\\' + gp.Path AS NVARCHAR(500))
    FROM GroupPath gp
    JOIN CDN.TwrGrupy par
        ON par.TwG_GIDTyp = -16 AND par.TwG_GIDNumer = gp.ParentNumer
    WHERE gp.ParentNumer IS NOT NULL AND gp.ParentNumer <> 0
),
GrupaNazwa AS (
    -- Tylko ścieżki zawierające 'Cmentarz' (kategorie cmentarne)
    SELECT TwrNumer, '\\' + MIN(Path) AS [Grupa kartoteki]
    FROM GroupPath
    WHERE (ParentNumer IS NULL OR ParentNumer = 0)
      AND Path LIKE '%Cmentarz%'
    GROUP BY TwrNumer
)
"""

def _sql_group(group_numer=GROUP_NUMER, group_typ=GROUP_TYP):
    return f"""
WITH GroupTree AS (
    SELECT TwG_GIDNumer, TwG_GIDTyp
    FROM CDN.TwrGrupy
    WHERE TwG_GIDNumer = {group_numer} AND TwG_GIDTyp = {group_typ}
    UNION ALL
    SELECT g.TwG_GIDNumer, g.TwG_GIDTyp
    FROM CDN.TwrGrupy g
    INNER JOIN GroupTree gt
        ON g.TwG_GrONumer = gt.TwG_GIDNumer
        AND g.TwG_GrOTyp  = gt.TwG_GIDTyp
        AND g.TwG_GIDTyp  = {group_typ}
),
Produkty AS (
    SELECT DISTINCT tk.Twr_GIDNumer, tk.Twr_GIDTyp
    FROM CDN.TwrKarty tk
    JOIN CDN.TwrGrupy tg
        ON tk.Twr_GIDNumer = tg.TwG_GIDNumer
        AND tk.Twr_GIDTyp  = tg.TwG_GIDTyp
    JOIN GroupTree gt
        ON tg.TwG_GrONumer = gt.TwG_GIDNumer
        AND tg.TwG_GrOTyp  = gt.TwG_GIDTyp
    WHERE tk.Twr_Archiwalny = 0
){_SQL_GROUP_PATH_CTE}
{_SQL_BODY}
"""


def _sql_excel(kody: list[str]) -> str:
    kody_list = ", ".join(f"'{k}'" for k in kody)
    return f"""
WITH Produkty AS (
    SELECT DISTINCT Twr_GIDNumer, Twr_GIDTyp
    FROM CDN.TwrKarty
    WHERE Twr_Archiwalny = 0 AND Twr_Kod IN ({kody_list})
){_SQL_GROUP_PATH_CTE}
{_SQL_BODY}
"""


def _find_photo(kod: str) -> tuple:
    """Zwraca (nazwa_pliku|None, zaladowane: bool, ma_png: bool)."""
    png = os.path.join(PHOTOS_DIR, kod + ".png")
    jpg = os.path.join(PHOTOS_DIR, kod + ".jpg")
    has_png = os.path.exists(png)
    has_jpg = os.path.exists(jpg)
    zaladowane = has_png or has_jpg
    if has_png:
        filename = kod + ".png"
    elif has_jpg:
        filename = kod + ".jpg"
    else:
        filename = None
    return filename, zaladowane, has_png


def _read_excel_codes(path: str) -> list[str]:
    """Odczytuje kody produktów z Excela (kolumna 'Kod' lub pierwsza kolumna)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_idx = headers.index("Kod") if "Kod" in headers else 0
    kody = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_idx]
        if val:
            kody.append(str(val).strip())
    return kody


IMG_HEIGHT_PX = 75   # wysokość miniatury w pikselach
IMG_WIDTH_PX  = 75   # szerokość miniatury w pikselach
ROW_HEIGHT_PT = 58   # wysokość wiersza w punktach (~77px, z marginesem)
COL_PHOTO_IDX = COLUMNS_ORDER.index("Zdjęcie") + 1  # nr kolumny (1-based)


def _export_excel(rows: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ofertowanie"

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF")

    # Nagłówki
    for col_idx, col_name in enumerate(COLUMNS_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dane + obrazki
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS_ORDER, start=1):
            # Kolumna Zdjęcie: nie wpisujemy tekstu — bedzie obrazek
            if col_name == "Zdjęcie":
                continue
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top")

        # Obrazek
        photo_filename = row.get("Zdjęcie")
        if photo_filename:
            photo_path = os.path.join(PHOTOS_DIR, photo_filename)
            if os.path.exists(photo_path):
                try:
                    img = XlImage(photo_path)
                    marker = AnchorMarker(col=COL_PHOTO_IDX - 1, colOff=0, row=row_idx - 1, rowOff=0)
                    size = XDRPositiveSize2D(pixels_to_EMU(IMG_WIDTH_PX), pixels_to_EMU(IMG_HEIGHT_PX))
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                    ws.add_image(img)
                    ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT
                except Exception:
                    pass  # uszkodzony plik — pomijamy, nie crashujemy

    # Szerokości kolumn
    col_widths = {
        "Kod": 14, "Zdjęcie": 12, "Nazwa": 45, "Grupa kartoteki": 20, "Kod EAN": 16,
        "ZAKUP": 10, "CZAS PALENIA / DZIAŁANIA": 14, "GRAMATURA WKŁADU": 14,
        "SZEROKOŚĆ NETTO PRODUKTU": 14, "WYSOKOŚĆ NETTO PRODUKTU": 14,
        "SZEROKOŚĆ BRUTTO OPAKOWANIA": 14, "warstwa": 10, "opak.": 10,
        "karton": 10, "paleta": 10,
        "Czy zdjęcie załadowane do systemu": 16, "Czy zdjęcie zrobione png": 14,
        "Cena 100": 12, "Cena Brico": 12, "Cena PSB": 12,
    }
    for col_idx, col_name in enumerate(COLUMNS_ORDER, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLUMNS_ORDER))}1"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Eksport ofertowania z ERP do Excel")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--group", action="store_true",
                       help="Pobierz produkty z grupy 10_OFERTY (domyślnie)")
    group.add_argument("--excel", metavar="PLIK",
                       help="Pobierz produkty z listy kodów w pliku Excel")
    parser.add_argument("--output", metavar="PLIK",
                        default=f"output/ofertowanie_{date.today()}.xlsx",
                        help="Ścieżka pliku wyjściowego (domyślnie output/ofertowanie_YYYY-MM-DD.xlsx)")
    args = parser.parse_args()

    # Budowanie SQL
    if args.excel:
        print(f"Tryb: Excel — wczytuje kody z {args.excel}")
        kody = _read_excel_codes(args.excel)
        if not kody:
            print("ERROR: brak kodów w pliku Excel.")
            sys.exit(1)
        print(f"  Znaleziono {len(kody)} kodów.")
        sql = _sql_excel(kody)
    else:
        print("Tryb: Grupa 10_OFERTY")
        sql = _sql_group()

    # Zapytanie ERP
    print("Odpytuje ERP...")
    result = run_query(sql, inject_top=None)
    if not result["ok"]:
        print(f"ERROR ERP: {result['error']['message']}")
        sys.exit(1)

    columns = result["data"]["columns"]
    db_rows = result["data"]["rows"]
    print(f"  Pobrano {len(db_rows)} produktów.")

    # Składanie wierszy + kolumny Python
    rows = []
    for db_row in db_rows:
        row = dict(zip(columns, db_row))
        kod = row.get("Kod", "")
        filename, zaladowane, has_png = _find_photo(kod)
        row["Zdjęcie"] = filename
        row["Czy zdjęcie załadowane do systemu"] = "Tak" if zaladowane else "Nie"
        row["Czy zdjęcie zrobione png"] = "Tak" if has_png else "Nie"
        rows.append(row)

    # Eksport
    print(f"Zapisuje do {args.output}...")
    _export_excel(rows, args.output)
    print(f"OK: {len(rows)} wierszy zapisanych do {args.output}")


if __name__ == "__main__":
    main()
