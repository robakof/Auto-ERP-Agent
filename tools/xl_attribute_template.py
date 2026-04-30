"""Generuje Excel template z atrybutami produktów z CDN.AtrybutyKlasy (GIDTyp=16).

Format (zgodny ze wzorem):
  Wiersz 1:  (puste) | "Atrybut / Akronim →" | "Typ" | ...
  Wiersze 2+: KOD_XL | NAZWA_ATRYBUTU | TYP | wartość1 | wartość2 | ...
  (jeden wiersz = jeden produkt × jeden atrybut)

Użycie:
  python tools/xl_attribute_template.py
  python tools/xl_attribute_template.py --output sciezka/template.xlsx
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent.parent))
load_dotenv()

from tools.lib.sql_client import SqlClient

DEFAULT_OUTPUT = Path("documents/Wzory plików/Atrybuty produktów - template.xlsx")

TYPE_LABELS = {1: "TAK / NIE", 2: "tekst", 3: "liczba", 4: "lista*"}

_QUERY_ATTRS = """
SELECT ak.AtK_Nazwa, ak.AtK_Typ, ak.AtK_Wielowart, ak.AtK_Zamknieta
FROM CDN.AtrybutyKlasy ak
INNER JOIN CDN.AtrybutyObiekty ao ON ao.AtO_AtKId = ak.AtK_ID
WHERE ao.AtO_GIDTyp = 16
ORDER BY ak.AtK_Nazwa
"""

_QUERY_PRODUCTS_ALL = """
SELECT Twr_Kod, Twr_Nazwa1
FROM CDN.TwrKarty
WHERE Twr_Archiwalny = 0
ORDER BY Twr_Kod
"""

_QUERY_PRODUCTS_GROUP = """
WITH GTree AS (
    SELECT TwG_GIDNumer, TwG_GIDTyp
    FROM CDN.TwrGrupy
    WHERE TwG_GIDNumer = {group_id} AND TwG_GIDTyp = -16
    UNION ALL
    SELECT g.TwG_GIDNumer, g.TwG_GIDTyp
    FROM CDN.TwrGrupy g
    INNER JOIN GTree gt
        ON g.TwG_GrONumer = gt.TwG_GIDNumer
        AND g.TwG_GrOTyp  = gt.TwG_GIDTyp
        AND g.TwG_GIDTyp  = -16
)
SELECT DISTINCT tk.Twr_Kod, tk.Twr_Nazwa1
FROM CDN.TwrKarty tk
JOIN CDN.TwrGrupy tg
    ON tk.Twr_GIDNumer = tg.TwG_GIDNumer AND tk.Twr_GIDTyp = tg.TwG_GIDTyp
JOIN GTree gt2
    ON tg.TwG_GrONumer = gt2.TwG_GIDNumer AND tg.TwG_GrOTyp = gt2.TwG_GIDTyp
WHERE tk.Twr_Archiwalny = 0
ORDER BY tk.Twr_Kod
"""

_QUERY_EXISTING_VALUES = """
SELECT tk.Twr_Kod, ak.AtK_Nazwa, a.Atr_Wartosc
FROM CDN.Atrybuty a
JOIN CDN.AtrybutyKlasy ak ON ak.AtK_Id = a.Atr_AtkId
JOIN CDN.TwrKarty tk ON tk.Twr_GIDNumer = a.Atr_ObiNumer AND tk.Twr_GIDTyp = a.Atr_ObiTyp
WHERE a.Atr_ObiTyp = 16
ORDER BY tk.Twr_Kod, ak.AtK_Nazwa
"""

_QUERY_PRODUCTS_BY_AKRONIMY = """
SELECT Twr_Kod, Twr_Nazwa1
FROM CDN.TwrKarty
WHERE Twr_Kod IN ({placeholders})
ORDER BY Twr_Kod
"""

_QUERY_VALUES_BY_AKRONIMY = """
SELECT tk.Twr_Kod, ak.AtK_Nazwa, a.Atr_Wartosc
FROM CDN.Atrybuty a
JOIN CDN.AtrybutyKlasy ak ON ak.AtK_Id = a.Atr_AtkId
JOIN CDN.TwrKarty tk ON tk.Twr_GIDNumer = a.Atr_ObiNumer AND tk.Twr_GIDTyp = a.Atr_ObiTyp
WHERE a.Atr_ObiTyp = 16
  AND tk.Twr_Kod IN ({placeholders})
ORDER BY tk.Twr_Kod, ak.AtK_Nazwa
"""

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_KOD_FILL    = PatternFill("solid", fgColor="2E75B6")
_KOD_FONT    = Font(bold=True, color="FFFFFF", size=11)
_TYPE_FILL   = PatternFill("solid", fgColor="D6E4F0")
_TYPE_FONT   = Font(italic=True, color="2F5597", size=10)
_LIST_FILL   = PatternFill("solid", fgColor="FFF2CC")
_ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")


def _fetch(client: SqlClient, query: str) -> list:
    result = client.execute(query, inject_top=None)
    if not result["ok"]:
        raise RuntimeError(result["error"]["message"])
    return result["rows"]


def generate_template(output: Path) -> dict:
    client = SqlClient()

    group_id_raw = os.getenv("XL_ATTR_GROUP_ID", "").strip()
    if group_id_raw.isdigit():
        query_products = _QUERY_PRODUCTS_GROUP.format(group_id=int(group_id_raw))
    else:
        query_products = _QUERY_PRODUCTS_ALL

    attrs = _fetch(client, _QUERY_ATTRS)
    products = _fetch(client, query_products)
    existing_raw = _fetch(client, _QUERY_EXISTING_VALUES)

    # {(twr_kod, attr_name): [val1, val2, ...]}
    existing: dict[tuple, list] = defaultdict(list)
    for twr_kod, attr_name, val in existing_raw:
        if val:
            existing[(twr_kod.strip(), attr_name.strip())].append(val)

    wb = Workbook()
    ws = wb.active
    ws.title = "Atrybuty produktów"

    # --- Wiersz 1: nagłówek ---
    ws.cell(1, 1, "").fill = _KOD_FILL
    h_attr = ws.cell(1, 2, "Atrybut / Akronim →")
    h_attr.fill = _HEADER_FILL
    h_attr.font = _HEADER_FONT
    h_attr.alignment = Alignment(horizontal="center")
    h_typ = ws.cell(1, 3, "Typ")
    h_typ.fill = _HEADER_FILL
    h_typ.font = _HEADER_FONT
    h_typ.alignment = Alignment(horizontal="center")

    h_val1 = ws.cell(1, 4, "Wartość 1")
    h_val1.fill = PatternFill("solid", fgColor="375623")
    h_val1.font = Font(bold=True, color="FFFFFF", size=11)
    h_val1.alignment = Alignment(horizontal="center")

    h_val2 = ws.cell(1, 5, "Wartość 2")
    h_val2.fill = PatternFill("solid", fgColor="375623")
    h_val2.font = Font(bold=True, color="FFFFFF", size=11)
    h_val2.alignment = Alignment(horizontal="center")

    # --- Wiersze danych: jeden wiersz = jeden produkt × jeden atrybut ---
    row_idx = 2
    for prod_i, (twr_kod, twr_nazwa) in enumerate(products):
        for attr_i, (attr_name, typ, wielowart, zamknieta) in enumerate(attrs):
            type_label = TYPE_LABELS.get(typ, str(typ))
            if zamknieta:
                type_label += " (zamknięta)"

            alt = (prod_i % 2 == 1)
            bg = _ALT_FILL if alt else None

            kod_cell = ws.cell(row_idx, 1, twr_kod)
            kod_cell.font = Font(bold=True, size=10)
            if bg:
                kod_cell.fill = bg

            attr_cell = ws.cell(row_idx, 2, attr_name)
            if bg:
                attr_cell.fill = bg

            type_cell = ws.cell(row_idx, 3, type_label)
            type_cell.fill = _LIST_FILL if typ == 4 else (_TYPE_FILL if not bg else bg)
            type_cell.font = _TYPE_FONT
            type_cell.alignment = Alignment(horizontal="center")

            vals = existing.get((twr_kod.strip(), attr_name.strip()), [])
            for v_i, v in enumerate(vals):
                cell = ws.cell(row_idx, 4 + v_i, v)
                if typ == 4:
                    cell.fill = _LIST_FILL
                elif bg:
                    cell.fill = bg

            row_idx += 1

    # --- Wymiary kolumn ---
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 22

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "ok": True,
        "data": {"path": str(output), "attributes": len(attrs), "products": len(products)},
        "error": None,
    }


def generate_for_akronimy(akronimy: list[str], output: Path) -> dict:
    """Generuje Excel z aktualnymi wartościami dla podanych akronimów."""
    if not akronimy:
        return {"ok": False, "data": None, "error": {"type": "EMPTY_INPUT", "message": "Brak akronimów"}}

    client = SqlClient()
    placeholders = ",".join(["?"] * len(akronimy))

    attrs_result = client.execute(_QUERY_ATTRS, inject_top=None)
    if not attrs_result["ok"]:
        return {"ok": False, "data": None, "error": attrs_result["error"]}
    attrs = attrs_result["rows"]

    try:
        conn = client.get_connection()
        cursor = conn.cursor()
        cursor.execute(_QUERY_PRODUCTS_BY_AKRONIMY.format(placeholders=placeholders), akronimy)
        products = [list(r) for r in cursor.fetchall()]
        cursor.execute(_QUERY_VALUES_BY_AKRONIMY.format(placeholders=placeholders), akronimy)
        vals_rows = [list(r) for r in cursor.fetchall()]
    except Exception as exc:
        return {"ok": False, "data": None, "error": {"type": "SQL_ERROR", "message": str(exc)}}

    existing: dict[tuple, list] = defaultdict(list)
    for twr_kod, attr_name, val in vals_rows:
        if val:
            existing[(twr_kod.strip(), attr_name.strip())].append(val)

    wb = Workbook()
    ws = wb.active
    ws.title = "Atrybuty produktów"

    ws.cell(1, 1, "").fill = _KOD_FILL
    h_attr = ws.cell(1, 2, "Atrybut / Akronim →")
    h_attr.fill = _HEADER_FILL; h_attr.font = _HEADER_FONT
    h_attr.alignment = Alignment(horizontal="center")
    h_typ = ws.cell(1, 3, "Typ")
    h_typ.fill = _HEADER_FILL; h_typ.font = _HEADER_FONT
    h_typ.alignment = Alignment(horizontal="center")
    for col, label in [(4, "Wartość 1"), (5, "Wartość 2")]:
        c = ws.cell(1, col, label)
        c.fill = PatternFill("solid", fgColor="375623")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center")

    row_idx = 2
    for prod_i, (twr_kod, _) in enumerate(products):
        for attr_i, (attr_name, typ, wielowart, zamknieta) in enumerate(attrs):
            type_label = TYPE_LABELS.get(typ, str(typ))
            if zamknieta:
                type_label += " (zamknięta)"
            alt = (prod_i % 2 == 1)
            bg = _ALT_FILL if alt else None

            kod_cell = ws.cell(row_idx, 1, twr_kod)
            kod_cell.font = Font(bold=True, size=10)
            if bg:
                kod_cell.fill = bg
            attr_cell = ws.cell(row_idx, 2, attr_name)
            if bg:
                attr_cell.fill = bg
            type_cell = ws.cell(row_idx, 3, type_label)
            type_cell.fill = _LIST_FILL if typ == 4 else (_TYPE_FILL if not bg else bg)
            type_cell.font = _TYPE_FONT
            type_cell.alignment = Alignment(horizontal="center")

            vals = existing.get((twr_kod.strip(), attr_name.strip()), [])
            for v_i, v in enumerate(vals):
                cell = ws.cell(row_idx, 4 + v_i, v)
                if typ == 4:
                    cell.fill = _LIST_FILL
                elif bg:
                    cell.fill = bg
            row_idx += 1

    for col, w in zip("ABCDE", [16, 38, 22, 22, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 22

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    not_found = sorted(set(a.upper() for a in akronimy) - {r[0].upper() for r in products})
    return {
        "ok": True,
        "data": {"path": str(output), "products": len(products), "not_found": not_found},
        "error": None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generuj Excel template atrybutów produktów")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--akronimy", default=None,
                        help="Lista akronimów oddzielonych przecinkiem (tryb selektywny)")
    parser.add_argument("--akronimy-file", type=Path, default=None,
                        help="Plik tekstowy z akronimami (jeden per linia)")
    args = parser.parse_args()

    if args.akronimy or args.akronimy_file:
        akronimy: list[str] = []
        if args.akronimy:
            akronimy += [a.strip() for a in args.akronimy.split(",") if a.strip()]
        if args.akronimy_file:
            akronimy += [l.strip() for l in args.akronimy_file.read_text(encoding="utf-8").splitlines() if l.strip()]
        default_out = Path(f"documents/human/reports/xl_attribute_query_{__import__('datetime').date.today().strftime('%Y%m%d')}.xlsx")
        out = args.output if args.output != DEFAULT_OUTPUT else default_out
        result = generate_for_akronimy(akronimy, out)
    else:
        result = generate_template(args.output)

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
