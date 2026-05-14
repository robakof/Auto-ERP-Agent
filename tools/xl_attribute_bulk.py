"""Bulk update atrybutów produktów z pliku Excel.

Format pliku (zgodny ze wzorem):
  Wiersz 1:  (puste) | "Atrybut / Akronim →" | "Typ" | ...
  Wiersze 2+: KOD_XL | NAZWA_ATRYBUTU | TYP | wartość1 | wartość2 | ...
  (jeden wiersz = jeden produkt × jeden atrybut)
  Brak wartości w D+ = pomiń (nie aktualizuj).

Użycie:
  python tools/xl_attribute_bulk.py --file atrybuty.xlsx
  python tools/xl_attribute_bulk.py --file atrybuty.xlsx --report raport.xlsx
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.lib import xl_session
from tools.lib.sql_client import SqlClient
from tools.xl_attribute_set import delete_attributes, set_attribute

_QUERY_CLASS_NAMES = """
SELECT DISTINCT KlasaAtrybutu
FROM AIOP.vAtrybutyTowarow
"""

_STATUS_FILLS = {
    "OK":        PatternFill("solid", fgColor="C6EFCE"),
    "BŁĄD":      PatternFill("solid", fgColor="FFC7CE"),
    "POMINIĘTY": PatternFill("solid", fgColor="FFEB9C"),
}


def _load_class_map(client: SqlClient) -> tuple[dict, dict | None]:
    result = client.execute(_QUERY_CLASS_NAMES, inject_top=None)
    if not result["ok"]:
        return {}, result["error"]
    return {row[0].strip().lower(): row[0] for row in result["rows"]}, None


def _parse_excel(path: Path) -> tuple[list, list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]


def _write_report(path: Path, results: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"

    headers = ["Akronim", "Atrybut", "Wartość", "Status", "Komunikat"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for row_idx, r in enumerate(results, 2):
        ws.cell(row_idx, 1, r["akronim"])
        ws.cell(row_idx, 2, r["class"])
        ws.cell(row_idx, 3, r["value"] if r["value"] is not None else "(pusta)")
        ws.cell(row_idx, 4, r["status"])
        ws.cell(row_idx, 5, r.get("message", ""))
        fill = _STATUS_FILLS.get(r["status"])
        if fill:
            for col in range(1, 6):
                ws.cell(row_idx, col).fill = fill

    for col, width in zip(range(1, 6), [18, 34, 20, 12, 40]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    wb.save(path)


def bulk_update(file: Path, operator: str | None = None, report: Path | None = None,
                update: bool = False) -> dict:
    header, data_rows = _parse_excel(file)

    if not header or not data_rows:
        return {"ok": False, "data": None, "error": {"type": "EMPTY_FILE", "message": "Plik jest pusty"}}

    client = SqlClient()
    class_map, err = _load_class_map(client)
    if err:
        return {"ok": False, "data": None, "error": err}

    # Tryb --update: ustal produkty do podmiany i usuń ich atrybuty przed insertem
    failed_akronimy: set[str] = set()
    if update:
        akronimy_to_update: set[str] = set()
        for row in data_rows:
            if not row or row[0] is None or str(row[0]).strip() == "":
                continue
            has_value = any(
                row[i] is not None and str(row[i]).strip() != ""
                for i in range(3, len(row))
            )
            if has_value:
                akronimy_to_update.add(str(row[0]).strip())
        for akronim in akronimy_to_update:
            del_res = delete_attributes(akronim)
            if not del_res["ok"]:
                failed_akronimy.add(akronim)

    results = []
    total = success = skipped = failed = 0

    try:
        for row in data_rows:
            if not row or row[0] is None or str(row[0]).strip() == "":
                continue

            akronim = str(row[0]).strip()
            attr_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

            if not attr_raw:
                continue

            # wartości od kolumny D (indeks 3) wzwyż — zbierz wszystkie niepuste
            values = [
                str(row[i]).strip()
                for i in range(3, len(row))
                if row[i] is not None and str(row[i]).strip() != ""
            ]

            total += 1

            if not values:
                skipped += 1
                results.append({"akronim": akronim, "class": attr_raw, "value": None, "status": "POMINIĘTY"})
                continue

            if akronim in failed_akronimy:
                failed += 1
                results.append({"akronim": akronim, "class": attr_raw, "value": ", ".join(values) if values else None,
                                 "status": "BŁĄD", "message": f"Błąd usuwania atrybutów: {akronim}"})
                continue

            exact_name = class_map.get(attr_raw.lower())
            if exact_name is None:
                failed += 1
                results.append({"akronim": akronim, "class": attr_raw, "value": ", ".join(values),
                                 "status": "BŁĄD", "message": f"Nieznana klasa atrybutu: '{attr_raw}'"})
                continue

            # wstaw każdą wartość osobno (obsługa wielowartościowych)
            row_ok = True
            for v in values:
                res = set_attribute(exact_name, v, akronim, obj_type=16, operator=operator)
                if not res["ok"]:
                    row_ok = False
                    failed += 1
                    results.append({"akronim": akronim, "class": attr_raw, "value": v,
                                     "status": "BŁĄD", "message": res["error"]["message"]})

            if row_ok:
                success += 1
                results.append({"akronim": akronim, "class": attr_raw,
                                 "value": ", ".join(values), "status": "OK"})
    finally:
        try:
            xl_session.logout()
        except Exception:
            pass

    if report:
        _write_report(path=report, results=results)

    return {
        "ok": True,
        "data": {
            "total": total, "success": success, "skipped": skipped, "failed": failed,
            "report": str(report) if report else None,
            "results": results,
        },
        "error": None,
    }


def main() -> None:
    today = date.today().strftime("%Y%m%d")
    default_report = Path(f"documents/human/reports/xl_attribute_bulk_{today}.xlsx")

    parser = argparse.ArgumentParser(description="Bulk update atrybutów produktów z Excel")
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--operator", default=None)
    parser.add_argument("--report", type=Path, default=default_report)
    parser.add_argument("--update", action="store_true",
                        help="Tryb aktualizacji: usuń istniejące atrybuty przed insertem")
    args = parser.parse_args()

    result = bulk_update(args.file, args.operator, args.report, args.update)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
