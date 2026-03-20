"""
wycena_generate.py — Generator Wyceny Zniczy.

Pobiera produkty CZNI z wybranej grupy oferty ERP, generuje plik Excel
z 16 wierszami BOM na produkt. Wynik oparty na kopii szablonu xlsm.
Zapis przez xlwings (COM) — zachowuje tabele, formuły i makra Excela.

CLI:
    python tools/wycena_generate.py --offer-group-id 10729 --client-name "AUCHAN"

Opcje:
    --offer-group-id INT    GIDNumer grupy oferty z CDN.TwrGrupy
    --client-name STR       Nazwa klienta wpisywana w kolumnę A
    --template PATH         Domyślnie: "Wycena 2026 Otorowo Szablon.xlsm"
    --output PATH           Domyślnie: "Wycena {nazwa_oferty}.xlsm" (nazwa z ERP)
"""

import argparse
import logging
import shutil
import sys
from pathlib import Path

import openpyxl
import xlwings as xw

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.lib.sql_client import SqlClient
from tools.lib.wycena_bom import build_bom_rows

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

_PROJECT_ROOT = Path(__file__).parent.parent
TEMPLATE_PATH = _PROJECT_ROOT / "Wycena 2026 Otorowo Szablon.xlsm"
DEKLE_PATH = _PROJECT_ROOT / "dekle.xlsx"
SPODY_PATH = _PROJECT_ROOT / "spody.xlsx"
TACKA_PATH = _PROJECT_ROOT / "tacka.xlsx"
WKLADY_PATH = _PROJECT_ROOT / "waga składak.xlsx"
SHEET_NAME = "Wycena Zniczy"
DATA_START_ROW = 4

# Indeksy kolumn (1-based: A=1, B=2, E=5, G=7, H=8, J=10)
COL_A = 1
COL_B = 2
COL_E = 5
COL_G = 7
COL_H = 8
COL_J = 10

# Kolumny z formułami w szablonie — wymagają FillDown po wpisaniu danych
# (szablon ma formuły tylko w 16 przykładowych wierszach)
FORMULA_COLS = [3, 4, 6, 9, 11, 12, 13, 14, 15, 17, 18, 19, 20, 21, 22]


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------

SQL_OFFER_NAME = """
SELECT TwG_Nazwa
FROM CDN.TwrGrupy
WHERE TwG_GIDNumer = {offer_group_id}
"""

SQL_PRODUCTS = """
SELECT TwG_Kod AS produkt_kod, TwG_Nazwa AS produkt_nazwa
FROM CDN.TwrGrupy
WHERE TwG_GrONumer = {offer_group_id}
  AND TwG_GIDTyp = 16
  AND TwG_Kod LIKE 'CZNI%'
ORDER BY TwG_Kod
"""

SQL_UNITS = """
SELECT j.TwJ_JmZ, j.TwJ_PrzeliczL
FROM CDN.TwrKarty tw
JOIN CDN.TwrJm j ON j.TwJ_TwrNumer = tw.Twr_GIDNumer
                 AND j.TwJ_TwrTyp = tw.Twr_GIDTyp
WHERE tw.Twr_Kod = '{produkt_kod}'
  AND j.TwJ_JmZ IN ('opak.', 'paleta', 'warstwa')
"""

SQL_EAN = """
SELECT TOP 1 ean.TwK_Kod
FROM CDN.TwrKarty tw
JOIN CDN.TwrKody ean ON ean.TwK_TwrNumer = tw.Twr_GIDNumer
WHERE tw.Twr_Kod = '{produkt_kod}'
  AND ean.TwK_Domyslny = 1
"""

SQL_GRAMATURA_WKLADU = """
SELECT a.Atr_Wartosc
FROM CDN.Atrybuty a
JOIN CDN.TwrKarty tw ON tw.Twr_GIDNumer = a.Atr_ObiNumer
                    AND tw.Twr_GIDTyp = a.Atr_ObiTyp
WHERE a.Atr_AtkId = 10
  AND tw.Twr_Kod = '{produkt_kod}'
"""

SQL_SREDNICA = """
SELECT a.Atr_Wartosc
FROM CDN.Atrybuty a
JOIN CDN.TwrKarty tw ON tw.Twr_GIDNumer = a.Atr_ObiNumer
                     AND tw.Twr_GIDTyp = a.Atr_ObiTyp
WHERE a.Atr_AtkId = 56
  AND tw.Twr_Kod = '{produkt_kod}'
"""

SQL_SZKLO_3DIG = """
SELECT TOP 1 Twr_Kod
FROM CDN.TwrKarty
WHERE Twr_Kod LIKE 'SZ%'
  AND Twr_Nazwa LIKE '%' + SUBSTRING('{produkt_nazwa}', PATINDEX('%[0-9][0-9][0-9]%', '{produkt_nazwa}'), 3) + '%'
"""

SQL_SZKLO_PREFIX = """
SELECT TOP 1 Twr_Kod
FROM CDN.TwrKarty
WHERE Twr_Kod LIKE 'SZ%'
  AND Twr_Nazwa LIKE '%' + LEFT('{produkt_nazwa}', CHARINDEX(' ', '{produkt_nazwa}' + ' ', CHARINDEX(' ', '{produkt_nazwa}' + ' ') + 1) - 1) + '%'
"""


# ---------------------------------------------------------------------------
# ERP queries
# ---------------------------------------------------------------------------

def _fetch_offer_name(client: SqlClient, offer_group_id: int) -> str:
    result = client.execute(SQL_OFFER_NAME.format(offer_group_id=offer_group_id), inject_top=1)
    if result["ok"] and result["rows"]:
        return result["rows"][0][0]
    return str(offer_group_id)


def _fetch_products(client: SqlClient, offer_group_id: int) -> list[dict]:
    sql = SQL_PRODUCTS.format(offer_group_id=offer_group_id)
    result = client.execute(sql, inject_top=None)
    if not result["ok"]:
        log.error("Błąd pobierania produktów: %s", result["error"]["message"])
        return []
    return [{"kod": row[0], "nazwa": row[1]} for row in result["rows"]]


def _fetch_units(client: SqlClient, produkt_kod: str) -> tuple[float | None, float | None, float | None]:
    sql = SQL_UNITS.format(produkt_kod=produkt_kod.replace("'", "''"))
    result = client.execute(sql, inject_top=None)
    paletka = None
    paleta = None
    warstwa = None
    if result["ok"]:
        for row in result["rows"]:
            jmz, przel = row[0], row[1]
            try:
                val = float(str(przel).replace(",", "."))
            except (ValueError, TypeError):
                val = None
            if jmz == "opak.":
                paletka = val
            elif jmz == "paleta":
                paleta = val
            elif jmz == "warstwa":
                warstwa = val
    if paletka is None:
        log.warning("Brak jednostki 'opak.' dla %s", produkt_kod)
    if warstwa is None:
        log.warning("Brak jednostki 'warstwa' dla %s", produkt_kod)
    return paletka, paleta, warstwa


def _fetch_szklo(client: SqlClient, produkt_nazwa: str) -> str | None:
    safe_nazwa = produkt_nazwa.replace("'", "''")

    if any(c.isdigit() for c in produkt_nazwa):
        sql = SQL_SZKLO_3DIG.format(produkt_nazwa=safe_nazwa)
        result = client.execute(sql, inject_top=1)
        if result["ok"] and result["rows"]:
            return result["rows"][0][0]

    sql = SQL_SZKLO_PREFIX.format(produkt_nazwa=safe_nazwa)
    result = client.execute(sql, inject_top=1)
    if result["ok"] and result["rows"]:
        return result["rows"][0][0]

    log.warning("Nie znaleziono Szkła dla: %s", produkt_nazwa)
    return None


# ---------------------------------------------------------------------------
# EAN
# ---------------------------------------------------------------------------

def _fetch_ean(client: SqlClient, produkt_kod: str) -> str | None:
    sql = SQL_EAN.format(produkt_kod=produkt_kod.replace("'", "''"))
    result = client.execute(sql, inject_top=1)
    if result["ok"] and result["rows"]:
        return str(result["rows"][0][0])
    return None


# ---------------------------------------------------------------------------
# Paletka (Tacka) — ładowanie tabeli i dopasowanie
# ---------------------------------------------------------------------------

def _load_tacka(path: Path) -> dict[int, dict[str, str]]:
    """Wczytuje tacka.xlsx (Arkusz2) → {na_warstwie: {'ceim': kod, 'bez': kod}}."""
    if not path.exists():
        raise FileNotFoundError(f"Plik tacka.xlsx nie istnieje: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Arkusz2"]
    result: dict[int, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kod, nazwa, na_warstwie = row[0], row[1], row[2]
        if kod is None or na_warstwie is None:
            continue
        try:
            nw = int(na_warstwie)
        except (ValueError, TypeError):
            continue
        nazwa_lower = (nazwa or "").lower()
        is_ceim = "ceim" in nazwa_lower or ("nadruk" in nazwa_lower and "bez" not in nazwa_lower)
        typ = "ceim" if is_ceim else "bez"
        result.setdefault(nw, {})[typ] = str(kod)
    wb.close()
    log.info("Wczytano %d rozmiarow tacek z %s", len(result), path.name)
    return result


def _find_paletka(
    tacka: dict[int, dict[str, str]],
    paletka: float | None,
    warstwa: float | None,
    ean: str | None,
    produkt_kod: str,
) -> str | None:
    if paletka is None or warstwa is None:
        log.warning("Brak danych do obliczenia Na warstwie dla %s -- Paletka pusta", produkt_kod)
        return None
    na_warstwie = int(round(warstwa / paletka))
    if na_warstwie not in tacka:
        log.warning("Brak tacki dla Na warstwie=%d dla %s -- Paletka pusta", na_warstwie, produkt_kod)
        return None
    is_ceim = ean is not None and ean.startswith("59077025")
    typ = "ceim" if is_ceim else "bez"
    warianty = tacka[na_warstwie]
    if typ in warianty:
        return warianty[typ]
    fallback = next(iter(warianty.values()))
    log.warning("Brak wariantu '%s' dla Na warstwie=%d u %s -- uzyto %s", typ, na_warstwie, produkt_kod, fallback)
    return fallback


# ---------------------------------------------------------------------------
# Wkład — ładowanie tabeli i dopasowanie
# ---------------------------------------------------------------------------

def _load_wklady(path: Path) -> dict[int, str]:
    """Wczytuje waga skladak.xlsx (Arkusz1) -> {waga_g: kod}."""
    if not path.exists():
        raise FileNotFoundError(f"Plik waga skladak.xlsx nie istnieje: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Arkusz1"]
    result: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        waga, kod = row[0], row[1]
        if waga is None or kod is None:
            continue
        try:
            result[int(waga)] = str(kod)
        except (ValueError, TypeError):
            continue
    wb.close()
    log.info("Wczytano %d wkladow z %s", len(result), path.name)
    return result


def _fetch_gramatura_wkladu(client: SqlClient, produkt_kod: str) -> int | None:
    sql = SQL_GRAMATURA_WKLADU.format(produkt_kod=produkt_kod.replace("'", "''"))
    result = client.execute(sql, inject_top=1)
    if result["ok"] and result["rows"]:
        try:
            return int(float(str(result["rows"][0][0]).replace(",", ".")))
        except (ValueError, TypeError):
            pass
    return None


def _find_wklad(
    wklady: dict[int, str],
    gramatura: int | None,
    produkt_kod: str,
) -> str | None:
    if gramatura is None:
        return None
    kod = wklady.get(gramatura)
    if kod is None:
        log.warning("Brak wkladu dla gramatura=%dg dla %s -- Wklad pusty", gramatura, produkt_kod)
    return kod


# ---------------------------------------------------------------------------
# Dekiel — ładowanie tabeli i dopasowanie
# ---------------------------------------------------------------------------

def _load_dekle(path: Path) -> dict[float, list[tuple[str, str]]]:
    """Wczytuje dekle.xlsx (Arkusz2) → {srednica: [(kod, nazwa), ...]}."""
    if not path.exists():
        raise FileNotFoundError(f"Plik dekle.xlsx nie istnieje: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Arkusz2"]
    result: dict[float, list[tuple[str, str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kod, nazwa, srednica = row[0], row[1], row[2]
        if kod is None or srednica is None:
            continue
        try:
            sr = float(srednica)
        except (ValueError, TypeError):
            continue
        result.setdefault(sr, []).append((str(kod), str(nazwa) if nazwa else ""))
    wb.close()
    log.info("Wczytano %d rozmiarów dekli z %s", len(result), path.name)
    return result


def _fetch_srednica(client: SqlClient, produkt_kod: str) -> float | None:
    sql = SQL_SREDNICA.format(produkt_kod=produkt_kod.replace("'", "''"))
    result = client.execute(sql, inject_top=1)
    if result["ok"] and result["rows"]:
        try:
            return float(str(result["rows"][0][0]).replace(",", "."))
        except (ValueError, TypeError):
            pass
    return None


def _find_dekiel(
    dekle: dict[float, list[tuple[str, str]]],
    srednica: float | None,
    produkt_kod: str,
) -> tuple[str | None, float | None]:
    """Zwraca (kod_dekla, matched_size_cm) lub (None, None) gdy brak dopasowania."""
    if srednica is None:
        log.warning("Brak średnicy otworu dla %s — Dekiel pusty", produkt_kod)
        return None, None
    # Najbliższy rozmiar >= średnicy produktu
    dopasowany = min((s for s in dekle if s >= srednica), default=None)
    if dopasowany is None:
        log.warning("Brak dekla >= %.1f cm dla %s — Dekiel pusty", srednica, produkt_kod)
        return None, None
    if dopasowany != srednica:
        log.info("Dekiel: %.1f cm -> %.1f cm (najblizszy w gore) dla %s", srednica, dopasowany, produkt_kod)
    kandydaci = dekle[dopasowany]
    rapcewicz = [kod for kod, nazwa in kandydaci if "rapcewicz" in nazwa.lower()]
    kod = min(rapcewicz) if rapcewicz else min(k for k, _ in kandydaci)
    return kod, dopasowany


# ---------------------------------------------------------------------------
# Spód — ładowanie tabeli i dopasowanie
# ---------------------------------------------------------------------------

def _load_spody(path: Path) -> dict[int, list[tuple[str, str]]]:
    """Wczytuje spody.xlsx (Arkusz2) → {rozmiar_mm: [(kod, nazwa), ...]}."""
    if not path.exists():
        raise FileNotFoundError(f"Plik spody.xlsx nie istnieje: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Arkusz2"]
    result: dict[int, list[tuple[str, str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kod, nazwa, rozmiar = row[0], row[1], row[2]
        if kod is None or rozmiar is None:
            continue
        try:
            rm = int(rozmiar)
        except (ValueError, TypeError):
            continue
        result.setdefault(rm, []).append((str(kod), str(nazwa) if nazwa else ""))
    wb.close()
    log.info("Wczytano %d rozmiarów spodów z %s", len(result), path.name)
    return result


def _find_spod(
    spody: dict[int, list[tuple[str, str]]],
    dekiel_size_cm: float | None,
    produkt_kod: str,
) -> str | None:
    """Zwraca kod spodu o następnym rozmiarze powyżej rozmiaru dekla."""
    if dekiel_size_cm is None:
        log.warning("Brak rozmiaru dekla dla %s — Spód pusty", produkt_kod)
        return None
    dekiel_mm = round(dekiel_size_cm * 10)
    # Najbliższy rozmiar ściśle > rozmiaru dekla
    dopasowany = min((s for s in spody if s > dekiel_mm), default=None)
    if dopasowany is None:
        log.warning("Brak spodu > %d mm dla %s — Spód pusty", dekiel_mm, produkt_kod)
        return None
    log.info("Spod: dekiel %d mm -> spod %d mm dla %s", dekiel_mm, dopasowany, produkt_kod)
    kandydaci = spody[dopasowany]
    return min(kod for kod, _ in kandydaci)


# ---------------------------------------------------------------------------
# Budowanie danych BOM
# ---------------------------------------------------------------------------

def _build_column_data(
    products: list[dict],
    client: SqlClient,
    dekle: dict,
    spody: dict,
    tacka: dict,
    wklady: dict,
    offer_group_id: int,
    client_name: str,
) -> dict[int, list]:
    """Zwraca słownik {numer_kolumny: lista_wartości} dla wszystkich produktów."""
    cols = {COL_A: [], COL_B: [], COL_E: [], COL_G: [], COL_H: [], COL_J: []}

    for prod in products:
        kod = prod["kod"]
        nazwa = prod["nazwa"]

        paletka, paleta, warstwa = _fetch_units(client, kod)
        ean = _fetch_ean(client, kod)
        szklo_akronim = _fetch_szklo(client, nazwa)
        srednica = _fetch_srednica(client, kod)
        dekiel_akronim, dekiel_size_cm = _find_dekiel(dekle, srednica, kod)
        spod_akronim = _find_spod(spody, dekiel_size_cm, kod)
        paletka_akronim = _find_paletka(tacka, paletka, warstwa, ean, kod)
        gramatura = _fetch_gramatura_wkladu(client, kod)
        wklad_akronim = _find_wklad(wklady, gramatura, kod)

        bom_rows = build_bom_rows(
            paletka=paletka,
            paleta=paleta,
            szklo_akronim=szklo_akronim,
            dekiel_akronim=dekiel_akronim,
            spod_akronim=spod_akronim,
            paletka_akronim=paletka_akronim,
            wklad_akronim=wklad_akronim,
            offer_group_id=offer_group_id,
        )

        for i, bom in enumerate(bom_rows):
            cols[COL_A].append(client_name if i == 0 else None)
            cols[COL_B].append(kod)
            cols[COL_E].append(bom.wlasciwosc)
            cols[COL_G].append(bom.nazwa)
            cols[COL_H].append(bom.akronim)
            cols[COL_J].append(bom.mianownik)

    return cols


# ---------------------------------------------------------------------------
# Excel write przez xlwings (COM)
# ---------------------------------------------------------------------------

def _write_to_excel(output: Path, col_data: dict[int, list]) -> None:
    """Otwiera plik przez COM Excela, zapisuje kolumny, wypełnia formuły, zamyka."""
    n_rows = len(col_data[COL_A])
    last_row = DATA_START_ROW + n_rows - 1
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(str(output.resolve()))
        ws = wb.sheets[SHEET_NAME]

        for col_idx, values in col_data.items():
            col_letter = _col_letter(col_idx)
            start_cell = f"{col_letter}{DATA_START_ROW}"
            ws.range(start_cell).options(transpose=True).value = values

        # FillDown na kolumnach z formułami — replikuje auto-fill Excela
        # (szablon ma formuły tylko w przykładowych wierszach)
        for col in FORMULA_COLS:
            ws.range((DATA_START_ROW, col), (last_row, col)).api.FillDown()
        log.info("FillDown wykonany dla %d kolumn formuł, wiersze %d–%d",
                 len(FORMULA_COLS), DATA_START_ROW, last_row)

        wb.save()
        wb.close()
        log.info("Zapisano %d wierszy przez xlwings", n_rows)
    finally:
        app.quit()


def _col_letter(n: int) -> str:
    """Zamienia numer kolumny (1-based) na literę (A, B, ..., Z, AA, ...)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate(offer_group_id: int, client_name: str, template: Path, output: Path | None) -> Path:
    if not template.exists():
        raise FileNotFoundError(f"Szablon nie istnieje: {template}")

    client = SqlClient()

    offer_name = _fetch_offer_name(client, offer_group_id)
    log.info("Oferta: %s (%s)", offer_name, offer_group_id)

    if output is None:
        safe_name = offer_name.replace("/", "-").replace("\\", "-")
        output = Path(f"Wycena {safe_name}.xlsm")

    log.info("Plik wynikowy: %s", output)
    try:
        shutil.copy2(template, output)
    except PermissionError:
        raise PermissionError(
            f"Nie można nadpisać '{output}' — plik jest otwarty w Excelu. "
            "Zamknij plik lub podaj inną ścieżkę przez --output."
        )

    products = _fetch_products(client, offer_group_id)
    log.info("Produktów: %d", len(products))

    dekle = _load_dekle(DEKLE_PATH)
    spody = _load_spody(SPODY_PATH)
    tacka = _load_tacka(TACKA_PATH)
    wklady = _load_wklady(WKLADY_PATH)
    col_data = _build_column_data(products, client, dekle, spody, tacka, wklady, offer_group_id, client_name)

    _write_to_excel(output, col_data)

    total_rows = len(col_data[COL_A])
    log.info("Gotowe: %s (%d produktów, %d wierszy)", output, len(products), total_rows)
    return output


def main():
    parser = argparse.ArgumentParser(description="Generator Wyceny Zniczy")
    parser.add_argument("--offer-group-id", type=int, required=True, dest="offer_group_id",
                        help="GIDNumer grupy oferty z CDN.TwrGrupy")
    parser.add_argument("--client-name", required=True, dest="client_name",
                        help="Nazwa klienta wpisywana w kolumnę A")
    parser.add_argument("--template", default=None,
                        help=f"Ścieżka do szablonu (domyślnie: {TEMPLATE_PATH})")
    parser.add_argument("--output", default=None,
                        help="Ścieżka do pliku wynikowego (domyślnie: Wycena {nazwa_oferty}.xlsm)")
    args = parser.parse_args()

    template = Path(args.template) if args.template else TEMPLATE_PATH
    output = Path(args.output) if args.output else None

    try:
        result = generate(
            offer_group_id=args.offer_group_id,
            client_name=args.client_name,
            template=template,
            output=output,
        )
        print(f"OK: {result}")
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
