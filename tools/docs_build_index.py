"""
build_index.py — Budowa indeksu SQLite FTS5 z pliku XLSM ERP.

CLI:
    python tools/build_index.py [--xlsm PATH] [--db PATH]

Uruchamiany jednorazowo (lub przy aktualizacji dokumentacji).
Zawsze buduje indeks od zera — idempotentne.
"""

import argparse
import os
import sys
from collections import defaultdict
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

# Umożliwia import tools.db zarówno przy uruchomieniu skryptu jak i przy imporcie w testach
sys.path.insert(0, str(Path(__file__).parent.parent))
from tools.db import get_db

load_dotenv()

XLSM_PATH = os.getenv("XLSM_PATH", "")

# Kolumny arkusza Tabele (nagłówek na wierszu 1, dane od wiersza 2)
COL_TABELE = {
    "table_id": 0,
    "table_label": 1,
    "table_name": 2,    # CDN.XXX
    "prefix": 3,
    "description": 5,
}

# Kolumny arkusza Kolumny (nagłówek szukany po wartości "Numer tabeli")
COL_KOLUMNY = {
    "table_id": 0,
    "table_label": 1,
    "table_name": 2,
    "col_id": 3,
    "col_label": 4,
    "col_name": 5,
    "role": 6,
    "is_useful": 7,
    "preferred": 8,
    "data_type": 9,
    "description": 11,
}

# Kolumny arkusza Relacje (nagłówek na wierszu 1)
COL_RELACJE = {
    "source_table": 11,  # CDN.XXX — pełna nazwa ze schematem
    "source_col": 4,
    "target_table": 2,   # bez schematu — normalizowane do CDN.XXX
    "target_col": 5,
}

# Kolumny arkusza Słownik wartości kolumn (nagłówek na wierszu 1)
COL_SLOWNIK = {
    "table_name": 2,    # CDN.XXX
    "col_name": 5,
    "value": 6,
    "meaning": 7,
}

# Kolumny arkusza Przykładowe wartości kolumn (nagłówek na wierszu 1)
COL_PRZYKLADOWE = {
    "table_name": 2,    # CDN.XXX
    "col_name": 5,
    "sample_value": 6,
}


_EXCEL_ERRORS = {"None", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NUM!", "#DIV/0!", "#NULL!"}


def _str(val) -> str:
    """Bezpieczna konwersja wartości komórki na string. None/błędy Excela → ''."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in _EXCEL_ERRORS else s


def _normalize_table(name: str) -> str:
    """Dodaje prefiks CDN. jeśli tabela nie ma schematu (brak kropki)."""
    name = name.strip()
    return name if "." in name else f"CDN.{name}"


def parse_tables(ws) -> list[dict]:
    """Parsuje arkusz Tabele → lista słowników."""
    tables = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # nagłówek
            continue
        if row[COL_TABELE["table_id"]] is None:
            continue
        table_name = _str(row[COL_TABELE["table_name"]])
        if not table_name:
            continue
        tables.append({
            "table_id": row[COL_TABELE["table_id"]],
            "table_name": table_name,
            "table_label": _str(row[COL_TABELE["table_label"]]),
            "prefix": _str(row[COL_TABELE["prefix"]]),
            "description": _str(row[COL_TABELE["description"]]),
        })
    return tables


def parse_columns(ws) -> list[dict]:
    """Parsuje arkusz Kolumny → lista słowników. Header szukany po 'Numer tabeli'."""
    columns = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "Numer tabeli":
                header_found = True
            continue
        if row[COL_KOLUMNY["table_id"]] is None:
            continue
        col_name = _str(row[COL_KOLUMNY["col_name"]])
        if not col_name:
            continue
        col_label = _str(row[COL_KOLUMNY["col_label"]])
        columns.append({
            "table_name": _str(row[COL_KOLUMNY["table_name"]]),
            "table_label": _str(row[COL_KOLUMNY["table_label"]]),
            "col_id": row[COL_KOLUMNY["col_id"]],
            "col_name": col_name,
            "col_label": "" if col_label == "0" else col_label,
            "role": _str(row[COL_KOLUMNY["role"]]),
            "is_useful": _str(row[COL_KOLUMNY["is_useful"]]),
            "preferred": _str(row[COL_KOLUMNY["preferred"]]),
            "data_type": _str(row[COL_KOLUMNY["data_type"]]),
            "description": _str(row[COL_KOLUMNY["description"]]),
        })
    return columns


def parse_relations(ws) -> list[dict]:
    """Parsuje arkusz Relacje → lista słowników."""
    relations = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # nagłówek
            continue
        source_table = _str(row[COL_RELACJE["source_table"]])
        if not source_table:
            continue
        relations.append({
            "source_table": source_table,
            "source_col": _str(row[COL_RELACJE["source_col"]]),
            "target_table": _normalize_table(_str(row[COL_RELACJE["target_table"]])),
            "target_col": _str(row[COL_RELACJE["target_col"]]),
        })
    return relations


def parse_value_dicts(ws) -> dict[tuple, str]:
    """Parsuje arkusz Słownik → dict[(table_name, col_name)] = 'wartość=znaczenie | ...'"""
    collected: dict[tuple, list[str]] = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        table_name = _str(row[COL_SLOWNIK["table_name"]])
        col_name = _str(row[COL_SLOWNIK["col_name"]])
        value = _str(row[COL_SLOWNIK["value"]])
        meaning = _str(row[COL_SLOWNIK["meaning"]])
        if not (table_name and col_name):
            continue
        entry = f"{value}={meaning}" if meaning else value
        collected[(table_name, col_name)].append(entry)
    return {k: " | ".join(v) for k, v in collected.items()}


def parse_sample_values(ws) -> dict[tuple, str]:
    """Parsuje arkusz Przykładowe → dict[(table_name, col_name)] = 'v1, v2, ...'"""
    collected: dict[tuple, list[str]] = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        table_name = _str(row[COL_PRZYKLADOWE["table_name"]])
        col_name = _str(row[COL_PRZYKLADOWE["col_name"]])
        sample = _str(row[COL_PRZYKLADOWE["sample_value"]])
        if table_name and col_name and sample:
            collected[(table_name, col_name)].append(sample)
    return {k: ", ".join(v[:10]) for k, v in collected.items()}


def parse_gid_types(html_path: str) -> list[dict]:
    """Parsuje e_typy.html → lista typów GID z polami gid_type, internal_name, symbol, description."""

    class _Parser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.types: list[dict] = []
            self._row: list[str] = []
            self._cell = ""
            self._in_td = False

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self._row = []
            elif tag == "td":
                self._in_td = True
                self._cell = ""

        def handle_data(self, data):
            if self._in_td:
                self._cell += data

        def handle_endtag(self, tag):
            if tag == "td":
                self._in_td = False
                self._row.append(self._cell.strip())
            elif tag == "tr":
                row = self._row
                if len(row) == 5:
                    try:
                        gid_type = int(row[2])
                        self.types.append({
                            "gid_type": gid_type,
                            "internal_name": row[0].strip(),
                            "symbol": row[3].strip(),
                            "description": row[4].strip().rstrip(",").strip(),
                        })
                    except ValueError:
                        pass  # wiersz nagłówkowy lub niepełny

    path = Path(html_path)
    if not path.exists():
        return []
    parser = _Parser()
    parser.feed(path.read_text(encoding="utf-8"))
    return parser.types


def build_schema(conn) -> None:
    """Tworzy schemat SQLite (DROP IF EXISTS + CREATE). Idempotentne."""
    conn.executescript("""
        DROP TABLE IF EXISTS gid_types_fts;
        DROP TABLE IF EXISTS gid_types;
        DROP TABLE IF EXISTS columns_fts;
        DROP TABLE IF EXISTS relations;
        DROP TABLE IF EXISTS columns;
        DROP TABLE IF EXISTS tables;

        CREATE TABLE tables (
            table_id    INTEGER PRIMARY KEY,
            table_name  TEXT NOT NULL,
            table_label TEXT,
            prefix      TEXT,
            description TEXT
        );

        CREATE TABLE columns (
            col_id        INTEGER,
            table_name    TEXT NOT NULL,
            table_label   TEXT,
            col_name      TEXT NOT NULL,
            col_label     TEXT,
            role          TEXT,
            is_useful     TEXT,
            preferred     TEXT,
            data_type     TEXT,
            description   TEXT,
            value_dict    TEXT,
            sample_values TEXT
        );

        CREATE TABLE relations (
            source_table TEXT,
            source_col   TEXT,
            target_table TEXT,
            target_col   TEXT
        );

        CREATE VIRTUAL TABLE columns_fts USING fts5(
            table_name,
            table_label,
            col_name,
            col_label,
            description,
            is_useful     UNINDEXED,
            preferred     UNINDEXED,
            data_type     UNINDEXED,
            value_dict    UNINDEXED,
            sample_values UNINDEXED,
            tokenize = 'unicode61 remove_diacritics 2'
        );

        CREATE TABLE gid_types (
            gid_type      INTEGER,
            internal_name TEXT,
            symbol        TEXT,
            description   TEXT
        );

        CREATE VIRTUAL TABLE gid_types_fts USING fts5(
            gid_type_text,
            internal_name,
            symbol,
            description,
            tokenize = 'unicode61 remove_diacritics 2'
        );
    """)


def insert_data(
    conn,
    tables: list[dict],
    columns: list[dict],
    relations: list[dict],
    value_dicts: dict[tuple, str],
    sample_values: dict[tuple, str],
    gid_types: list[dict] | None = None,
) -> None:
    """Wstawia dane do tabel SQLite i buduje indeks FTS5."""
    conn.executemany(
        "INSERT INTO tables VALUES (:table_id, :table_name, :table_label, :prefix, :description)",
        tables,
    )

    for col in columns:
        key = (col["table_name"], col["col_name"])
        col["value_dict"] = value_dicts.get(key, "")
        col["sample_values"] = sample_values.get(key, "")

    conn.executemany(
        """INSERT INTO columns VALUES (
            :col_id, :table_name, :table_label, :col_name, :col_label, :role,
            :is_useful, :preferred, :data_type, :description,
            :value_dict, :sample_values
        )""",
        columns,
    )

    conn.executemany(
        "INSERT INTO relations VALUES (:source_table, :source_col, :target_table, :target_col)",
        relations,
    )

    conn.execute("""
        INSERT INTO columns_fts
            (table_name, table_label, col_name, col_label, description,
             is_useful, preferred, data_type, value_dict, sample_values)
        SELECT table_name, table_label, col_name, col_label, description,
               is_useful, preferred, data_type, value_dict, sample_values
        FROM columns
    """)

    for gt in (gid_types or []):
        conn.execute(
            "INSERT INTO gid_types VALUES (:gid_type, :internal_name, :symbol, :description)",
            gt,
        )
        conn.execute(
            "INSERT INTO gid_types_fts VALUES (?, ?, ?, ?)",
            (str(gt["gid_type"]), gt["internal_name"], gt["symbol"], gt["description"]),
        )

    conn.commit()


def _find_sheet(wb, fragment: str) -> str | None:
    """Szuka arkusza po fragmencie nazwy (obsługa problemów z kodowaniem)."""
    return next((s for s in wb.sheetnames if fragment in s), None)


def build_index(xlsm_path: str, db_path: str | None = None) -> None:
    """Główna funkcja: ładuje XLSM i buduje docs.db."""
    print(f"[build_index] Wczytywanie: {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=False, data_only=True)

    tables = parse_tables(wb["Tabele"])
    print(f"  Tabele:    {len(tables)}")

    columns = parse_columns(wb["Kolumny"])
    print(f"  Kolumny:   {len(columns)}")

    relations = parse_relations(wb["Relacje"])
    print(f"  Relacje:   {len(relations)}")

    # Nazwy arkuszy z polskimi znakami mogą być różnie kodowane — szukamy po fragmencie
    slownik_name = _find_sheet(wb, "ownik warto")
    przykladowe_name = _find_sheet(wb, "adowe warto")

    value_dicts = parse_value_dicts(wb[slownik_name]) if slownik_name else {}
    sample_values = parse_sample_values(wb[przykladowe_name]) if przykladowe_name else {}
    print(f"  Słowniki:  {len(value_dicts)} kolumn")
    print(f"  Przykłady: {len(sample_values)} kolumn")

    erp_docs = os.getenv("ERP_DOCS_PATH", "./erp_docs")
    html_path = Path(erp_docs) / "raw" / "Dokumnetacja bazy" / "e_typy.html"
    gid_types = parse_gid_types(str(html_path))
    print(f"  GIDTypy:   {len(gid_types)}")

    conn = get_db(db_path)
    build_schema(conn)
    insert_data(conn, tables, columns, relations, value_dicts, sample_values, gid_types)
    conn.close()
    print(f"[build_index] Gotowe: {db_path or 'domyslna sciezka'}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Buduje indeks SQLite FTS5 z pliku XLSM ERP.")
    parser.add_argument("--xlsm", default=XLSM_PATH, help="Ścieżka do pliku .xlsm")
    parser.add_argument("--db", default=None, help="Ścieżka do docs.db (opcjonalna)")
    args = parser.parse_args()

    if not args.xlsm:
        print("BŁĄD: brak ścieżki XLSM. Ustaw XLSM_PATH w .env lub podaj --xlsm.")
        sys.exit(1)

    build_index(args.xlsm, args.db)


if __name__ == "__main__":
    main()
