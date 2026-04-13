"""
offer_kerti_data.py — warstwa danych dla generatora oferty szkła Kerti.

Odpowiedzialność:
- Odczyt listy produktów z Excel (Akronim produktu, Nowość)
- Zapytania ERP: TwrKarty + TwrGrupy + Atrybuty + TwrJm (paleta)
- Rozwiązanie ścieżki zdjęcia
- Zwraca list[KertiGlassData] posortowane wg serii, potem kodu
"""

import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl

PHOTOS_DIR = r"C:\CEIM\Zdjęcia"

ATK_WYSOKOSC  = 12
ATK_SR_OTWORU = 56
ATK_SR_SPODU  = 16


@dataclass
class KertiGlassData:
    kod: str
    nazwa: str
    seria: str
    wysokosc_cm: Optional[float]
    sr_otworu_cm: Optional[float]
    sr_spodu_cm: Optional[float]
    ilosc_paleta: Optional[int]
    nowosc: bool
    zdjecie_path: Optional[str]


def _find_photo(kod: str) -> Optional[str]:
    for ext in (".png", ".jpg"):
        path = os.path.join(PHOTOS_DIR, kod + ext)
        if os.path.exists(path):
            return path
    return None


def _read_excel(excel_path: str) -> dict[str, dict]:
    """Zwraca {kod: {nowosc, seria}} z Excela.

    Kolumny:
      - 'Akronim produktu' (wymagana)
      - 'Nowość' TRUE/FALSE (opcjonalna)
      - 'Seria' (opcjonalna — nadpisuje grupę ERP)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_kod    = headers.index("Akronim produktu")
    col_nowosc = headers.index("Nowość") if "Nowość" in headers else None
    col_seria  = headers.index("Seria")  if "Seria"  in headers else None

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kod = row[col_kod]
        if not kod:
            continue
        nowosc = bool(row[col_nowosc]) if col_nowosc is not None else False
        seria  = str(row[col_seria]).strip() if col_seria is not None and row[col_seria] else None
        result[str(kod).strip()] = {"nowosc": nowosc, "seria": seria}

    if not result:
        raise ValueError("Plik Excel nie zawiera produktów.")
    return result


def _query_karty(kody_sql: str):
    from sql_query import run_query
    sql = f"""
        SELECT Twr_Kod, Twr_Nazwa, Twr_GIDNumer
        FROM CDN.TwrKarty
        WHERE Twr_Kod IN ({kody_sql})
    """
    r = run_query(sql, inject_top=None)
    if not r["ok"]:
        raise RuntimeError(f"ERP TwrKarty: {r['error']['message']}")
    return {row[0]: {"nazwa": row[1], "gid": row[2]} for row in r["data"]["rows"]}


def _query_grupy(gid_list: str) -> dict:
    from sql_query import run_query
    sql = f"""
        SELECT br.TwG_GIDNumer, MAX(RTRIM(grp.TwG_Nazwa))
        FROM CDN.TwrGrupy br
        LEFT JOIN CDN.TwrGrupy grp
            ON grp.TwG_GIDTyp = -16 AND grp.TwG_GIDNumer = br.TwG_GrONumer
        WHERE br.TwG_GIDTyp = 16 AND br.TwG_GIDNumer IN ({gid_list})
          AND RTRIM(grp.TwG_Nazwa) <> 'Surowce'
        GROUP BY br.TwG_GIDNumer
    """
    r = run_query(sql, inject_top=None)
    if not r["ok"]:
        return {}
    return {row[0]: row[1] or "" for row in r["data"]["rows"]}


def _query_atrybuty(gid_list: str) -> dict:
    from sql_query import run_query
    sql = f"""
        SELECT Atr_ObiNumer, Atr_AtkId, Atr_Wartosc
        FROM CDN.Atrybuty
        WHERE Atr_ObiNumer IN ({gid_list})
          AND Atr_AtkId IN ({ATK_WYSOKOSC}, {ATK_SR_OTWORU}, {ATK_SR_SPODU})
    """
    r = run_query(sql, inject_top=None)
    atr: dict = {}
    if r["ok"]:
        for row in r["data"]["rows"]:
            gid, atk_id, val = row[0], row[1], row[2]
            atr.setdefault(gid, {})[atk_id] = float(val) if val else None
    return atr


def _query_palety(gid_list: str) -> dict:
    from sql_query import run_query
    sql = f"""
        SELECT TwJ_TwrNumer, TwJ_PrzeliczL
        FROM CDN.TwrJm
        WHERE TwJ_TwrNumer IN ({gid_list})
          AND TwJ_JmZ = 'paleta'
    """
    r = run_query(sql, inject_top=None)
    if not r["ok"]:
        return {}
    return {row[0]: int(row[1]) if row[1] else None for row in r["data"]["rows"]}


def load_kerti_products(excel_path: str) -> list[KertiGlassData]:
    """
    Ładuje produkty Kerti z Excela (Akronim + Nowość) + ERP.

    Returns:
        list[KertiGlassData] posortowane wg seria, kod
    """
    sys.path.insert(0, str(Path(__file__).parent))

    excel = _read_excel(excel_path)
    kody_sql = ", ".join(f"'{k}'" for k in excel)

    karty = _query_karty(kody_sql)
    gid_list = ", ".join(str(v["gid"]) for v in karty.values())

    grupy   = _query_grupy(gid_list)
    atryb   = _query_atrybuty(gid_list)
    palety  = _query_palety(gid_list)

    products = []
    for kod, meta in excel.items():
        karta = karty.get(kod)
        if not karta:
            continue
        gid = karta["gid"]
        atr = atryb.get(gid, {})
        seria = meta["seria"] or grupy.get(gid, "")
        products.append(KertiGlassData(
            kod=kod,
            nazwa=karta["nazwa"],
            seria=seria,
            wysokosc_cm=atr.get(ATK_WYSOKOSC),
            sr_otworu_cm=atr.get(ATK_SR_OTWORU),
            sr_spodu_cm=atr.get(ATK_SR_SPODU),
            ilosc_paleta=palety.get(gid),
            nowosc=meta["nowosc"],
            zdjecie_path=_find_photo(kod),
        ))

    return products
