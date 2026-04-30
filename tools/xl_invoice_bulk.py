"""Import wsadowy faktur zakupowych FZ z katalogu XML do Comarch XL.

Skanuje katalog, parsuje pliki KSeF FA(3), importuje przez XL API.
Jeden błędny plik nie zatrzymuje pozostałych.

Użycie:
  python tools/xl_invoice_bulk.py --dir katalog/z/xml
  python tools/xl_invoice_bulk.py --dir katalog/z/xml --report raport.xlsx
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.xl_invoice_parser import parse_ksef_xml
from tools.xl_invoice_set import set_invoice

_STATUS_FILLS = {
    "OK":        PatternFill("solid", fgColor="C6EFCE"),
    "BŁĄD":      PatternFill("solid", fgColor="FFC7CE"),
    "POMINIĘTO": PatternFill("solid", fgColor="FFEB9C"),
}


def _write_report(path: Path, rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Import FZ"

    headers = ["Plik", "Nr faktury", "Kontrahent (NIP)", "Kwota brutto", "Status", "Komunikat"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for row_idx, r in enumerate(rows, 2):
        ws.cell(row_idx, 1, r["plik"])
        ws.cell(row_idx, 2, r.get("nr_obcy", ""))
        ws.cell(row_idx, 3, r.get("nip", ""))
        ws.cell(row_idx, 4, r.get("kwota_brutto", ""))
        ws.cell(row_idx, 5, r["status"])
        ws.cell(row_idx, 6, r.get("komunikat", ""))
        fill = _STATUS_FILLS.get(r["status"])
        if fill:
            for col in range(1, 7):
                ws.cell(row_idx, col).fill = fill

    for col, width in zip(range(1, 7), [28, 22, 18, 14, 12, 50]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def bulk_import(dir_path: Path, report: Path | None = None) -> dict:
    xml_files = sorted(dir_path.glob("*.xml"))

    rows = []
    inserted = skipped = failed = 0

    for xml_path in xml_files:
        parse_result = parse_ksef_xml(xml_path)

        if not parse_result["ok"]:
            failed += 1
            rows.append({
                "plik": xml_path.name,
                "status": "BŁĄD",
                "komunikat": parse_result["error"],
            })
            continue

        inv = parse_result["data"]
        set_result = set_invoice(inv)

        if not set_result["ok"]:
            failed += 1
            rows.append({
                "plik": xml_path.name,
                "nr_obcy": inv.nr_obcy,
                "nip": inv.nip_sprzedawcy,
                "kwota_brutto": str(inv.suma_brutto),
                "status": "BŁĄD",
                "komunikat": set_result["error"]["message"],
            })
            continue

        action = set_result["data"]["action"]
        if action == "skipped":
            skipped += 1
            status = "POMINIĘTO"
            komunikat = "Faktura już istnieje w ERP"
        else:
            inserted += 1
            status = "OK"
            komunikat = f"doc_id={set_result['data']['doc_id']}"

        rows.append({
            "plik": xml_path.name,
            "nr_obcy": inv.nr_obcy,
            "nip": inv.nip_sprzedawcy,
            "kwota_brutto": str(inv.suma_brutto),
            "status": status,
            "komunikat": komunikat,
        })

    if report:
        _write_report(report, rows)

    return {
        "ok": True,
        "data": {
            "total": len(xml_files),
            "inserted": inserted,
            "skipped": skipped,
            "failed": failed,
            "report": str(report) if report else None,
            "results": rows,
        },
        "error": None,
    }


def main() -> None:
    today = date.today().strftime("%Y%m%d")
    default_report = Path(f"documents/human/reports/xl_invoice_bulk_{today}.xlsx")

    parser = argparse.ArgumentParser(description="Wsadowy import FZ z katalogu XML do Comarch XL")
    parser.add_argument("--dir", required=True, type=Path, help="Katalog z plikami XML")
    parser.add_argument("--report", type=Path, default=default_report,
                        help="Ścieżka raportu Excel (domyślnie: documents/human/reports/)")
    args = parser.parse_args()

    result = bulk_import(args.dir, args.report)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
