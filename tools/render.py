"""Universal renderer for mrowisko.db views.

Usage:
    python tools/render.py backlog --format xlsx
    python tools/render.py backlog --format md --status planned --area Bot Dev Arch
    python tools/render.py backlog --format json
    python tools/render.py suggestions --format md --status open --author erp_specialist
    python tools/render.py inbox --role developer --format md
    python tools/render.py session-log --role developer --format json --limit 50
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.lib.agent_bus import AgentBus

DB_PATH = "mrowisko.db"

# --- View definitions ---

VIEWS = {
    "backlog": {
        "title": "Backlog",
        "columns": ["id", "title", "area", "value", "effort", "status", "created_at"],
    },
    "suggestions": {
        "title": "Suggestions",
        "columns": ["id", "type", "author", "title", "status", "created_at"],
    },
    "inbox": {
        "title": "Inbox",
        "columns": ["id", "sender", "type", "content", "status", "created_at"],
    },
    "session-log": {
        "title": "Session Log",
        "columns": ["id", "role", "content", "session_id", "created_at"],
    },
    "messages": {
        "title": "Messages",
        "columns": ["id", "sender", "recipient", "type", "content", "status", "created_at", "read_at"],
    },
}

SESSION_TRACE_TOOL_COLUMNS = ["id", "tool_name", "input_summary", "is_error", "tokens_out", "timestamp"]
SESSION_TRACE_TOKEN_COLUMNS = ["turn_index", "input_tokens", "output_tokens", "cache_read_tokens", "cache_create_tokens", "duration_ms", "timestamp"]

VALUE_COLORS = {"wysoka": "C6EFCE", "srednia": "FFEB9C", "niska": "FFC7CE"}
STATUS_COLORS = {
    "planned": "FFFFFF", "in_progress": "FFEB9C", "done": "C6EFCE", "cancelled": "F2F2F2", "deferred": "DCE6F1",
    "open": "FFFFFF", "in_backlog": "FFEB9C", "implemented": "C6EFCE", "rejected": "FFC7CE",
    "unread": "FFFFFF", "read": "F2F2F2", "archived": "F2F2F2",
}

# --- Data fetch ---

def fetch(view: str, bus: AgentBus, args: argparse.Namespace) -> list[dict]:
    if view == "backlog":
        status = getattr(args, "status", "planned")
        items = bus.get_backlog(status=None if status == "all" else status)
        area = getattr(args, "area", None)
        if area:
            items = [i for i in items if i.get("area") in area]
        return items
    if view == "suggestions":
        return bus.get_suggestions(
            status=getattr(args, "status", None),
            author=getattr(args, "author", None),
            type=getattr(args, "type", None),
        )
    if view == "inbox":
        return bus.get_inbox(
            role=args.role,
            status=getattr(args, "status", "unread"),
        )
    if view == "session-log":
        return bus.get_session_log(
            role=args.role,
            limit=getattr(args, "limit", 20),
        )
    if view == "messages":
        return bus.get_messages(
            sender=getattr(args, "sender", None),
            recipient=getattr(args, "recipient", None),
            status=getattr(args, "status", None),
            limit=getattr(args, "limit", 200),
        )
    return []

# --- Renderers ---

def render_json(data: list[dict], title: str, output: Path) -> None:
    payload = {"title": title, "data": data, "count": len(data)}
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def render_backlog_md(data: list[dict], title: str, output: Path) -> None:
    """Backlog md: summary tables grouped by effort/value, then detailed sections."""
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")
    lines = [f"# {title} — {today}\n", f"*{len(data)} pozycji*\n", "---\n"]

    def table_rows(items):
        rows = ["| id | tytuł | obszar | wartość | effort |",
                "|----|-------|--------|---------|--------|"]
        for r in items:
            rows.append(f"| {r.get('id','')} | {r.get('title','')} | {r.get('area','')} | {r.get('value','')} | {r.get('effort','')} |")
        return rows

    groups = [
        ("Szybkie strzały (wysoka wartość, mała praca)",
         [r for r in data if r.get("value") == "wysoka" and r.get("effort") == "mala"]),
        ("Wysoka wartość, średnia praca",
         [r for r in data if r.get("value") == "wysoka" and r.get("effort") == "srednia"]),
        ("Wysoka wartość, duża praca",
         [r for r in data if r.get("value") == "wysoka" and r.get("effort") == "duza"]),
        ("Średnia wartość, mała praca",
         [r for r in data if r.get("value") == "srednia" and r.get("effort") == "mala"]),
        ("Średnia wartość, średnia/duża praca",
         [r for r in data if r.get("value") == "srednia" and r.get("effort") in ("srednia", "duza")]),
        ("Pozostałe",
         [r for r in data if r.get("value") not in ("wysoka", "srednia") or r.get("effort") not in ("mala", "srednia", "duza")]),
    ]

    for heading, items in groups:
        if not items:
            continue
        lines.append(f"## {heading}\n")
        lines.extend(table_rows(items))
        lines.append("")

    lines.append("---\n")
    lines.append("## Szczegóły\n")

    for row in data:
        item_id = row.get("id")
        item_title = row.get("title", "")
        lines.append(f"### [{item_id}] {item_title}")
        meta = []
        for col in ["area", "value", "effort", "status", "created_at"]:
            val = row.get(col)
            if val:
                if col == "created_at":
                    val = str(val)[:10]
                meta.append(f"**{col}:** {val}")
        if meta:
            lines.append("  ".join(meta))
        content = (row.get("content") or "").strip()
        if content:
            lines.append(f"\n{content}")
        lines.append("")

    output.write_text("\n".join(lines), encoding="utf-8")


def render_suggestions_md(data: list[dict], title: str, output: Path) -> None:
    """Suggestions md: table at top (grouped by type), full content sections below."""
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")
    lines = [f"# {title} — {today}\n", f"*{len(data)} sugestii*\n", "---\n"]

    TYPE_ORDER = ["rule", "tool", "discovery", "observation"]
    TYPE_LABELS = {
        "rule": "Zasady (rule)",
        "tool": "Narzędzia (tool)",
        "discovery": "Odkrycia (discovery)",
        "observation": "Obserwacje (observation)",
    }

    def table_rows(items):
        rows = [
            "| id | autor | tytuł | status | data |",
            "|----|-------|-------|--------|------|",
        ]
        for r in items:
            date_val = str(r.get("created_at", ""))[:10]
            rows.append(
                f"| {r.get('id','')} | {r.get('author','')} "
                f"| {r.get('title','')} | {r.get('status','')} | {date_val} |"
            )
        return rows

    grouped = {t: [r for r in data if r.get("type") == t] for t in TYPE_ORDER}
    other = [r for r in data if r.get("type") not in TYPE_ORDER]
    if other:
        grouped["observation"] = grouped.get("observation", []) + other

    for type_key in TYPE_ORDER:
        items = grouped.get(type_key, [])
        if not items:
            continue
        lines.append(f"## {TYPE_LABELS[type_key]}\n")
        lines.extend(table_rows(items))
        lines.append("")

    lines.append("---\n")
    lines.append("## Treści\n")

    for type_key in TYPE_ORDER:
        items = grouped.get(type_key, [])
        if not items:
            continue
        lines.append(f"### {TYPE_LABELS[type_key]}\n")
        for row in items:
            sid = row.get("id")
            stitle = row.get("title") or ""
            author = row.get("author", "")
            status = row.get("status", "")
            date_val = str(row.get("created_at", ""))[:10]
            lines.append(f"#### [{sid}] {stitle}")
            lines.append(f"**autor:** {author}  **status:** {status}  **data:** {date_val}")
            content = (row.get("content") or "").strip()
            if content:
                lines.append(f"\n{content}")
            lines.append("")

    output.write_text("\n".join(lines), encoding="utf-8")


def render_md(data: list[dict], columns: list[str], title: str, output: Path) -> None:
    """Human-readable md: each item as a section with metadata + content (if present)."""
    META_COLS = ["id", "area", "value", "effort", "status", "created_at", "sender", "author", "role"]
    CONTENT_COLS = ["content", "title"]

    lines = [f"# {title} — {len(data)} pozycji\n"]
    for row in data:
        item_title = str(row.get("title") or row.get("id") or "")
        item_id = row.get("id")
        heading = f"## [{item_id}] {item_title}" if item_id and item_title and item_title != str(item_id) else f"## {item_title or item_id}"
        lines.append(heading)

        meta = []
        for col in META_COLS:
            if col in columns and row.get(col) is not None:
                val = str(row[col])
                if col == "created_at":
                    val = val[:10]
                meta.append(f"**{col}:** {val}")
        if meta:
            lines.append("  ".join(meta))

        content = row.get("content") or ""
        if content and "content" not in columns:
            lines.append(f"\n{content.strip()}")
        elif content:
            lines.append(f"\n{content.strip()}")

        lines.append("\n---\n")

    output.write_text("\n".join(lines), encoding="utf-8")


def render_xlsx(data: list[dict], columns: list[str], title: str, output: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            val = item.get(col)
            if col == "created_at" and val:
                val = val[:10]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col == "value" and val in VALUE_COLORS:
                cell.fill = PatternFill("solid", fgColor=VALUE_COLORS[val])
            if col == "status" and val in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[val])

    for col_idx, col in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 60 if col in ("title", "content") else 14

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(data)+1}"
    wb.save(str(output))


def render_session_trace_xlsx(trace: dict, output: Path) -> None:
    """Render session trace as multi-sheet XLSX: Summary, ToolCalls, TokenUsage."""
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill("solid", fgColor="FFC7CE")

    def write_header(ws, columns):
        for col, name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=name.upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    session = trace["session"]
    tool_calls = trace["tool_calls"]
    token_usage = trace["token_usage"]
    total_input = sum(t.get("input_tokens") or 0 for t in token_usage)
    total_output = sum(t.get("output_tokens") or 0 for t in token_usage)
    total_cache = sum(t.get("cache_read_tokens") or 0 for t in token_usage)
    total_duration = sum(t.get("duration_ms") or 0 for t in token_usage)
    error_count = sum(1 for tc in tool_calls if tc.get("is_error"))
    summary_rows = [
        ("session_id", session.get("id")),
        ("claude_session_id", session.get("claude_session_id")),
        ("role", session.get("role")),
        ("started_at", session.get("started_at")),
        ("ended_at", session.get("ended_at")),
        ("transcript_path", session.get("transcript_path")),
        ("tool_calls_total", len(tool_calls)),
        ("tool_calls_errors", error_count),
        ("turns_total", len(token_usage)),
        ("input_tokens_total", total_input),
        ("output_tokens_total", total_output),
        ("cache_read_tokens_total", total_cache),
        ("duration_ms_total", total_duration),
    ]
    write_header(ws_sum, ["metric", "value"])
    for row_idx, (metric, value) in enumerate(summary_rows, 2):
        ws_sum.cell(row=row_idx, column=1, value=metric)
        ws_sum.cell(row=row_idx, column=2, value=value)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 60

    # Sheet 2: ToolCalls
    ws_tc = wb.create_sheet("ToolCalls")
    write_header(ws_tc, SESSION_TRACE_TOOL_COLUMNS)
    for row_idx, tc in enumerate(tool_calls, 2):
        for col_idx, col in enumerate(SESSION_TRACE_TOOL_COLUMNS, 1):
            cell = ws_tc.cell(row=row_idx, column=col_idx, value=tc.get(col))
            if col == "is_error" and tc.get("is_error"):
                cell.fill = error_fill
    ws_tc.column_dimensions["B"].width = 20
    ws_tc.column_dimensions["C"].width = 60
    ws_tc.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SESSION_TRACE_TOOL_COLUMNS))}{len(tool_calls)+1}"

    # Sheet 3: TokenUsage
    ws_tu = wb.create_sheet("TokenUsage")
    write_header(ws_tu, SESSION_TRACE_TOKEN_COLUMNS)
    for row_idx, tu in enumerate(token_usage, 2):
        for col_idx, col in enumerate(SESSION_TRACE_TOKEN_COLUMNS, 1):
            ws_tu.cell(row=row_idx, column=col_idx, value=tu.get(col))
    ws_tu.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SESSION_TRACE_TOKEN_COLUMNS))}{len(token_usage)+1}"

    wb.save(str(output))

# --- CLI ---

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Render mrowisko.db views to md/xlsx/json")
    parser.add_argument("--db", default=DB_PATH)
    sub = parser.add_subparsers(dest="view", required=True)

    for view in ["backlog", "suggestions"]:
        p = sub.add_parser(view)
        p.add_argument("--format", required=True, choices=["md", "xlsx", "json"])
        p.add_argument("--output", default=None)
        if view == "backlog":
            p.add_argument("--status", default="planned",
                           help="Filter by status (default: planned). Use 'all' to show everything.")
            p.add_argument("--area", nargs="+", default=None)
        else:
            p.add_argument("--status", default=None)
            p.add_argument("--author", default=None)
            p.add_argument("--type", default=None,
                           choices=["rule", "tool", "discovery", "observation"])

    p = sub.add_parser("messages")
    p.add_argument("--format", required=True, choices=["md", "xlsx", "json"])
    p.add_argument("--output", default=None)
    p.add_argument("--sender", default=None)
    p.add_argument("--recipient", default=None)
    p.add_argument("--status", default=None)
    p.add_argument("--limit", type=int, default=200)

    for view in ["inbox", "session-log"]:
        p = sub.add_parser(view)
        p.add_argument("--format", required=True, choices=["md", "xlsx", "json"])
        p.add_argument("--output", default=None)
        p.add_argument("--role", required=True)
        p.add_argument("--status", default="unread")
        if view == "session-log":
            p.add_argument("--limit", type=int, default=20)

    p = sub.add_parser("session-trace")
    p.add_argument("--session", required=True, help="Our session ID (12 hex)")
    p.add_argument("--output", default=None)

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    bus = AgentBus(db_path=args.db)
    view = args.view

    if view == "session-trace":
        trace = bus.get_session_trace(args.session)
        if not trace:
            print(f"Sesja '{args.session}' nie znaleziona w DB.", file=sys.stderr)
            sys.exit(1)
        output = Path(args.output) if args.output else Path("views") / f"session_trace_{args.session}.xlsx"
        output.parent.mkdir(parents=True, exist_ok=True)
        render_session_trace_xlsx(trace, output)
        tc_count = len(trace["tool_calls"])
        tu_count = len(trace["token_usage"])
        print(f"{output} ({tc_count} tool_calls, {tu_count} turns)")
        return

    cfg = VIEWS[view]
    data = fetch(view, bus, args)

    ext = args.format
    output = Path(args.output) if args.output else Path("views") / f"{view}.{ext}"
    output.parent.mkdir(parents=True, exist_ok=True)

    if ext == "json":
        render_json(data, cfg["title"], output)
    elif ext == "md":
        if view == "backlog":
            render_backlog_md(data, cfg["title"], output)
        elif view == "suggestions":
            render_suggestions_md(data, cfg["title"], output)
        else:
            render_md(data, cfg["columns"], cfg["title"], output)
    elif ext == "xlsx":
        render_xlsx(data, cfg["columns"], cfg["title"], output)

    print(f"{output} ({len(data)} pozycji)")


if __name__ == "__main__":
    main()
