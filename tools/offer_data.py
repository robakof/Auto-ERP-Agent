"""
offer_data.py — warstwa danych dla generatora ofert katalogowych.

Odpowiedzialność:
- Odczyt listy produktów z Excel (Akronim, EAN, Nazwa, Cena sprzedaży)
- Zapytania ERP: TwrKarty + TwrJm (opak.) + Atrybuty (wysokość, AtK_ID=12)
- Parsowanie czasu palenia z nazwy produktu
- Rozwiązanie ścieżki zdjęcia (.jpg → .png → None)
- Zwraca list[ProductData]
"""

import re
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


PHOTOS_DIRS = [
    r"D:\UdzialySieciowe\ZDJĘCIA\ZDJĘCIA PRODUKTÓW\jpg do systemu",
    r"D:\UdzialySieciowe\ZDJĘCIA\ZDJĘCIA PRODUKTÓW",
]

BURNING_TIME_HOURS = {
    2.0: 48,
    3.0: 72,
    4.0: 96,
    5.0: 120,
    5.5: 132,
    6.0: 144,
    6.5: 156,
}

CURRENCY = {
    "pl": "zł",
    "en": "EUR",
    "ro": "EUR",
}


@dataclass
class ProductData:
    kod: str
    nazwa: str
    wysokosc: str
    wysokosc_val: float  # wartość numeryczna [cm] — do skalowania zdjęć
    czas_palenia: str
    ilosc_opak: str
    cena: str
    zdjecie_path: Optional[str]


def _parse_burning_time(nazwa: str, lang: str) -> str:
    """Wyciąga czas palenia z nazwy produktu i zwraca sformatowany string."""
    match = re.search(r"(\d+(?:[,\.]\d+)?)\s*(?:szeroki)?$", nazwa.strip())
    if not match:
        return ""
    days_str = match.group(1).replace(",", ".")
    days = float(days_str)
    hours = BURNING_TIME_HOURS.get(days, int(days * 24))

    days_display = match.group(1)  # oryginalna forma (np. "5,5")

    return f"~{hours} h"


def _find_photo(kod: str) -> Optional[str]:
    """Szuka pliku zdjęcia: najpierw w podkatalogu, potem w katalogu wyżej. Formaty: .jpg, .png."""
    for directory in PHOTOS_DIRS:
        for ext in (".jpg", ".png"):
            path = os.path.join(directory, kod + ext)
            if os.path.exists(path):
                return path
    return None


def _format_price(price_float: float, lang: str) -> str:
    """Formatuje cenę: 2 miejsca dziesiętne, separator przecinek + waluta."""
    currency = CURRENCY.get(lang, "zł")
    price_str = f"{price_float:.2f}".replace(".", ",")
    return f"{price_str} {currency}"


def load_products(excel_path: str, lang: str = "pl") -> list[ProductData]:
    """
    Ładuje produkty z Excela, uzupełnia danymi z ERP.

    Args:
        excel_path: ścieżka do pliku OFerta katalogowa.xlsx
        lang: język (pl/en/ro) — wpływa na format czasu palenia i walutę

    Returns:
        list[ProductData] — lista produktów gotowych do PDF
    """
    # --- 1. Odczyt Excel ---
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_kod = headers.index("Akronim produktu")
    col_cena = headers.index("Cena sprzedaży")

    excel_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kod = row[col_kod]
        cena = row[col_cena]
        if kod:
            excel_rows[kod] = float(cena) if cena else 0.0

    if not excel_rows:
        raise ValueError("Plik Excel nie zawiera produktów.")

    kody = list(excel_rows.keys())
    kody_sql = ", ".join(f"'{k}'" for k in kody)

    # --- 2. Zapytania ERP ---
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from sql_query import run_query

    # Karta towaru
    sql_karty = f"""
        SELECT Twr_Kod, Twr_Nazwa, Twr_GIDNumer
        FROM CDN.TwrKarty
        WHERE Twr_Kod IN ({kody_sql})
    """
    result_karty = run_query(sql_karty, inject_top=None)
    if not result_karty["ok"]:
        raise RuntimeError(f"ERP TwrKarty error: {result_karty['error']['message']}")

    karty = {row[0]: {"nazwa": row[1], "gid": row[2]} for row in result_karty["data"]["rows"]}

    # Jednostki pomocnicze (opakowanie)
    gid_list = ", ".join(str(v["gid"]) for v in karty.values())
    sql_jm = f"""
        SELECT j.TwJ_TwrNumer, j.TwJ_PrzeliczL
        FROM CDN.TwrJm j
        WHERE j.TwJ_TwrNumer IN ({gid_list})
          AND j.TwJ_JmZ = 'opak.'
    """
    result_jm = run_query(sql_jm, inject_top=None)
    opak = {}
    if result_jm["ok"]:
        for row in result_jm["data"]["rows"]:
            opak[row[0]] = row[1]

    # Atrybuty — wysokość netto (AtK_ID = 12)
    sql_atr = f"""
        SELECT a.Atr_ObiNumer, a.Atr_Wartosc
        FROM CDN.Atrybuty a
        WHERE a.Atr_ObiNumer IN ({gid_list})
          AND a.Atr_AtkId = 12
    """
    result_atr = run_query(sql_atr, inject_top=None)
    wysokosci = {}
    if result_atr["ok"]:
        for row in result_atr["data"]["rows"]:
            wysokosci[row[0]] = row[1]

    # --- 3. Składanie ProductData ---
    products = []
    for kod in kody:
        if kod not in karty:
            continue
        karta = karty[kod]
        gid = karta["gid"]
        nazwa = karta["nazwa"]

        wys_val = wysokosci.get(gid, "")
        wysokosc = f"{wys_val} cm" if wys_val else "—"
        try:
            wysokosc_num = float(wys_val) if wys_val else 0.0
        except (ValueError, TypeError):
            wysokosc_num = 0.0

        ilosc_val = opak.get(gid, "")
        ilosc = f"{ilosc_val} szt." if ilosc_val else "—"

        czas = _parse_burning_time(nazwa, lang)
        cena = _format_price(excel_rows[kod], lang)
        zdjecie = _find_photo(kod)

        products.append(ProductData(
            kod=kod,
            nazwa=nazwa,
            wysokosc=wysokosc,
            wysokosc_val=wysokosc_num,
            czas_palenia=czas,
            ilosc_opak=ilosc,
            cena=cena,
            zdjecie_path=zdjecie,
        ))

    return products
