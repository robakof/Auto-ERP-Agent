"""pp_bom.py — Odczyt BOM i wydajności z arkusza 'Wycena Zniczy' pliku Wycena PQ.xlsm.

Kolumny (1-indexed):
  B=2: czni_kod        (Akronim produktu)
  F=6: grupa           (Podstawowa / Pakowanie / Robocizna / Inne koszty / ...)
  H=8: surowiec_kod    (Akronim surowca)
  I=9: surowiec_nazwa
  J=10: mianownik      (divisor: zapotrzebowanie = ilosc_czni / mianownik)
                       dla wierszy Roboczogodzina Otorowo: mianownik = szt/h (wydajność)

Nagłówki w wierszu 3, dane od wiersza 4.
"""
import warnings
from dataclasses import dataclass
from pathlib import Path

import openpyxl

SHEET_NAME = "Wycena Zniczy"


@dataclass
class EfficiencyEntry:
    czni_kod: str
    units_per_hour: float   # col J (mianownik) = szt/h dla 1 osoby


@dataclass
class BomEntry:
    czni_kod: str
    surowiec_kod: str
    surowiec_nazwa: str
    mianownik: float


def load_bom(xlsm_path: Path) -> dict[str, list[BomEntry]]:
    """Wczytuje BOM z arkusza 'Wycena Zniczy'.

    Klucz: czni_kod → lista BomEntry (jeden wpis na surowiec).
    Rzuca FileNotFoundError jeśli plik nie istnieje.
    Rzuca ValueError jeśli brak arkusza 'Wycena Zniczy'.
    """
    xlsm_path = Path(xlsm_path)
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Plik BOM nie istnieje: {xlsm_path}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(  # noqa: F821
            xlsm_path, read_only=True, data_only=True, keep_vba=False
        )

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Brak arkusza '{SHEET_NAME}' w pliku: {xlsm_path}")

    ws = wb[SHEET_NAME]
    result: dict[str, list[BomEntry]] = {}
    skipped = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        czni_kod = row[1]   # B
        grupa    = row[5]   # F
        sur_kod  = row[7]   # H
        sur_nazwa = row[8]  # I
        mianownik = row[9]  # J

        if not czni_kod:
            continue
        if not sur_kod:
            skipped += 1
            continue
        if isinstance(grupa, str) and grupa.startswith("Koszt"):
            skipped += 1
            continue
        if not _is_valid_number(mianownik):
            warnings.warn(
                f"[BOM] Mianownik '{mianownik}' dla {czni_kod}/{sur_kod} — pominięto",
                stacklevel=2,
            )
            skipped += 1
            continue

        entry = BomEntry(
            czni_kod=str(czni_kod),
            surowiec_kod=str(sur_kod),
            surowiec_nazwa=str(sur_nazwa) if sur_nazwa else "",
            mianownik=float(mianownik),
        )
        result.setdefault(str(czni_kod), []).append(entry)

    wb.close()
    return result


def load_efficiency(xlsm_path: Path) -> dict[str, EfficiencyEntry]:
    """Wyciąga wydajność (szt/h) z wierszy grupy 'Robocizna'.

    Na CZNI mogą przypadać 2 wiersze Robocizna: J=1 (stawka kosztu)
    i J=właściwa wydajność (np. 30, 50). Bierzemy max(J) per CZNI.

    Klucz: czni_kod. Wartość: EfficiencyEntry.
    Rzuca FileNotFoundError jeśli plik nie istnieje.
    Rzuca ValueError jeśli brak arkusza 'Wycena Zniczy'.
    """
    xlsm_path = Path(xlsm_path)
    if not xlsm_path.exists():
        raise FileNotFoundError(f"Plik BOM nie istnieje: {xlsm_path}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(
            xlsm_path, read_only=True, data_only=True, keep_vba=False
        )

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Brak arkusza '{SHEET_NAME}' w pliku: {xlsm_path}")

    ws = wb[SHEET_NAME]
    # Zbieramy max(J) per CZNI — eliminuje wiersz J=1 (stawka kosztu)
    best: dict[str, float] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        czni_kod = row[1]   # B
        grupa    = row[5]   # F
        mianownik = row[9]  # J

        if not czni_kod:
            continue
        if grupa != "Robocizna":
            continue
        if not _is_valid_number(mianownik):
            warnings.warn(
                f"[BOM] units_per_hour '{mianownik}' dla {czni_kod} — pominięto",
                stacklevel=2,
            )
            continue

        kod = str(czni_kod)
        val = float(mianownik)
        if val > best.get(kod, 0):
            best[kod] = val

    wb.close()
    return {
        kod: EfficiencyEntry(czni_kod=kod, units_per_hour=val)
        for kod, val in best.items()
    }


def _is_valid_number(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val))
        return True
    except (ValueError, TypeError):
        return False
