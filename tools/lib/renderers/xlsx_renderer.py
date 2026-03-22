"""Excel (XLSX) renderers for mrowisko.db views."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .base import VALUE_COLORS, STATUS_COLORS

# Column definitions for session trace sheets
SESSION_TRACE_TOOL_COLUMNS = ["id", "tool_name", "input_summary", "is_error", "tokens_out", "timestamp"]
SESSION_TRACE_TOKEN_COLUMNS = ["turn_index", "input_tokens", "output_tokens", "cache_read_tokens", "cache_create_tokens", "duration_ms", "timestamp"]


def render_xlsx(data: list[dict], columns: list[str], title: str, output: Path) -> None:
    """Render data as Excel workbook with formatted headers and colored cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            val = item.get(col)
            if col == "created_at" and val:
                val = val[:10]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col == "value" and val in VALUE_COLORS:
                cell.fill = PatternFill("solid", fgColor=VALUE_COLORS[val])
            if col == "status" and val in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[val])

    for col_idx, col in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 60 if col in ("title", "content") else 14

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(data)+1}"
    wb.save(str(output))


def render_session_trace_xlsx(trace: dict, output: Path) -> None:
    """Render session trace as multi-sheet XLSX: Summary, ToolCalls, TokenUsage."""
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill("solid", fgColor="FFC7CE")

    def write_header(ws, columns):
        for col, name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=name.upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    session = trace["session"]
    tool_calls = trace["tool_calls"]
    token_usage = trace["token_usage"]
    total_input = sum(t.get("input_tokens") or 0 for t in token_usage)
    total_output = sum(t.get("output_tokens") or 0 for t in token_usage)
    total_cache = sum(t.get("cache_read_tokens") or 0 for t in token_usage)
    total_duration = sum(t.get("duration_ms") or 0 for t in token_usage)
    error_count = sum(1 for tc in tool_calls if tc.get("is_error"))
    summary_rows = [
        ("session_id", session.get("id")),
        ("claude_session_id", session.get("claude_session_id")),
        ("role", session.get("role")),
        ("started_at", session.get("started_at")),
        ("ended_at", session.get("ended_at")),
        ("transcript_path", session.get("transcript_path")),
        ("tool_calls_total", len(tool_calls)),
        ("tool_calls_errors", error_count),
        ("turns_total", len(token_usage)),
        ("input_tokens_total", total_input),
        ("output_tokens_total", total_output),
        ("cache_read_tokens_total", total_cache),
        ("duration_ms_total", total_duration),
    ]
    write_header(ws_sum, ["metric", "value"])
    for row_idx, (metric, value) in enumerate(summary_rows, 2):
        ws_sum.cell(row=row_idx, column=1, value=metric)
        ws_sum.cell(row=row_idx, column=2, value=value)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 60

    # Sheet 2: ToolCalls
    ws_tc = wb.create_sheet("ToolCalls")
    write_header(ws_tc, SESSION_TRACE_TOOL_COLUMNS)
    for row_idx, tc in enumerate(tool_calls, 2):
        for col_idx, col in enumerate(SESSION_TRACE_TOOL_COLUMNS, 1):
            cell = ws_tc.cell(row=row_idx, column=col_idx, value=tc.get(col))
            if col == "is_error" and tc.get("is_error"):
                cell.fill = error_fill
    ws_tc.column_dimensions["B"].width = 20
    ws_tc.column_dimensions["C"].width = 60
    ws_tc.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SESSION_TRACE_TOOL_COLUMNS))}{len(tool_calls)+1}"

    # Sheet 3: TokenUsage
    ws_tu = wb.create_sheet("TokenUsage")
    write_header(ws_tu, SESSION_TRACE_TOKEN_COLUMNS)
    for row_idx, tu in enumerate(token_usage, 2):
        for col_idx, col in enumerate(SESSION_TRACE_TOKEN_COLUMNS, 1):
            ws_tu.cell(row=row_idx, column=col_idx, value=tu.get(col))
    ws_tu.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SESSION_TRACE_TOKEN_COLUMNS))}{len(token_usage)+1}"

    wb.save(str(output))
