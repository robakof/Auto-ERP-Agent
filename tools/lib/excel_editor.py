"""
ExcelEditor — modyfikacja istniejących plików xlsx: zapis wartości do wskazanych komórek.

Używana przez: excel_write_cells.py
"""

from pathlib import Path

from openpyxl import load_workbook


class ExcelEditor:
    def __init__(self, path: Path):
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Plik nie istnieje: {path}")
        self._wb = load_workbook(path)

    def close(self):
        self._wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def save(self, path: Path) -> None:
        self._wb.save(path)

    def write_cells(
        self,
        sheet: str,
        key_column: str,
        target_column: str,
        data: dict,
    ) -> dict:
        """Zapisuje wartości z data do target_column, dopasowując wiersze po key_column.

        Args:
            sheet: nazwa arkusza
            key_column: kolumna identyfikująca wiersz (np. CDN_Pole)
            target_column: kolumna do zapisu (np. Komentarz)
            data: {wartość_klucza: wartość_do_zapisu}

        Returns:
            {"ok": True, "data": {"updated": N, "skipped": M}, "error": None}
        """
        if sheet not in self._wb.sheetnames:
            return {
                "ok": False,
                "data": None,
                "error": {
                    "type": "SHEET_NOT_FOUND",
                    "message": f"Arkusz '{sheet}' nie istnieje. Dostępne: {list(self._wb.sheetnames)}",
                },
            }

        ws = self._wb[sheet]
        rows = list(ws.iter_rows())
        if not rows:
            return {"ok": True, "data": {"updated": 0, "skipped": len(data)}, "error": None}

        headers = [cell.value for cell in rows[0]]

        if key_column not in headers:
            return {
                "ok": False,
                "data": None,
                "error": {
                    "type": "COLUMN_NOT_FOUND",
                    "message": f"Kolumna klucza '{key_column}' nie istnieje. Dostępne: {headers}",
                },
            }

        if target_column not in headers:
            return {
                "ok": False,
                "data": None,
                "error": {
                    "type": "COLUMN_NOT_FOUND",
                    "message": f"Kolumna docelowa '{target_column}' nie istnieje. Dostępne: {headers}",
                },
            }

        key_idx = headers.index(key_column)
        target_idx = headers.index(target_column)

        remaining = dict(data)
        updated = 0

        for row in rows[1:]:
            key_val = row[key_idx].value
            if key_val in remaining:
                row[target_idx].value = remaining.pop(key_val)
                updated += 1

        skipped = len(remaining)
        return {"ok": True, "data": {"updated": updated, "skipped": skipped}, "error": None}
