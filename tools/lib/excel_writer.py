"""
ExcelWriter — współdzielona logika zapisu plików xlsx.

Używana przez: excel_export.py, excel_export_bi.py
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelWriter:
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
    MAX_COL_WIDTH = 50

    TABLE_STYLE = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    def __init__(self):
        self._wb = Workbook()
        self._first = True
        self._table_counter = 0

    def add_sheet(self, name: str, columns: list[str], rows: list[list]) -> None:
        """Dodaje arkusz z sformatowanym nagłówkiem i danymi."""
        if self._first:
            ws = self._wb.active
            ws.title = name
            self._first = False
        else:
            ws = self._wb.create_sheet(name)

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        ws.freeze_panes = "A2"

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(col_name)
            for row in rows:
                val = str(row[col_idx - 1]) if col_idx - 1 < len(row) and row[col_idx - 1] is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, self.MAX_COL_WIDTH)

        if columns:
            last_col = get_column_letter(len(columns))
            last_row = len(rows) + 1
            self._table_counter += 1
            safe_name = re.sub(r"[^A-Za-z0-9_]", "_", name)
            table_name = f"T_{safe_name}_{self._table_counter}"
            table = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
            table.tableStyleInfo = self.TABLE_STYLE
            ws.add_table(table)

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(path)
