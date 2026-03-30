"""pp_export.py — Eksport planowania produkcji do Excel (do 7 arkuszy).

Arkusz 1: Zamówienia CZNI
Arkusz 2: Zapotrzebowanie surowców
Arkusz 3: Stany OTOR_SUR
Arkusz 4: Gap Analysis (czerwone wiersze gdzie Brak > 0)
Arkusz 5: Zamówienia z priorytetem (gdy schedule podany)
Arkusz 6: Harmonogram (gdy schedule podany)
Arkusz 7: Gantt (gdy schedule podany)
"""
from datetime import date
from pathlib import Path

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from tools.lib.excel_writer import ExcelWriter
from tools.lib.pp_gap import GapRow
from tools.lib.pp_schedule import ScheduleSlot

_RED_FILL    = PatternFill("solid", fgColor="FFFF9999")  # braki surowców
_YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")  # priorytet
_GREEN_FILL  = PatternFill("solid", fgColor="FF92D050")  # normalna produkcja
_GANTT_RED   = PatternFill("solid", fgColor="FFFF0000")  # po deadline

_COLS_ORDERS = [
    "Nr_Zamowienia", "Data_Realizacji", "Kontrahent_Kod", "Kontrahent_Nazwa",
    "Towar_Kod", "Towar_Nazwa", "Ilosc", "Jednostka", "Opis",
]
_COLS_DEMAND = ["Surowiec_Kod", "Surowiec_Nazwa", "Ilosc_Potrzebna"]
_COLS_SUPPLY = ["Towar_Kod", "Towar_Nazwa", "Jednostka", "Stan"]
_COLS_GAP    = ["Surowiec_Kod", "Surowiec_Nazwa", "Potrzeba", "Dostepne", "Brak"]


def export_plan(
    demand_rows: list[dict],
    supply: dict[str, float],
    gaps: list[GapRow],
    supply_rows: list[dict],
    output_path: Path,
    schedule: list[ScheduleSlot] | None = None,
    priorities: set[str] | None = None,
    start_date: date | None = None,
) -> None:
    """Zapisuje Excel z 4-7 arkuszami.

    Arkusze 5-7 (harmonogram) dodawane gdy schedule jest podany.
    """
    writer = ExcelWriter()
    _add_orders(writer, demand_rows)
    _add_demand(writer, gaps)
    _add_supply(writer, supply_rows)
    _add_gap(writer, gaps)
    if schedule is not None:
        _add_priority_orders(writer, demand_rows, priorities or set())
        _add_schedule(writer, schedule)
        _add_gantt(writer, schedule, start_date or date.today())
    try:
        writer.save(output_path)
    except PermissionError:
        raise RuntimeError(f"Nie można zapisać: {output_path}. Zamknij plik w Excelu.")


def _add_orders(writer: ExcelWriter, rows: list[dict]) -> None:
    data = [[r.get(c) for c in _COLS_ORDERS] for r in rows]
    data.sort(key=lambda r: (r[1] is None, r[1]))
    writer.add_sheet("Zamówienia CZNI", _COLS_ORDERS, data)


def _add_demand(writer: ExcelWriter, gaps: list[GapRow]) -> None:
    data = [[g.surowiec_kod, g.surowiec_nazwa, g.potrzeba] for g in gaps]
    data.sort(key=lambda r: r[0])
    writer.add_sheet("Zapotrzebowanie surowców", _COLS_DEMAND, data)


def _add_supply(writer: ExcelWriter, supply_rows: list[dict]) -> None:
    data = [[r.get(c) for c in _COLS_SUPPLY] for r in supply_rows]
    data.sort(key=lambda r: (r[0] or ""))
    writer.add_sheet("Stany OTOR_SUR", _COLS_SUPPLY, data)


def _add_gap(writer: ExcelWriter, gaps: list[GapRow]) -> None:
    data = [
        [g.surowiec_kod, g.surowiec_nazwa, g.potrzeba, g.dostepne, g.brak]
        for g in gaps
    ]
    writer.add_sheet("Gap Analysis", _COLS_GAP, data)
    _highlight_gap_rows(writer, gaps)


def _highlight_gap_rows(writer: ExcelWriter, gaps: list[GapRow]) -> None:
    """Czerwone tło dla wierszy gdzie Brak > 0 w arkuszu Gap Analysis."""
    wb = writer._wb
    ws = wb["Gap Analysis"]
    n_cols = len(_COLS_GAP)
    for row_idx, gap in enumerate(gaps, start=2):
        if gap.brak > 0:
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _RED_FILL


# ── Arkusz 5: Zamówienia z priorytetem ───────────────────────────────────────

_COLS_PRIORITY = [
    "Nr_Zamowienia", "Data_Realizacji", "Kontrahent_Kod",
    "Towar_Kod", "Towar_Nazwa", "Ilosc", "Priorytet",
]


def _add_priority_orders(
    writer: ExcelWriter,
    demand_rows: list[dict],
    priorities: set[str],
) -> None:
    """Arkusz 5: zamówienia z kolumną Priorytet (pusta = Run 1, "1" = Run 2)."""
    data = []
    for r in demand_rows:
        nr = r.get("Nr_Zamowienia")
        prio = "1" if nr in priorities else ""
        data.append([
            nr,
            r.get("Data_Realizacji"),
            r.get("Kontrahent_Kod"),
            r.get("Towar_Kod"),
            r.get("Towar_Nazwa"),
            r.get("Ilosc"),
            prio,
        ])
    data.sort(key=lambda r: (r[1] is None, r[1]))
    writer.add_sheet("Zamówienia z priorytetem", _COLS_PRIORITY, data)


# ── Arkusz 6: Harmonogram ────────────────────────────────────────────────────

_COLS_SCHEDULE = [
    "Data", "CZNI_Kod", "CZNI_Nazwa", "Ilosc", "Godziny",
    "Zamowienia", "Deadline", "Priorytet",
]


def _add_schedule(writer: ExcelWriter, schedule: list[ScheduleSlot]) -> None:
    """Arkusz 6: harmonogram dzienny, żółte wiersze z priorytetem."""
    data = sorted(schedule, key=lambda s: s.date)
    rows = [
        [
            s.date,
            s.czni_kod,
            s.czni_nazwa,
            round(s.qty, 2),
            round(s.hours_needed, 2),
            ", ".join(str(n) for n in s.order_numbers),
            s.deadline,
            "1" if s.priority else "",
        ]
        for s in data
    ]
    writer.add_sheet("Harmonogram", _COLS_SCHEDULE, rows)
    _highlight_schedule_rows(writer, data)


def _highlight_schedule_rows(writer: ExcelWriter, slots: list[ScheduleSlot]) -> None:
    """Żółte tło dla wierszy z priorytetem w arkuszu Harmonogram."""
    ws = writer._wb["Harmonogram"]
    n_cols = len(_COLS_SCHEDULE)
    for row_idx, slot in enumerate(slots, start=2):
        if slot.priority:
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _YELLOW_FILL


# ── Arkusz 7: Gantt ──────────────────────────────────────────────────────────

def _add_gantt(
    writer: ExcelWriter,
    schedule: list[ScheduleSlot],
    start_date: date,
) -> None:
    """Arkusz 7: Gantt (wiersze=CZNI, kolumny=daty, wartości=godziny).

    Kolory: zielony=normalne, żółty=priorytet, czerwony=slot po deadline.
    """
    if not schedule:
        return

    dates = sorted({s.date for s in schedule})
    czni_kody = sorted({s.czni_kod for s in schedule})
    date_to_col = {d: i + 2 for i, d in enumerate(dates)}

    ws = writer._wb.create_sheet("Gantt")

    # Nagłówek: kolumna A = CZNI_Kod, reszta = daty
    ws.cell(row=1, column=1, value="CZNI_Kod")
    for d, col_idx in date_to_col.items():
        ws.cell(row=1, column=col_idx, value=d)

    czni_to_row = {k: i + 2 for i, k in enumerate(czni_kody)}
    for k in czni_kody:
        ws.cell(row=czni_to_row[k], column=1, value=k)

    # Wypełnienie i kolorowanie
    for slot in schedule:
        row_idx = czni_to_row[slot.czni_kod]
        col_idx = date_to_col[slot.date]
        cell = ws.cell(row=row_idx, column=col_idx)
        existing = cell.value or 0
        cell.value = round(existing + slot.hours_needed, 2)
        cell.fill = _gantt_fill(slot)

    # Szerokość kolumn
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10


def _gantt_fill(slot: ScheduleSlot) -> PatternFill:
    if slot.deadline and slot.date > slot.deadline:
        return _GANTT_RED
    if slot.priority:
        return _YELLOW_FILL
    return _GREEN_FILL
