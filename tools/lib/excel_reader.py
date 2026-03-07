"""
ExcelReader — współdzielona logika odczytu plików xlsx: statystyki i dane wierszami.

Używana przez: excel_read_stats.py, excel_read_rows.py
"""

from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, path: Path):
        if not Path(path).exists():
            raise FileNotFoundError(f"Plik nie istnieje: {path}")
        self._wb = load_workbook(path, read_only=True, data_only=True)

    def close(self):
        self._wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def _get_ws(self, sheet_name: str | None):
        if sheet_name is None:
            ws = self._wb.active
            return ws, ws.title
        if sheet_name not in self._wb.sheetnames:
            raise KeyError(
                f"Arkusz '{sheet_name}' nie istnieje. Dostępne: {list(self._wb.sheetnames)}"
            )
        return self._wb[sheet_name], sheet_name

    def _parse(self, ws) -> tuple[list[str], list[list]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(h) if h is not None else f"Col{i + 1}" for i, h in enumerate(rows[0])]
        data_rows = [list(r) for r in rows[1:]]
        return headers, data_rows

    def _filter_cols(
        self,
        headers: list[str],
        data_rows: list[list],
        columns: list[str] | None,
    ) -> tuple[list[str], list[list]]:
        if columns is None:
            return headers, data_rows
        indices = [i for i, h in enumerate(headers) if h in columns]
        filtered_headers = [headers[i] for i in indices]
        filtered_rows = [[row[i] if i < len(row) else None for i in indices] for row in data_rows]
        return filtered_headers, filtered_rows

    def read_stats(
        self,
        sheet_name: str | None = None,
        max_unique: int = 20,
        columns: list[str] | None = None,
    ) -> dict:
        """Zwraca statystyki kolumn (total, distinct, null_count, values/sample)."""
        try:
            ws, actual_sheet = self._get_ws(sheet_name)
        except KeyError as e:
            return {"ok": False, "data": None, "error": {"type": "SHEET_NOT_FOUND", "message": str(e)}}

        headers, data_rows = self._parse(ws)

        if not headers:
            return {"ok": True, "data": {"sheet": actual_sheet, "row_count": 0, "columns": []}, "error": None}

        filtered_headers, filtered_data = self._filter_cols(headers, data_rows, columns)
        row_count = len(filtered_data)

        col_stats = []
        for col_idx, col_name in enumerate(filtered_headers):
            values = [row[col_idx] if col_idx < len(row) else None for row in filtered_data]
            null_count = sum(1 for v in values if v is None)
            non_null = [v for v in values if v is not None]

            seen: dict = {}
            for v in non_null:
                key = str(v)
                if key not in seen:
                    seen[key] = True
            distinct_values = list(seen.keys())
            distinct_count = len(distinct_values)

            stat: dict = {
                "name": col_name,
                "total": row_count,
                "distinct": distinct_count,
                "null_count": null_count,
            }
            if distinct_count <= max_unique:
                stat["values"] = distinct_values
                stat["sample"] = None
            else:
                stat["values"] = None
                stat["sample"] = distinct_values[:10]
            col_stats.append(stat)

        return {
            "ok": True,
            "data": {"sheet": actual_sheet, "row_count": row_count, "columns": col_stats},
            "error": None,
        }

    def read_rows(
        self,
        sheet_name: str | None = None,
        columns: list[str] | None = None,
    ) -> dict:
        """Zwraca wszystkie wiersze arkusza jako listę list."""
        try:
            ws, actual_sheet = self._get_ws(sheet_name)
        except KeyError as e:
            return {"ok": False, "data": None, "error": {"type": "SHEET_NOT_FOUND", "message": str(e)}}

        headers, data_rows = self._parse(ws)

        if not headers:
            return {
                "ok": True,
                "data": {"sheet": actual_sheet, "row_count": 0, "columns": [], "rows": []},
                "error": None,
            }

        filtered_headers, filtered_data = self._filter_cols(headers, data_rows, columns)

        return {
            "ok": True,
            "data": {
                "sheet": actual_sheet,
                "row_count": len(filtered_data),
                "columns": filtered_headers,
                "rows": filtered_data,
            },
            "error": None,
        }
