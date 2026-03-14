import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from tools.lib.agent_bus import AgentBus
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

bus = AgentBus(db_path="mrowisko.db")
items = bus.get_backlog()

# --- MD ---
lines = ["# Backlog\n"]
lines.append(f"*{len(items)} pozycji*\n")
lines.append("| # | Tytuł | Obszar | Wartość | Praca | Status |")
lines.append("|---|---|---|---|---|---|")
for i in items:
    lines.append(f"| {i['id']} | {i['title'][:60]} | {i['area'] or ''} | {i['value'] or ''} | {i['effort'] or ''} | {i['status']} |")

Path("backlog_view.md").write_text("\n".join(lines), encoding="utf-8")
print("backlog_view.md zapisany")

# --- XLSX ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backlog"

headers = ["ID", "Tytuł", "Obszar", "Wartość", "Praca", "Status", "Utworzono"]
header_fill = PatternFill("solid", fgColor="2D6A9F")
header_font = Font(bold=True, color="FFFFFF")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

value_colors = {"wysoka": "C6EFCE", "srednia": "FFEB9C", "niska": "FFC7CE"}
status_colors = {"planned": "FFFFFF", "in_progress": "FFEB9C", "done": "C6EFCE", "cancelled": "F2F2F2"}

for row, i in enumerate(items, 2):
    ws.cell(row=row, column=1, value=i["id"])
    ws.cell(row=row, column=2, value=i["title"])
    ws.cell(row=row, column=3, value=i["area"] or "")
    val_cell = ws.cell(row=row, column=4, value=i["value"] or "")
    if i["value"] in value_colors:
        val_cell.fill = PatternFill("solid", fgColor=value_colors[i["value"]])
    ws.cell(row=row, column=5, value=i["effort"] or "")
    status_cell = ws.cell(row=row, column=6, value=i["status"])
    if i["status"] in status_colors:
        status_cell.fill = PatternFill("solid", fgColor=status_colors[i["status"]])
    ws.cell(row=row, column=7, value=i["created_at"][:10])

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12

ws.auto_filter.ref = f"A1:G{len(items)+1}"

wb.save("backlog_view.xlsx")
print("backlog_view.xlsx zapisany")
